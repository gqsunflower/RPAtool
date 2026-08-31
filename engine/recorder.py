"""
MacroRecorder: 「次はどの操作を追加しますか?」のように対話形式で質問しながら、
Excel・PDF・Webサイト(ブラウザ)・エクスプローラー(ファイル/フォルダ)・
実行ファイル(exe/py)の操作を組み合わせて1つの一連の流れとして記録し、
config/macros.json (+ config/intents.json) に保存する。

Excel/PDF/エクスプローラーの各手順は、実際にその場でファイルに対して実行して
確認してから登録する。Webサイトの手順も同様に、実際にブラウザへ実行して確認
してから登録する(いわゆる「動作確認しながらレコーディングする」方式)。
Webのボタン等はCSSセレクタではなく「表示されているであろう文字」を手がかりにする
browser_handler の *_by_text 系アクションのみを使用し、UI変更に強くしている。

操作間違いへの対応:
- 目印やパスを聞かれた時点で「キャンセル」と入力すると、その操作は記録せずに
  メニューへ戻る(_ActionCancelled)。
- メインメニューの「直前の操作を取り消す」で、既に記録してしまった直前の1手順を
  取り消せる(_undo_last_step)。

Excel・PDF・Webサイト・エクスプローラー・実行ファイルの操作を同じ1つのマクロの
中で自由に組み合わせられるため、「Excelから読み込む → PDFを結合する →
フォルダへ移動する → 登録済みサイトにアップロードする」のような一連の流れを
まとめて1回の登録で作れる。既に個別に登録済みのマクロ同士を後からつなげたい
場合は、main.pyの「パイプライン作成」機能を使う。

破壊的な操作への配慮:
- ファイルの移動・上書きは、実際に試して確認する過程でも本当に発生する
  (元に戻せない場合がある)ため、移動系の操作は登録前にもう一段階
  「本当に実行してよいか」を確認する。
- exe/pyの実行は、事前に config/exec_whitelist.json へ登録した
  ファイルのみ実行できる(ブラウザのサイトホワイトリストと同じ考え方)。
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter

from engine.backup import backup_file
from handlers.browser_handler import (
    BrowserHandler,
    ElementNotFoundError,
    SiteNotWhitelistedError,
)
from handlers.excel_handler import ExcelHandler
from handlers.explorer_handler import ExplorerHandler
from handlers.pdf_handler import PdfHandler
from handlers.desktop_handler import DesktopHandler
from handlers.list_handler import ListHandler
from handlers.process_handler import ProcessHandler
from handlers.text_handler import TextHandler

CANCEL_WORDS = {"キャンセル", "cancel", "戻る", "やめる"}


def _offset_template(var_name: str, offset: int) -> str:
    """{{var_name}} に対するオフセット付きテンプレート文字列を組み立てる。
    offset=0なら "{{var_name}}"、それ以外は "{{var_name+N}}"/"{{var_name-N}}"。
    """
    if offset == 0:
        return "{{" + var_name + "}}"
    sign = "+" if offset > 0 else "-"
    return "{{" + var_name + sign + str(abs(offset)) + "}}"


class _ActionCancelled(Exception):
    """記録中の操作を、目印/パス入力の時点で取りやめたことを表す。"""


class MacroRecorder:
    def __init__(self, config_dir: Path, browser: str = "chrome"):
        self.config_dir = Path(config_dir)
        self.browser = BrowserHandler(self.config_dir / "whitelist_urls.json", headless=False, browser=browser)
        self.excel = ExcelHandler()
        self.pdf = PdfHandler()
        self.explorer = ExplorerHandler()
        self.process = ProcessHandler(self.config_dir / "exec_whitelist.json")
        self.desktop = DesktopHandler()
        self.text = TextHandler()
        self.list = ListHandler()
        # 記録中に実際に動作確認できた変数・リストの値(変数名 -> 値)。
        # レコーダーが各手順を動作確認した「その場の実値」を記録しておき、
        # GUIの右側パネル等でユーザーに一覧表示するために使う。あくまで
        # 記録時点のスナップショットであり、実行時に必ず同じ値になるとは限らない
        # (次回実行時にExcelの中身が変わっていれば最終行等は変わる)。
        self.variables: dict[str, Any] = {}
        self.steps: list[dict] = []
        self.required_slots: list[str] = []
        self._site_opened = False
        self._auto_var_counter = 0

    def _next_auto_var(self, prefix: str) -> str:
        """最終行/最終列などを自動取得する際に使う、衝突しない変数名を発行する。"""
        self._auto_var_counter += 1
        return f"_auto_{prefix}_{self._auto_var_counter}"

    def record_variable(self, name: str, value: Any) -> None:
        """store_asで変数として保存する手順を、実際に動作確認できたときに呼ぶ。
        記録時点で確認できた実値を self.variables に記録しておき、GUIの
        変数一覧パネル等で確認できるようにする(あくまで記録時点のスナップ
        ショットであり、実行のたびに変わりうる値はその都度更新される)。
        """
        if name:
            self.variables[name] = value

    # ---------- 共通の入力ヘルパー ----------

    def _ask(self, prompt: str) -> str:
        return input(prompt).strip()

    def _ask_target(self, prompt: str) -> str:
        """Webの目印など、キャンセル可能な入力。CANCEL_WORDS が入力された場合は
        この操作の記録全体を取りやめる(_ActionCancelled を送出)。
        """
        val = input(prompt).strip()
        if val in CANCEL_WORDS:
            raise _ActionCancelled()
        return val

    def _ask_sluttable_value(self, label: str) -> tuple[str, str] | None:
        """ファイルパスやシート名など、Excel/PDFの値を1つ聞く。
        「キャンセル」が入力されたら None を返す。
        戻り値: (動作確認に使う実際の値, macros.jsonに書き込むparam値)

        入力した値の中に "{{変数名}}" が含まれている場合(前の手順で
        store_as した結果を埋め込みたい場合)は、それをそのままテンプレートとして
        登録し、動作確認用にだけ別途「実際の値に置き換えたもの」を聞く。
        例: 値として "A1:B{{last_row}}" と入力すると、そのまま登録され、
        動作確認では「A1:B10」のような具体的な値で試せる。
        "{{last_row+1}}" のように変数名の後ろに +数値/-数値 を付けると、
        その場で加減算した値を使える(「最終行の次の行」等)。
        値全体がスロットの場合は従来どおり「スロット名」欄で指定する
        (スロットは実行のたびに外部から与える値、{{変数名}}の直接入力は
        前の手順の結果を埋め込む値、という使い分け)。
        """
        value = input(
            f"  {label}を入力してください"
            f"(前の手順の結果を埋め込む場合は {{変数名}}、次の行なら {{変数名+1}} のように書けます。"
            f"間違えた場合は「キャンセル」): "
        ).strip()
        if value in CANCEL_WORDS:
            return None

        if "{{" in value and "}}" in value:
            test_value = self._ask(
                f"  動作確認用に、実際の値に置き換えたものを入力してください(例: {value}): "
            )
            return test_value, value

        slot_name = self._ask(
            "  この値は実行するたびに変わりますか? 変わる場合はスロット名を、固定値ならそのままEnter: "
        )
        if slot_name:
            if slot_name not in self.required_slots:
                self.required_slots.append(slot_name)
            return value, "{{" + slot_name + "}}"
        return value, value

    def _ask_verification(self) -> dict:
        """この手順が本当に成功したかを、次にどう確認するかを聞く(Web操作用)。
        「ちゃんとボタンが押せていたか」を毎回確かめるための仕組みで、
        ここで確認方法を決めておかないと後続の手順が誤った状態のまま
        先に進んでしまう(=とんでもない操作につながる)ことを防ぐ。
        """
        print("  この操作が『本当に成功したか』をどう確認しますか?")
        print("    1) この後、特定の文字が新しく表示されるはず")
        print("    2) この後、画面のURLが変わるはず")
        print("    3) 確認しない(ほぼ失敗しない操作の場合。後から個別に変更も可能)")
        choice = self._ask("  番号> ")
        if choice == "1":
            text = self._ask("  表示されるはずの文字を入力してください: ")
            if text:
                return {"type": "text_appears", "value": text, "timeout": 10}
        elif choice == "2":
            return {"type": "url_changes", "timeout": 10}
        return {"type": "none"}

    def _ask_obstruction_wait(self) -> float:
        """広告のポップアップ等、別の要素にクリックが妨害される可能性がある操作かを聞き、
        手動で閉じるのを待つ最大秒数を設定する(既定は待機せず即エラー)。
        """
        raw = self._ask(
            "  広告等の別要素にクリックが妨害される可能性がありますか?"
            " 手動で閉じるのを待つ最大秒数を入力してください(不要ならそのままEnter): "
        ).strip()
        if not raw:
            return 0
        try:
            return max(float(raw), 0)
        except ValueError:
            return 0

    def _ask_retry(self) -> dict:
        """一時的な読み込み遅延などで失敗しやすい操作かどうかを聞き、
        自動リトライ回数を設定する。ほぼ確実に成功する操作は0のままでよい。
        """
        raw = self._ask(
            "  読み込み待ちなどで失敗しやすい操作ですか?"
            " 自動で再試行する回数を入力してください(不要ならそのままEnter): "
        )
        if not raw:
            return {"count": 0, "interval_seconds": 2}
        try:
            count = max(int(raw), 0)
        except ValueError:
            return {"count": 0, "interval_seconds": 2}
        return {"count": count, "interval_seconds": 2}

    # ---------- Webサイト: サイト選択(初回操作時のみ) ----------

    def _choose_site(self) -> str:
        sites = self.browser._sites
        if not sites:
            print("  登録済みサイトがまだありません。新しいサイトを登録します。")
            return self._register_new_site()

        print("操作するサイトを選んでください:")
        keys = list(sites.keys())
        for i, key in enumerate(keys, start=1):
            print(f"  {i}. {key} ({sites[key]['url']})")
        print("  0. 新しいサイトを登録する")

        choice = self._ask("番号> ")
        if choice.strip() == "0":
            return self._register_new_site()
        try:
            idx = int(choice) - 1
            if idx < 0:
                raise ValueError
            return keys[idx]
        except (ValueError, IndexError):
            print("  入力が正しくありません。もう一度選んでください。\n")
            return self._choose_site()

    def _register_new_site(self) -> str:
        site_key = self._ask("サイトの識別名を入力してください(半角英数字、例: portal): ")
        url = self._ask("サイトのURL(操作を始める画面)を入力してください: ")
        self.browser.register_site(site_key, url)
        print(f"  → '{site_key}' をホワイトリストに登録しました\n")
        return site_key

    def _ensure_site_opened(self) -> bool:
        """Webサイト操作を初めて追加するときだけサイトを開く(遅延オープン)。
        既に開いている場合は何もしない。
        """
        if self._site_opened:
            return True
        site_key = self._choose_site()
        try:
            self.browser.open_registered_site(site_key)
        except SiteNotWhitelistedError as e:
            print(f"  ⚠ {e}\n")
            return False
        self.steps.append({
            "handler": "browser",
            "action": "open_registered_site",
            "params": {"site_key": site_key},
        })
        self._site_opened = True
        print(f"  → '{site_key}' を開きました。\n")
        return True

    # ---------- メインループ ----------

    def record(self) -> str | None:
        print("=== 操作の登録(レコーダー) ===")
        print("Excel・PDF・Webサイト・エクスプローラー・実行ファイル(exe/py)・")
        print("デスクトップ(画面の画像でクリック等)・テキスト加工・リスト・")
        print("制御構文(For/If/Goto)の操作を組み合わせて、一連の流れとして")
        print("記録できます。")
        print("(目印やパスの入力時に「キャンセル」と入力すると、その操作は記録せず戻れます)\n")

        base_step_count = 0  # まだ何も記録していない状態(これより前には戻れない)

        while True:
            print("次はどの操作を追加しますか?")
            print("  1) Excel操作を追加する")
            print("  2) PDF操作を追加する")
            print("  3) Webサイト操作を追加する(ブラウザ)")
            print("  4) エクスプローラー操作を追加する(ファイル/フォルダ)")
            print("  5) 実行ファイル(exe/py)を追加する")
            print("  6) デスクトップ操作を追加する(画面の画像でクリック等)")
            print("  7) テキスト加工を追加する(切り出し・置換・日付付加など)")
            print("  8) リスト操作を追加する(配列の作成・追加・取得)")
            print("  9) 制御構文を追加する(For繰り返し・If判定・Goto)")
            print("  10) ここまでの操作を保存して終了する")
            print("  11) 保存せずに中止する")
            print("  12) 直前の操作を取り消す(操作を間違えた場合)")
            print("  13) 今の変数一覧を見る(記録時点で確認できた値)")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_excel_menu()
            elif choice == "2":
                self._record_pdf_menu()
            elif choice == "3":
                self._record_web_menu()
            elif choice == "4":
                self._record_explorer_menu()
            elif choice == "5":
                self._record_process_menu()
            elif choice == "6":
                self._record_desktop_menu()
            elif choice == "7":
                self._record_text_menu()
            elif choice == "8":
                self._record_list_menu()
            elif choice == "9":
                self._record_control_menu()
            elif choice == "10":
                return self._finish()
            elif choice == "11":
                if self._site_opened:
                    self.browser.close()
                print("登録を中止しました(何も保存していません)。")
                return None
            elif choice == "12":
                self._undo_last_step(base_step_count)
            elif choice == "13":
                self._print_variables()
            else:
                print("1〜13のいずれかを入力してください。\n")

    def _print_variables(self) -> None:
        if not self.variables:
            print("  (まだ変数はありません)\n")
            return
        print("  今の変数一覧(記録時点で確認できた値。実行時は変わる場合があります):")
        for name, value in self.variables.items():
            preview = repr(value)
            if len(preview) > 100:
                preview = preview[:100] + "..."
            print(f"    {name} ({type(value).__name__}) = {preview}")
        print()

    # ---------- 直前の操作の取り消し ----------

    def _undo_last_step(self, base_step_count: int) -> None:
        if len(self.steps) <= base_step_count:
            print("  これ以上は取り消せません(まだ何も記録していません)。\n")
            return

        removed = self.steps.pop()
        label = f"{removed['handler']}.{removed['action']} {removed.get('params', {})}"
        print(f"  → 直前の操作を取り消しました: {label}")

        if removed["handler"] == "browser" and removed["action"] == "open_registered_site":
            # サイトを開く手順自体を取り消した場合、次にWeb操作を追加するときは
            # 改めてサイト選択からやり直す
            self._site_opened = False

        removed_store_as = removed.get("store_as")
        if removed_store_as and removed_store_as in self.variables:
            del self.variables[removed_store_as]
            print(f"    (変数 '{removed_store_as}' の記録も取り消しました)")

        # このステップでしか使われていなかったスロットは required_slots からも外す
        removed_slot_refs = self._collect_slot_refs(removed.get("params", {}))
        for ref in removed_slot_refs:
            slot_name = ref[2:-2].strip()
            still_used = any(
                ref in self._collect_slot_refs(step.get("params", {}))
                for step in self.steps
            )
            if not still_used and slot_name in self.required_slots:
                self.required_slots.remove(slot_name)
                print(f"    (スロット '{slot_name}' も他で使われていないため削除しました)")

        print(f"  現在、記録済みの操作は {len(self.steps) - base_step_count} 件です。\n")

    @staticmethod
    def _collect_slot_refs(params: dict) -> list[str]:
        """params内(dictの値も含む)から "{{slot}}" 形式の参照をすべて集める。"""
        refs = []
        for v in params.values():
            if isinstance(v, str) and v.startswith("{{") and v.endswith("}}"):
                refs.append(v)
            elif isinstance(v, dict):
                refs.extend(
                    x for x in v.values()
                    if isinstance(x, str) and x.startswith("{{") and x.endswith("}}")
                )
            elif isinstance(v, list):
                refs.extend(
                    x for x in v
                    if isinstance(x, str) and x.startswith("{{") and x.endswith("}}")
                )
        return refs

    # ---------- Excel操作 ----------

    def _record_excel_menu(self) -> None:
        while True:
            print("Excelでは何をしますか?")
            print("  1) Excelファイルを開く")
            print("  2) シートを読み込む")
            print("  3) 読み込んだデータをCSVに書き出す")
            print("  4) セルに値を書き込む")
            print("  5) 別名で保存する")
            print("  6) PDFとして保存する(印刷範囲・ページ設定は事前設定済み前提)")
            print("  7) 組み込み済みのVBAマクロを実行する(.xlsm)")
            print("  8) 開いている別のExcelに切り替える(複数ファイルを行き来する)")
            print("  9) セル1つの値を読み込む")
            print("  10) 最終行を取得する")
            print("  11) セルをコピーして貼り付ける")
            print("  12) シートをコピーする")
            print("  13) シート名の一覧を取得する")
            print("  14) 印刷範囲を指定する")
            print("  15) 新規シートを追加する")
            print("  16) シート名を変更する")
            print("  17) セル範囲(1列/1行)をリストとして取得する")
            print("  18) 値でセルを検索する(アドレスを取得)")
            print("  19) 表の末尾に1行追記する")
            print("  20) 行を挿入する")
            print("  21) 行を削除する")
            print("  22) セル範囲の値を空にする")
            print("  23) シートを削除する")
            print("  24) 最終列を取得する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_excel_load()
            elif choice == "2":
                self._record_excel_read_sheet()
            elif choice == "3":
                self._record_excel_write_csv()
            elif choice == "4":
                self._record_excel_write_cells()
            elif choice == "5":
                self._record_excel_save_as()
            elif choice == "6":
                self._record_excel_save_as_pdf()
            elif choice == "7":
                self._record_excel_run_macro()
            elif choice == "8":
                self._record_excel_switch()
            elif choice == "9":
                self._record_excel_get_cell_value()
            elif choice == "10":
                self._record_excel_get_last_row()
            elif choice == "11":
                self._record_excel_copy_cell_range()
            elif choice == "12":
                self._record_excel_copy_sheet()
            elif choice == "13":
                self._record_excel_get_sheet_names()
            elif choice == "14":
                self._record_excel_set_print_area()
            elif choice == "15":
                self._record_excel_create_sheet()
            elif choice == "16":
                self._record_excel_rename_sheet()
            elif choice == "17":
                self._record_excel_get_range_as_list()
            elif choice == "18":
                self._record_excel_find_cell()
            elif choice == "19":
                self._record_excel_append_row()
            elif choice == "20":
                self._record_excel_insert_rows()
            elif choice == "21":
                self._record_excel_delete_rows()
            elif choice == "22":
                self._record_excel_clear_range()
            elif choice == "23":
                self._record_excel_delete_sheet()
            elif choice == "24":
                self._record_excel_get_last_column()
            elif choice == "0":
                return
            else:
                print("0〜24のいずれかを入力してください。\n")

    def _record_excel_load(self) -> None:
        result = self._ask_sluttable_value("開くExcelファイルのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        alias = self._ask(
            "  このExcelを後で識別するための名前(複数ファイルを開く場合に使用。"
            "空欄ならファイル名から自動で付けます): "
        ).strip()
        try:
            result_msg = self.excel.load_workbook(test_value, alias=alias or None)
            print(f"  → 実際に '{test_value}' を開いて確認できました。({result_msg})")
            params = {"path": param_value}
            if alias:
                params["alias"] = alias
            self.steps.append({"handler": "excel", "action": "load_workbook", "params": params})
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                params = {"path": param_value}
                if alias:
                    params["alias"] = alias
                self.steps.append({"handler": "excel", "action": "load_workbook", "params": params})
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。\n")

    def _record_excel_read_sheet(self) -> None:
        result = self._ask_sluttable_value("読み込むシート名")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        try:
            records = self.excel.read_sheet_to_records(test_value)
            print(f"  → 実際に読み込めました({len(records)}件)。")
            self.steps.append({
                "handler": "excel", "action": "read_sheet_to_records",
                "params": {"sheet_name": param_value},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_write_csv(self) -> None:
        result = self._ask_sluttable_value("出力するCSVファイルのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        try:
            self.excel.write_records_to_csv(test_value)
            print(f"  → 実際にCSVを書き出せました: {test_value}")
            self.steps.append({
                "handler": "excel", "action": "write_records_to_csv",
                "params": {"output_path": param_value},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_write_cells(self) -> None:
        sheet_result = self._ask_sluttable_value("書き込み対象のシート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result

        print("  セル参照の指定方法を選んでください:")
        print("    1) セル参照を直接指定する(例: B2)")
        print("    2) 最終行の続きに書き込む(列だけ指定。表の下に1行追加する場合等)")
        print("    3) 最終列の続きに書き込む(行だけ指定。表の右に1列追加する場合等)")
        basis = self._ask("  番号(空Enterで1)> ").strip() or "1"
        print()

        prereq_step: dict | None = None
        row_part: str | None = None
        row_test_base: int | None = None
        col_part: str | None = None
        col_test_base: str | None = None

        if basis == "2":
            col_scope = self._ask(
                "  最終行を判定する基準列(空欄ならシート全体の最終行。"
                "特定の列のデータの続きに書きたい場合は列を指定): "
            ).strip()
            offset_raw = self._ask("  最終行から何行後に書き込みますか?(空Enterで1=次の行): ").strip()
            try:
                offset = int(offset_raw) if offset_raw else 1
            except ValueError:
                offset = 1
            try:
                current_last_row = self.excel.get_last_row(sheet_test, column=col_scope or None)
            except Exception as e:  # noqa: BLE001
                print(f"  ⚠ 最終行の取得に失敗しました: {e}\n")
                return
            var_name = self._next_auto_var("last_row")
            prereq_step = {
                "handler": "excel", "action": "get_last_row",
                "params": {"sheet_name": sheet_param, "column": col_scope or None},
                "store_as": var_name,
            }
            self.record_variable(var_name, current_last_row)
            row_test_base = current_last_row + offset
            row_part = _offset_template(var_name, offset)
            print(f"  → 今の時点の最終行は{current_last_row}行目なので、"
                  f"動作確認では{row_test_base}行目に書き込みます。\n")
        elif basis == "3":
            row_scope = self._ask("  最終列を判定する基準行(行番号): ").strip()
            if not row_scope.isdigit():
                print("  → 行番号を数字で入力してください。この手順は登録しませんでした。\n")
                return
            offset_raw = self._ask("  最終列から何列後に書き込みますか?(空Enterで1=次の列): ").strip()
            try:
                offset = int(offset_raw) if offset_raw else 1
            except ValueError:
                offset = 1
            try:
                current_last_col = self.excel.get_last_column(sheet_test, row=row_scope)
            except Exception as e:  # noqa: BLE001
                print(f"  ⚠ 最終列の取得に失敗しました: {e}\n")
                return
            var_name = self._next_auto_var("last_col")
            prereq_step = {
                "handler": "excel", "action": "get_last_column",
                "params": {"sheet_name": sheet_param, "row": int(row_scope)},
                "store_as": var_name,
            }
            self.record_variable(var_name, current_last_col)
            col_idx_base = column_index_from_string(current_last_col) if current_last_col else 0
            if col_idx_base + offset < 1:
                print("  → 指定したオフセットでは列がA列より前になります。この手順は登録しませんでした。\n")
                return
            col_test_base = get_column_letter(col_idx_base + offset)
            col_part = _offset_template(var_name, offset)
            print(f"  → 今の時点の最終列は{current_last_col or '(無し)'}なので、"
                  f"動作確認では{col_test_base}列に書き込みます。\n")

        print("  書き込むセルと値を入力します(空Enterで入力終了)")
        cell_values_test: dict[str, str] = {}
        cell_values_param: dict[str, str] = {}
        while True:
            if basis == "2":
                axis_raw = self._ask(f"  列(例: B。行は{row_test_base}に固定。空Enterで入力終了): ").strip()
                if not axis_raw:
                    break
                try:
                    col_idx = int(axis_raw) if axis_raw.isdigit() else column_index_from_string(axis_raw.upper())
                    col_letter = get_column_letter(col_idx)
                except ValueError:
                    print("  列は 'B' のような列文字、または列番号で入力してください。\n")
                    continue
                cell_ref_test = f"{col_letter}{row_test_base}"
                cell_ref_param = f"{col_letter}{row_part}"
            elif basis == "3":
                axis_raw = self._ask(f"  行番号(列は{col_test_base}に固定。空Enterで入力終了): ").strip()
                if not axis_raw:
                    break
                if not axis_raw.isdigit():
                    print("  → 行番号は数字で入力してください。\n")
                    continue
                cell_ref_test = f"{col_test_base}{axis_raw}"
                cell_ref_param = f"{col_part}{axis_raw}"
            else:
                cell_ref_test = self._ask("  セル参照(例: B2。空Enterで入力終了): ")
                if not cell_ref_test:
                    break
                cell_ref_param = cell_ref_test

            result = self._ask_sluttable_value(f"セル {cell_ref_test} に書き込む値")
            if result is None:
                continue
            test_v, param_v = result
            cell_values_test[cell_ref_test] = test_v
            cell_values_param[cell_ref_param] = param_v

        if not cell_values_param:
            print("  → 値が1つも入力されなかったため、この手順は登録しませんでした。\n")
            return

        try:
            self.excel.write_cells(sheet_test, cell_values_test)
            print("  → 実際にセルへの書き込みを確認できました。")
            if prereq_step:
                self.steps.append(prereq_step)
            self.steps.append({
                "handler": "excel", "action": "write_cells",
                "params": {"sheet_name": sheet_param, "cell_values": cell_values_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_save_as(self) -> None:
        result = self._ask_sluttable_value("保存先のパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        try:
            self.excel.save_workbook_as(test_value)
            print(f"  → 実際に保存できました: {test_value}")
            self.steps.append({
                "handler": "excel", "action": "save_workbook_as",
                "params": {"output_path": param_value},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_save_as_pdf(self) -> None:
        print("  ※ 印刷範囲・ページ設定(用紙サイズ・拡大縮小・余白等)はこの機能では変更しません。")
        print("     事前にExcel側で設定を済ませておいてください。")
        print("  ※ この操作にはWindows上のExcel本体 + pywin32が必要です。")

        path_result = self._ask_sluttable_value("PDFへ書き出す元のExcelファイルのパス")
        if path_result is None:
            print("  → キャンセルしました。\n")
            return
        path_test, path_param = path_result

        out_result = self._ask_sluttable_value("出力するPDFファイルのパス")
        if out_result is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = out_result

        sheet_raw = self._ask("  特定のシートだけ書き出しますか? シート名を入力(ブック全体ならそのままEnter): ")
        sheet_param = sheet_raw or None

        try:
            self.excel.save_as_pdf(path_test, out_test, sheet_name=sheet_raw or None)
            print(f"  → 実際にPDFとして書き出せました: {out_test}")
            self.steps.append({
                "handler": "excel", "action": "save_as_pdf",
                "params": {"path": path_param, "output_path": out_param, "sheet_name": sheet_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                self.steps.append({
                    "handler": "excel", "action": "save_as_pdf",
                    "params": {"path": path_param, "output_path": out_param, "sheet_name": sheet_param},
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。\n")

    def _ask_store_as(self) -> str | None:
        """この手順の実行結果を、後の手順で使うための変数名を聞く。
        空欄なら使わない({{変数名}}の形で後続の手順のテキスト欄に入力すれば
        参照できるようになる)。
        """
        name = self._ask(
            "  この結果に名前を付けて後の手順で使いますか?"
            "(空欄なら使わない。使う場合の書き方: {{変数名}}): "
        ).strip()
        return name or None

    def _record_excel_run_macro(self) -> None:
        print("  ※ この操作にはWindows上のExcel本体 + pywin32が必要です。")
        path_result = self._ask_sluttable_value("マクロが組み込まれている.xlsmファイルのパス")
        if path_result is None:
            print("  → キャンセルしました。\n")
            return
        path_test, path_param = path_result

        macro_result = self._ask_sluttable_value("実行するマクロ名(Sub名)")
        if macro_result is None:
            print("  → キャンセルしました。\n")
            return
        macro_test, macro_param = macro_result

        try:
            self.excel.run_excel_macro(path_test, macro_test)
            print(f"  → 実際にマクロ '{macro_test}' を実行できました。")
            self.steps.append({
                "handler": "excel", "action": "run_excel_macro",
                "params": {"path": path_param, "macro_name": macro_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                self.steps.append({
                    "handler": "excel", "action": "run_excel_macro",
                    "params": {"path": path_param, "macro_name": macro_param},
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。\n")

    def _record_excel_switch(self) -> None:
        aliases = self.excel.list_open_workbooks()
        if len(aliases) < 2:
            print("  ⚠ 切り替え可能なExcelがまだ2つ以上開かれていません"
                  "(先に「1) Excelファイルを開く」で別名で2つ以上開いてください)\n")
            return
        print("開いているExcel:")
        for i, a in enumerate(aliases, start=1):
            marker = " (現在アクティブ)" if a == self.excel._active_alias else ""
            print(f"  {i}. {a}{marker}")
        choice = self._ask("番号> ")
        try:
            alias = aliases[int(choice) - 1]
        except (ValueError, IndexError):
            print("  入力が正しくありません\n")
            return
        self.excel.switch_workbook(alias)
        print(f"  → '{alias}' に切り替えました。")
        self.steps.append({"handler": "excel", "action": "switch_workbook", "params": {"alias": alias}})
        print("  → 登録しました。\n")

    def _record_excel_get_cell_value(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        cell_result = self._ask_sluttable_value("読み込むセル参照(例: B2)")
        if cell_result is None:
            print("  → キャンセルしました。\n")
            return
        cell_test, cell_param = cell_result

        try:
            value = self.excel.get_cell_value(sheet_test, cell_test)
            print(f"  → 実際に読み込めました: {value!r}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, value)
            step = {
                "handler": "excel", "action": "get_cell_value",
                "params": {"sheet_name": sheet_param, "cell_ref": cell_param},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_get_last_row(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        column = self._ask("  対象の列(例: A。空欄ならシート全体の最終行): ").strip()

        try:
            last_row = self.excel.get_last_row(sheet_test, column=column or None)
            print(f"  → 実際に取得できました: {last_row}行目")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, last_row)
            step = {
                "handler": "excel", "action": "get_last_row",
                "params": {"sheet_name": sheet_param, "column": column or None},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_get_last_column(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        row = self._ask("  対象の行番号(例: 1。空欄ならシート全体の最終列): ").strip()

        try:
            last_col = self.excel.get_last_column(sheet_test, row=row or None)
            print(f"  → 実際に取得できました: {last_col or '(該当する列なし)'}列目")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, last_col)
            step = {
                "handler": "excel", "action": "get_last_column",
                "params": {"sheet_name": sheet_param, "row": row or None},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _ask_excel_dest_cell(self, sheet_test: str, sheet_param: str) -> tuple[str, str, dict | None] | None:
        """貼り付け先の左上セルを聞く。直接指定のほか、「最終行の続き」
        「最終列の続き」を基準にした指定にも対応する(戻り値: (実値, テンプレート値,
        必要なら事前に挿入するget_last_row/get_last_column手順) または None=キャンセル)。
        """
        print("  貼り付け先の指定方法を選んでください:")
        print("    1) セルを直接指定する(例: D1)")
        print("    2) 最終行の続きに貼り付ける(列を指定)")
        print("    3) 最終列の続きに貼り付ける(行を指定)")
        basis = self._ask("  番号(空Enterで1)> ").strip() or "1"

        if basis == "2":
            col_raw = self._ask("  貼り付け先の列(例: D): ").strip()
            try:
                col_idx = int(col_raw) if col_raw.isdigit() else column_index_from_string(col_raw.upper())
                col_letter = get_column_letter(col_idx)
            except ValueError:
                print("  → 列の指定が不正です。\n")
                return None
            col_scope = self._ask(
                "  最終行を判定する基準列(空欄ならシート全体の最終行): "
            ).strip()
            offset_raw = self._ask("  最終行から何行後に貼り付けますか?(空Enterで1=次の行): ").strip()
            try:
                offset = int(offset_raw) if offset_raw else 1
            except ValueError:
                offset = 1
            try:
                current_last_row = self.excel.get_last_row(sheet_test, column=col_scope or None)
            except Exception as e:  # noqa: BLE001
                print(f"  ⚠ 最終行の取得に失敗しました: {e}\n")
                return None
            var_name = self._next_auto_var("last_row")
            prereq_step = {
                "handler": "excel", "action": "get_last_row",
                "params": {"sheet_name": sheet_param, "column": col_scope or None},
                "store_as": var_name,
            }
            self.record_variable(var_name, current_last_row)
            dest_row_test = current_last_row + offset
            print(f"  → 今の時点の最終行は{current_last_row}行目なので、"
                  f"動作確認では{col_letter}{dest_row_test}に貼り付けます。")
            return f"{col_letter}{dest_row_test}", f"{col_letter}{_offset_template(var_name, offset)}", prereq_step

        if basis == "3":
            row_raw = self._ask("  貼り付け先の行番号(例: 5): ").strip()
            if not row_raw.isdigit():
                print("  → 行番号を数字で入力してください。\n")
                return None
            offset_raw = self._ask("  最終列から何列後に貼り付けますか?(空Enterで1=次の列): ").strip()
            try:
                offset = int(offset_raw) if offset_raw else 1
            except ValueError:
                offset = 1
            try:
                current_last_col = self.excel.get_last_column(sheet_test, row=row_raw)
            except Exception as e:  # noqa: BLE001
                print(f"  ⚠ 最終列の取得に失敗しました: {e}\n")
                return None
            col_idx_base = column_index_from_string(current_last_col) if current_last_col else 0
            if col_idx_base + offset < 1:
                print("  → 指定したオフセットでは列がA列より前になります。\n")
                return None
            var_name = self._next_auto_var("last_col")
            prereq_step = {
                "handler": "excel", "action": "get_last_column",
                "params": {"sheet_name": sheet_param, "row": int(row_raw)},
                "store_as": var_name,
            }
            self.record_variable(var_name, current_last_col)
            dest_col_test = get_column_letter(col_idx_base + offset)
            print(f"  → 今の時点の最終列は{current_last_col or '(無し)'}なので、"
                  f"動作確認では{dest_col_test}{row_raw}に貼り付けます。")
            return f"{dest_col_test}{row_raw}", f"{_offset_template(var_name, offset)}{row_raw}", prereq_step

        result = self._ask_sluttable_value("貼り付け先の左上セル(例: D1)")
        if result is None:
            return None
        dest_test, dest_param = result
        return dest_test, dest_param, None

    def _record_excel_copy_cell_range(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        src_result = self._ask_sluttable_value("コピー元の範囲(例: A1:B3)")
        if src_result is None:
            print("  → キャンセルしました。\n")
            return
        src_test, src_param = src_result

        dest_result = self._ask_excel_dest_cell(sheet_test, sheet_param)
        if dest_result is None:
            print("  → キャンセルしました。\n")
            return
        dest_test, dest_param, prereq_step = dest_result

        print("  貼り付け方法を選んでください:")
        print("    1) 値のみ貼り付け(数式は計算済みの値になる)")
        print("    2) 数式の貼り付け(セル参照はそのままコピーされ、自動調整はされません)")
        print("    3) すべて貼り付け(値/数式 + 表示形式・フォント・塗りつぶし・罫線)")
        paste_choice = self._ask("  番号(空Enterで1: 値のみ)> ")
        paste_type = {"1": "values", "2": "formulas", "3": "all"}.get(paste_choice, "values")

        try:
            result = self.excel.copy_cell_range(sheet_test, src_test, dest_test, paste_type=paste_type)
            print(f"  → 実際にコピーできました: {result}")
            if prereq_step:
                self.steps.append(prereq_step)
            self.steps.append({
                "handler": "excel", "action": "copy_cell_range",
                "params": {
                    "sheet_name": sheet_param, "source_range": src_param,
                    "dest_cell": dest_param, "paste_type": paste_type,
                },
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_create_sheet(self) -> None:
        name_result = self._ask_sluttable_value("追加する新規シート名")
        if name_result is None:
            print("  → キャンセルしました。\n")
            return
        name_test, name_param = name_result
        idx_raw = self._ask("  挿入位置(0で先頭。空Enterで末尾に追加): ").strip()
        index = None
        if idx_raw:
            try:
                index = int(idx_raw)
            except ValueError:
                print("  数字で入力してください。末尾に追加します。")

        try:
            result = self.excel.create_sheet(name_test, index=index)
            print(f"  → 実際に追加できました: {result}")
            self.steps.append({
                "handler": "excel", "action": "create_sheet",
                "params": {"sheet_name": name_param, "index": index},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_rename_sheet(self) -> None:
        old_result = self._ask_sluttable_value("名前を変更する既存シート名")
        if old_result is None:
            print("  → キャンセルしました。\n")
            return
        old_test, old_param = old_result
        new_result = self._ask_sluttable_value("変更後の新しいシート名")
        if new_result is None:
            print("  → キャンセルしました。\n")
            return
        new_test, new_param = new_result

        try:
            result = self.excel.rename_sheet(old_test, new_test)
            print(f"  → 実際に変更できました: {result}")
            self.steps.append({
                "handler": "excel", "action": "rename_sheet",
                "params": {"old_name": old_param, "new_name": new_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_get_range_as_list(self) -> None:
        print("  1列(例: A1:A10)または1行(例: A1:J1)の範囲を指定してください。")
        print("  取得したリストはPythonと同じ0始まりの番号で参照できます"
              "({{変数名[0]}}が先頭要素)。")
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        range_result = self._ask_sluttable_value("範囲(1列または1行)")
        if range_result is None:
            print("  → キャンセルしました。\n")
            return
        range_test, range_param = range_result

        try:
            result = self.excel.get_range_as_list(sheet_test, range_test)
            preview = result[:5]
            more = "..." if len(result) > 5 else ""
            print(f"  → 実際に取得できました({len(result)}件): {preview}{more}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, result)
            step = {
                "handler": "excel", "action": "get_range_as_list",
                "params": {"sheet_name": sheet_param, "cell_range": range_param},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_copy_sheet(self) -> None:
        src_result = self._ask_sluttable_value("コピー元のシート名")
        if src_result is None:
            print("  → キャンセルしました。\n")
            return
        src_test, src_param = src_result
        new_result = self._ask_sluttable_value("新しいシート名")
        if new_result is None:
            print("  → キャンセルしました。\n")
            return
        new_test, new_param = new_result

        try:
            result = self.excel.copy_sheet(src_test, new_test)
            print(f"  → 実際にシートをコピーできました: {result}")
            self.steps.append({
                "handler": "excel", "action": "copy_sheet",
                "params": {"source_sheet_name": src_param, "new_sheet_name": new_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_get_sheet_names(self) -> None:
        try:
            names = self.excel.get_sheet_names()
            print(f"  → 実際に取得できました: {names}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, names)
            step = {"handler": "excel", "action": "get_sheet_names", "params": {}}
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_set_print_area(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        range_result = self._ask_sluttable_value("印刷範囲(例: A1:H30)")
        if range_result is None:
            print("  → キャンセルしました。\n")
            return
        range_test, range_param = range_result

        try:
            result = self.excel.set_print_area(sheet_test, range_test)
            print(f"  → 実際に設定できました: {result}")
            self.steps.append({
                "handler": "excel", "action": "set_print_area",
                "params": {"sheet_name": sheet_param, "cell_range": range_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_find_cell(self) -> None:
        sheet_result = self._ask_sluttable_value("検索対象のシート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        value_result = self._ask_sluttable_value("検索する値")
        if value_result is None:
            print("  → キャンセルしました。\n")
            return
        value_test, value_param = value_result
        column = self._ask("  検索する列(例: A。空欄ならシート全体を検索): ").strip()

        try:
            address = self.excel.find_cell(sheet_test, value_test, column=column or None)
            print(f"  → 実際に見つかりました: {address}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, address)
            step = {
                "handler": "excel", "action": "find_cell",
                "params": {"sheet_name": sheet_param, "value": value_param, "column": column or None},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_append_row(self) -> None:
        sheet_result = self._ask_sluttable_value("追記先のシート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result

        print("  追記する値を、左の列から順に入力します(空Enterで入力終了)")
        values_test: list[str] = []
        values_param: list[str] = []
        while True:
            result = self._ask_sluttable_value(f"{len(values_test) + 1}列目の値(空Enterで入力終了)")
            if result is None:
                continue
            test_v, param_v = result
            if not test_v and not param_v:
                break
            values_test.append(test_v)
            values_param.append(param_v)
        if not values_param:
            print("  → 値が1つも入力されなかったため、この手順は登録しませんでした。\n")
            return
        start_column = self._ask("  何列目から書き込みますか?(例: A。空Enterで A から): ").strip() or "A"

        try:
            result_msg = self.excel.append_row(sheet_test, values_test, start_column=start_column)
            print(f"  → 実際に追記できました: {result_msg}")
            self.steps.append({
                "handler": "excel", "action": "append_row",
                "params": {"sheet_name": sheet_param, "values": values_param, "start_column": start_column},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_insert_rows(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        row_result = self._ask_sluttable_value("何行目の前に挿入しますか(行番号)")
        if row_result is None:
            print("  → キャンセルしました。\n")
            return
        row_test, row_param = row_result
        count_raw = self._ask("  何行挿入しますか?(空Enterで1行): ").strip()
        count = int(count_raw) if count_raw.isdigit() else 1

        try:
            result = self.excel.insert_rows(sheet_test, int(row_test), count=count)
            print(f"  → 実際に挿入できました: {result}")
            self.steps.append({
                "handler": "excel", "action": "insert_rows",
                "params": {"sheet_name": sheet_param, "row": row_param, "count": count},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_delete_rows(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        row_result = self._ask_sluttable_value("何行目から削除しますか(行番号)")
        if row_result is None:
            print("  → キャンセルしました。\n")
            return
        row_test, row_param = row_result
        count_raw = self._ask("  何行削除しますか?(空Enterで1行): ").strip()
        count = int(count_raw) if count_raw.isdigit() else 1

        confirm = self._ask(
            f"  ⚠ これから実際にシート '{sheet_test}' の{row_test}行目から{count}行を削除して"
            f"動作確認します(元に戻せません)。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 削除を中止しました。この手順は登録しません。\n")
            return

        try:
            result = self.excel.delete_rows(sheet_test, int(row_test), count=count)
            print(f"  → 実際に削除できました: {result}")
            self.steps.append({
                "handler": "excel", "action": "delete_rows",
                "params": {"sheet_name": sheet_param, "row": row_param, "count": count},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_clear_range(self) -> None:
        sheet_result = self._ask_sluttable_value("対象シート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result
        range_result = self._ask_sluttable_value("空にする範囲(例: A2:C10)")
        if range_result is None:
            print("  → キャンセルしました。\n")
            return
        range_test, range_param = range_result

        try:
            result = self.excel.clear_range(sheet_test, range_test)
            print(f"  → 実際に空にできました: {result}")
            self.steps.append({
                "handler": "excel", "action": "clear_range",
                "params": {"sheet_name": sheet_param, "cell_range": range_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_excel_delete_sheet(self) -> None:
        sheet_result = self._ask_sluttable_value("削除するシート名")
        if sheet_result is None:
            print("  → キャンセルしました。\n")
            return
        sheet_test, sheet_param = sheet_result

        confirm = self._ask(
            f"  ⚠ これから実際にシート '{sheet_test}' を削除して動作確認します"
            f"(元に戻せません)。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 削除を中止しました。この手順は登録しません。\n")
            return

        try:
            result = self.excel.delete_sheet(sheet_test)
            print(f"  → 実際に削除できました: {result}")
            self.steps.append({
                "handler": "excel", "action": "delete_sheet",
                "params": {"sheet_name": sheet_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    # ---------- PDF操作 ----------

    def _record_pdf_menu(self) -> None:
        while True:
            print("PDFでは何をしますか?")
            print("  1) PDFのテキストを抽出する(ページ全体)")
            print("  2) 複数のPDFを結合する")
            print("  3) PDFを分割する")
            print("  4) PDFを回転する")
            print("  5) PDFをOCRでテキスト化する(ページ全体・スキャンPDF向け)")
            print("  6) 範囲を指定してテキストを取得する(座標%指定、OCR可)")
            print("  7) ページ範囲を1ファイルに抜き出す")
            print("  8) ページ数を取得する")
            print("  9) 表(罫線あり)をCSVとして抽出する")
            print("  10) 埋め込み画像を抽出する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_pdf_extract_text()
            elif choice == "2":
                self._record_pdf_merge()
            elif choice == "3":
                self._record_pdf_split()
            elif choice == "4":
                self._record_pdf_rotate()
            elif choice == "5":
                self._record_pdf_ocr()
            elif choice == "6":
                self._record_pdf_extract_area()
            elif choice == "7":
                self._record_pdf_extract_range()
            elif choice == "8":
                self._record_pdf_get_page_count()
            elif choice == "9":
                self._record_pdf_extract_tables()
            elif choice == "10":
                self._record_pdf_extract_images()
            elif choice == "0":
                return
            else:
                print("0〜10のいずれかを入力してください。\n")

    def _record_pdf_get_page_count(self) -> None:
        r1 = self._ask_sluttable_value("対象PDFのパス")
        if r1 is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = r1

        try:
            count = self.pdf.get_page_count(in_test)
            print(f"  → 実際に取得できました: {count}ページ")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, count)
            step = {"handler": "pdf", "action": "get_page_count", "params": {"input_path": in_param}}
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_extract_tables(self) -> None:
        r1 = self._ask_sluttable_value("表を抽出する元PDFのパス")
        if r1 is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = r1
        r2 = self._ask_sluttable_value("出力するCSVファイルのパス")
        if r2 is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = r2
        page_raw = self._ask("  対象ページ番号(空欄なら全ページ): ").strip()
        page_number = int(page_raw) if page_raw.isdigit() else None

        try:
            self.pdf.extract_tables(in_test, out_test, page_number=page_number)
            print(f"  → 実際に抽出できました: {out_test}")
            self.steps.append({
                "handler": "pdf", "action": "extract_tables",
                "params": {"input_path": in_param, "output_path": out_param, "page_number": page_number},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_extract_images(self) -> None:
        r1 = self._ask_sluttable_value("画像を抽出する元PDFのパス")
        if r1 is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = r1
        r2 = self._ask_sluttable_value("出力先フォルダのパス")
        if r2 is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = r2
        page_raw = self._ask("  対象ページ番号(空欄なら全ページ): ").strip()
        page_number = int(page_raw) if page_raw.isdigit() else None

        try:
            outputs = self.pdf.extract_images(in_test, out_test, page_number=page_number)
            print(f"  → 実際に抽出できました({len(outputs)}件)")
            self.steps.append({
                "handler": "pdf", "action": "extract_images",
                "params": {"input_path": in_param, "output_dir": out_param, "page_number": page_number},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_extract_text(self) -> None:
        r1 = self._ask_sluttable_value("抽出元PDFのパス")
        if r1 is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = r1

        r2 = self._ask_sluttable_value("出力するテキストファイルのパス")
        if r2 is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = r2

        try:
            self.pdf.extract_text(in_test, out_test)
            print(f"  → 実際にテキストを抽出できました: {out_test}")
            self.steps.append({
                "handler": "pdf", "action": "extract_text",
                "params": {"input_path": in_param, "output_path": out_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_merge(self) -> None:
        print("  結合するPDFのパスを1つずつ入力してください(空Enterで入力終了、2件以上必要です)")
        input_paths: list[str] = []
        while True:
            path = self._ask(f"  {len(input_paths) + 1}件目のPDFパス(空Enterで入力終了): ")
            if not path:
                break
            input_paths.append(path)

        if len(input_paths) < 2:
            print("  → 結合するPDFが2件未満のため、この手順は登録しませんでした。\n")
            return

        out_result = self._ask_sluttable_value("結合後の出力先パス")
        if out_result is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = out_result

        try:
            self.pdf.merge_pdfs(input_paths, out_test)
            print(f"  → 実際に結合できました: {out_test}")
            self.steps.append({
                "handler": "pdf", "action": "merge_pdfs",
                # パスの一覧は固定値のみ対応(スロット化は対象外)
                "params": {"input_paths": input_paths, "output_path": out_param},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_split(self) -> None:
        in_result = self._ask_sluttable_value("分割元PDFのパス")
        if in_result is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = in_result

        out_result = self._ask_sluttable_value("分割後の出力先フォルダ")
        if out_result is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = out_result

        raw = self._ask("  何ページごとに分割しますか?(数字。空Enterで1ページごと=全ページ個別出力): ")
        try:
            pages_per_file = int(raw) if raw else 1
        except ValueError:
            pages_per_file = 1

        range_raw = self._ask("  対象ページ範囲(例: 3,10 で3〜10ページのみ。空Enterで全ページ): ")
        start_page = end_page = None
        if range_raw:
            try:
                start_str, end_str = [x.strip() for x in range_raw.split(",")]
                start_page, end_page = int(start_str), int(end_str)
            except ValueError:
                print("  形式が正しくないため、全ページを対象にします。")

        pattern = self._ask(
            "  出力ファイル名のルール(空Enterで既定 '{stem}_part{part}' 。"
            "使えるプレースホルダ: {stem}{page}{page_end}{part}、例: invoice_{page:03d}): "
        ).strip() or None

        try:
            outputs = self.pdf.split_pdf(
                in_test, out_test, pages_per_file,
                start_page=start_page, end_page=end_page, filename_pattern=pattern,
            )
            print(f"  → 実際に{len(outputs)}件に分割できました。例: {outputs[0]}")
            self.steps.append({
                "handler": "pdf", "action": "split_pdf",
                "params": {
                    "input_path": in_param, "output_dir": out_param,
                    "pages_per_file": pages_per_file,
                    "start_page": start_page, "end_page": end_page,
                    "filename_pattern": pattern,
                },
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_extract_range(self) -> None:
        in_result = self._ask_sluttable_value("抜き出し元PDFのパス")
        if in_result is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = in_result

        out_result = self._ask_sluttable_value("出力先のパス")
        if out_result is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = out_result

        start_raw = self._ask("  開始ページ番号: ")
        end_raw = self._ask("  終了ページ番号: ")
        try:
            start_page, end_page = int(start_raw), int(end_raw)
        except ValueError:
            print("  ⚠ ページ番号は数字で入力してください\n")
            return

        try:
            self.pdf.extract_page_range(in_test, out_test, start_page, end_page)
            print(f"  → 実際に{start_page}〜{end_page}ページを抜き出せました: {out_test}")
            self.steps.append({
                "handler": "pdf", "action": "extract_page_range",
                "params": {
                    "input_path": in_param, "output_path": out_param,
                    "start_page": start_page, "end_page": end_page,
                },
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_rotate(self) -> None:
        in_result = self._ask_sluttable_value("回転元PDFのパス")
        if in_result is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = in_result

        out_result = self._ask_sluttable_value("回転後の出力先パス")
        if out_result is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = out_result

        raw = self._ask("  何度回転しますか?(90/180/270。空Enterで90度): ")
        try:
            degrees = int(raw) if raw else 90
        except ValueError:
            degrees = 90

        pages_raw = self._ask("  対象ページ番号(例: 1,3。空Enterで全ページ): ")
        pages: list[int] | None = None
        if pages_raw:
            try:
                pages = [int(x.strip()) for x in pages_raw.split(",") if x.strip()]
            except ValueError:
                print("  ページ番号の形式が正しくないため、全ページを対象にします。")
                pages = None

        try:
            self.pdf.rotate_pdf(in_test, out_test, degrees, pages)
            print(f"  → 実際に回転できました: {out_test}")
            self.steps.append({
                "handler": "pdf", "action": "rotate_pdf",
                "params": {
                    "input_path": in_param, "output_path": out_param,
                    "degrees": degrees, "pages": pages,
                },
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_pdf_ocr(self) -> None:
        print("  ※ OCRには pytesseract / pdf2image と、OS側の Tesseract OCR・Poppler が必要です。")
        in_result = self._ask_sluttable_value("OCR対象のPDFのパス")
        if in_result is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = in_result

        out_result = self._ask_sluttable_value("出力するテキストファイルのパス")
        if out_result is None:
            print("  → キャンセルしました。\n")
            return
        out_test, out_param = out_result

        lang_raw = self._ask("  OCRの言語を入力してください(空Enterで既定値 'jpn+eng'): ")
        language = lang_raw or "jpn+eng"

        try:
            self.pdf.ocr_pdf_to_text(in_test, out_test, language)
            print(f"  → 実際にOCRでテキストを抽出できました: {out_test}")
            self.steps.append({
                "handler": "pdf", "action": "ocr_pdf_to_text",
                "params": {"input_path": in_param, "output_path": out_param, "language": language},
            })
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                self.steps.append({
                    "handler": "pdf", "action": "ocr_pdf_to_text",
                    "params": {"input_path": in_param, "output_path": out_param, "language": language},
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。\n")

    def _record_pdf_extract_area(self) -> None:
        print("  ※ 基準点はページ左上を(0,0)とし、x/yはページ幅・高さに対する%(0〜100)で指定します。")
        print("     例: 右下1/4を取得したい場合 → x_left=50, x_right=100, y_upper=50, y_lower=100")
        in_result = self._ask_sluttable_value("対象PDFのパス")
        if in_result is None:
            print("  → キャンセルしました。\n")
            return
        in_test, in_param = in_result

        def ask_pct(label: str) -> float:
            raw = self._ask(f"  {label}(0〜100): ")
            try:
                return float(raw)
            except ValueError:
                return 0.0

        x_left = ask_pct("x_left")
        x_right = ask_pct("x_right")
        y_upper = ask_pct("y_upper")
        y_lower = ask_pct("y_lower")

        page_raw = self._ask("  対象ページ番号(空Enterで1ページ目): ")
        try:
            page_number = int(page_raw) if page_raw else 1
        except ValueError:
            page_number = 1

        ocr = self._ask("  文字が選択できないPDF(スキャン等)ですか? OCRを使う場合はy (y/N): ").lower() == "y"
        language = "jpn+eng"
        if ocr:
            lang_raw = self._ask("  OCRの言語(空Enterで既定値 'jpn+eng'): ")
            language = lang_raw or "jpn+eng"

        try:
            result = self.pdf.extract_text_in_area(
                in_test, x_left, x_right, y_upper, y_lower,
                page_number=page_number, ocr=ocr, ocr_language=language,
            )
            print(f"  → 実際に取得できました: {result[:150]!r}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, result)
            step = {
                "handler": "pdf", "action": "extract_text_in_area",
                "params": {
                    "input_path": in_param, "x_left": x_left, "x_right": x_right,
                    "y_upper": y_upper, "y_lower": y_lower, "page_number": page_number,
                    "ocr": ocr, "ocr_language": language,
                },
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    # ---------- Webサイト操作 ----------

    def _record_web_menu(self) -> None:
        if not self._ensure_site_opened():
            return
        while True:
            print("Webサイトでは何をしますか?")
            print("  1) ボタン/リンクをクリックする")
            print("  2) 入力欄に文字を入力する")
            print("  3) ドロップダウンから選択する")
            print("  4) 数秒待つ")
            print("  5) 今の画面をPDFとして保存する")
            print("  6) 画面から文字を読み取る(CSSセレクタ指定)")
            print("  7) 画面から属性値を読み取る(href/value等、CSSセレクタ指定)")
            print("  8) 画面から文字のリストを読み取る(表の1列など、CSSセレクタ指定)")
            print("  9) チェックボックスをON/OFFする")
            print("  10) ウィンドウサイズを指定する(pyautogui併用時の座標合わせ用)")
            print("  11) ウィンドウ位置を指定する(pyautogui併用時の座標合わせ用)")
            print("  12) 表示倍率(ズーム)を指定する")
            print("  13) 番号指定で操作する(表のセル読み取り・ボタン/入力欄/")
            print("      チェックボックス/トグル/プルダウンをプレビューして選ぶ)")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_click()
            elif choice == "2":
                self._record_type()
            elif choice == "3":
                self._record_select()
            elif choice == "4":
                self._record_wait()
            elif choice == "5":
                self._record_save_page_pdf()
            elif choice == "6":
                self._record_get_text_by_selector()
            elif choice == "7":
                self._record_get_attribute_by_selector()
            elif choice == "8":
                self._record_get_text_list_by_selector()
            elif choice == "9":
                self._record_check_checkbox()
            elif choice == "10":
                self._record_set_window_size()
            elif choice == "11":
                self._record_set_window_position()
            elif choice == "12":
                self._record_set_zoom()
            elif choice == "13":
                self._record_index_menu()
            elif choice == "0":
                return
            else:
                print("0〜13のいずれかを入力してください。\n")

    # ---------- 番号指定(インデックス)によるプレビュー付き操作 ----------

    def _pick_index(self, items: list[dict], describe) -> int | None:
        """list_*系の戻り値を describe(item)->str で整形して一覧表示し、
        番号を選んでもらう。キャンセル(空Enter)や無効な入力ならNoneを返す。
        """
        if not items:
            print("  (該当する要素が見つかりませんでした)\n")
            return None
        for item in items:
            print(f"  {item['index']}. {describe(item)}")
        raw = self._ask("  番号(空Enterでキャンセル)> ").strip()
        if not raw:
            return None
        try:
            idx = int(raw)
        except ValueError:
            print("  数字で入力してください。\n")
            return None
        if not any(it["index"] == idx for it in items):
            print("  その番号は一覧にありません。\n")
            return None
        return idx

    def _record_index_menu(self) -> None:
        while True:
            print("番号指定では何を操作しますか?"
                  "(表示テキストでの目印が使いにくい場合向け。まず一覧をプレビューしてから選びます)")
            print("  1) 表のセルを読み取る")
            print("  2) ボタン/リンクをクリックする")
            print("  3) 入力欄に入力する")
            print("  4) チェックボックスをON/OFFする")
            print("  5) トグルボタンをON/OFFする")
            print("  6) プルダウンから選択する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_index_table_cell()
            elif choice == "2":
                self._record_index_click()
            elif choice == "3":
                self._record_index_type()
            elif choice == "4":
                self._record_index_checkbox()
            elif choice == "5":
                self._record_index_toggle()
            elif choice == "6":
                self._record_index_select()
            elif choice == "0":
                return
            else:
                print("0〜6のいずれかを入力してください。\n")

    def _record_index_table_cell(self) -> None:
        try:
            tables = self.browser.list_tables()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        table_idx = self._pick_index(
            tables,
            lambda t: f"{t['rows']}行 x {t['columns']}列(先頭行のプレビュー: {', '.join(t['preview'])})",
        )
        if table_idx is None:
            print("  → キャンセルしました。\n")
            return
        row_raw = self._ask("  何行目を読み取りますか?(1始まり): ").strip()
        col_raw = self._ask("  何列目を読み取りますか?(1始まり): ").strip()
        if not row_raw.isdigit() or not col_raw.isdigit():
            print("  → 行・列は数字で入力してください。この手順は登録しませんでした。\n")
            return
        row, col = int(row_raw), int(col_raw)
        try:
            value = self.browser.get_table_cell_text(table_idx, row, col)
            print(f"  → プレビュー(実際に取得できた文字): {value!r}")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        if self._ask("  この内容で登録しますか? (Y/n): ").strip().lower() == "n":
            print("  → 登録しませんでした。\n")
            return
        store_as = self._ask_store_as()
        step = {
            "handler": "browser", "action": "get_table_cell_text",
            "params": {"table_index": table_idx, "row": row, "column": col},
        }
        if store_as:
            step["store_as"] = store_as
            self.record_variable(store_as, value)
        self.steps.append(step)
        print("  → 登録しました。\n")

    def _record_index_click(self) -> None:
        try:
            elements = self.browser.list_clickable_elements()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        idx = self._pick_index(elements, lambda el: f"[{el['tag']}] {el['text'] or '(表示文字なし)'}")
        if idx is None:
            print("  → キャンセルしました。\n")
            return
        try:
            self.browser.click_by_index(idx)
            print(f"  → 実際に{idx}番目をクリックして確認できました。")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() != "y":
                print("  → 登録しませんでした。\n")
                return
        verify_cfg = self._ask_verification()
        retry_cfg = self._ask_retry()
        obstruction_wait = self._ask_obstruction_wait()
        self.steps.append({
            "handler": "browser", "action": "click_by_index",
            "params": {"index": idx, "obstruction_wait_seconds": obstruction_wait},
            "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
        })
        print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    def _record_index_type(self) -> None:
        try:
            elements = self.browser.list_input_elements()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        idx = self._pick_index(
            elements,
            lambda el: f"[{el['type']}] {el['label'] or '(目印なし)'} (現在の値: {el['current_value']!r})",
        )
        if idx is None:
            print("  → キャンセルしました。\n")
            return
        result = self._ask_sluttable_value(f"{idx}番目の入力欄に入力する値")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_v, param_v = result
        press_enter = self._ask("  入力後にEnterキーで送信しますか? (y/N): ").lower() == "y"
        try:
            self.browser.type_by_index(idx, test_v, press_enter=press_enter)
            print(f"  → 実際に{idx}番目へ入力して確認できました。")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        verify_cfg = self._ask_verification()
        retry_cfg = self._ask_retry()
        self.steps.append({
            "handler": "browser", "action": "type_by_index",
            "params": {"index": idx, "value": param_v, "press_enter": press_enter},
            "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
        })
        print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    def _record_index_checkbox(self) -> None:
        try:
            elements = self.browser.list_checkbox_elements()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        idx = self._pick_index(
            elements,
            lambda el: f"{el['label'] or '(目印なし)'} (現在: {'ON' if el['checked'] else 'OFF'})",
        )
        if idx is None:
            print("  → キャンセルしました。\n")
            return
        checked = self._ask("  チェックを入れますか(N=外す)? (Y/n): ").strip().lower() != "n"
        try:
            self.browser.check_checkbox_by_index(idx, checked=checked)
            print(f"  → 実際に{idx}番目を{checked}にできました。")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        verify_cfg = self._ask_verification()
        retry_cfg = self._ask_retry()
        self.steps.append({
            "handler": "browser", "action": "check_checkbox_by_index",
            "params": {"index": idx, "checked": checked},
            "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
        })
        print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    def _record_index_toggle(self) -> None:
        try:
            elements = self.browser.list_toggle_elements()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        idx = self._pick_index(
            elements,
            lambda el: f"{el['label'] or '(目印なし)'} (現在: {'ON' if el['on'] else 'OFF'})",
        )
        if idx is None:
            print("  → キャンセルしました。\n")
            return
        on = self._ask("  ONにしますか(N=OFF)? (Y/n): ").strip().lower() != "n"
        try:
            self.browser.toggle_by_index(idx, on=on)
            print(f"  → 実際に{idx}番目を{on}にできました。")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        verify_cfg = self._ask_verification()
        retry_cfg = self._ask_retry()
        self.steps.append({
            "handler": "browser", "action": "toggle_by_index",
            "params": {"index": idx, "on": on},
            "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
        })
        print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    def _record_index_select(self) -> None:
        try:
            elements = self.browser.list_dropdown_elements()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        idx = self._pick_index(
            elements,
            lambda el: (
                f"{el['label'] or '(目印なし)'} 選択肢: {', '.join(el['options'])}"
                f" (現在: {el['selected']})"
            ),
        )
        if idx is None:
            print("  → キャンセルしました。\n")
            return
        option_text = self._ask("  選択したい項目の表示文字を入力してください: ").strip()
        try:
            self.browser.select_by_index(idx, option_text)
            print(f"  → 実際に{idx}番目で '{option_text}' を選択できました。")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        verify_cfg = self._ask_verification()
        retry_cfg = self._ask_retry()
        self.steps.append({
            "handler": "browser", "action": "select_by_index",
            "params": {"index": idx, "option_text": option_text},
            "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
        })
        print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    def _record_set_window_size(self) -> None:
        print("  ピクセルで指定するか、画面全体に対する割合(%)で指定するか選べます。")
        print("  割合を指定した場合はそちらが優先されます。")
        mode = self._ask("  割合(%)で指定しますか? (y/N): ").strip().lower()
        width = height = width_percent = height_percent = None
        if mode == "y":
            try:
                screen = self.browser.get_screen_size()
                print(f"  (参考: 画面解像度 {screen['width']}x{screen['height']})")
            except Exception:  # noqa: BLE001
                pass
            w_raw = self._ask("  幅の割合(%。空欄で変更しない): ").strip()
            h_raw = self._ask("  高さの割合(%。空欄で変更しない): ").strip()
            width_percent = float(w_raw) if w_raw else None
            height_percent = float(h_raw) if h_raw else None
        else:
            w_raw = self._ask("  幅(ピクセル。空欄で変更しない): ").strip()
            h_raw = self._ask("  高さ(ピクセル。空欄で変更しない): ").strip()
            width = int(w_raw) if w_raw.isdigit() else None
            height = int(h_raw) if h_raw.isdigit() else None

        try:
            result = self.browser.set_window_size(
                width=width, height=height, width_percent=width_percent, height_percent=height_percent
            )
            print(f"  → 実際に設定できました: {result}")
            self.steps.append({
                "handler": "browser", "action": "set_window_size",
                "params": {
                    "width": width, "height": height,
                    "width_percent": width_percent, "height_percent": height_percent,
                },
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_set_window_position(self) -> None:
        print("  画面の左上を基準(0,0)として、右方向・下方向がプラスです。")
        x_raw = self._ask("  X座標(間違えた場合は「キャンセル」): ")
        if x_raw in CANCEL_WORDS:
            print("  → キャンセルしました。\n")
            return
        y_raw = self._ask("  Y座標: ")
        try:
            x, y = int(x_raw), int(y_raw)
        except ValueError:
            print("  数字で入力してください。\n")
            return

        try:
            result = self.browser.set_window_position(x, y)
            print(f"  → 実際に設定できました: {result}")
            self.steps.append({
                "handler": "browser", "action": "set_window_position",
                "params": {"x": x, "y": y},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_set_zoom(self) -> None:
        print("  ※ 別のページに遷移すると設定はリセットされます"
              "(維持したい場合は遷移のたびに登録し直してください)。")
        raw = self._ask("  表示倍率を%で入力してください(空Enterで100=等倍): ").strip()
        try:
            percent = float(raw) if raw else 100.0
        except ValueError:
            print("  数字で入力してください。\n")
            return

        try:
            result = self.browser.set_zoom(percent)
            print(f"  → 実際に設定できました: {result}")
            self.steps.append({
                "handler": "browser", "action": "set_zoom",
                "params": {"percent": percent},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_get_text_by_selector(self) -> None:
        print("  F12の開発者ツールで調べたCSSセレクタ(class/id等)を指定してください。")
        selector = self._ask("  CSSセレクタ(間違えた場合は「キャンセル」): ")
        if selector in CANCEL_WORDS:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return
        try:
            text = self.browser.get_text_by_selector(selector)
            print(f"  → 実際に読み取れました: {text!r}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, text)
            step = {"handler": "browser", "action": "get_text_by_selector", "params": {"selector": selector}}
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_get_attribute_by_selector(self) -> None:
        print("  F12の開発者ツールで調べたCSSセレクタ(class/id等)を指定してください。")
        selector = self._ask("  CSSセレクタ(間違えた場合は「キャンセル」): ")
        if selector in CANCEL_WORDS:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return
        attribute = self._ask("  取得する属性名(例: href, value, data-id): ")
        try:
            value = self.browser.get_attribute_by_selector(selector, attribute)
            print(f"  → 実際に読み取れました: {value!r}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, value)
            step = {
                "handler": "browser", "action": "get_attribute_by_selector",
                "params": {"selector": selector, "attribute": attribute},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_get_text_list_by_selector(self) -> None:
        print("  一致するすべての要素の文字をリストとして読み取ります"
              "(例: 表の1列全体なら \"table tr td:nth-child(2)\")。")
        selector = self._ask("  CSSセレクタ(間違えた場合は「キャンセル」): ")
        if selector in CANCEL_WORDS:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return
        try:
            values = self.browser.get_text_list_by_selector(selector)
            preview = values[:5]
            more = "..." if len(values) > 5 else ""
            print(f"  → 実際に読み取れました({len(values)}件): {preview}{more}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, values)
            step = {
                "handler": "browser", "action": "get_text_list_by_selector",
                "params": {"selector": selector},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_check_checkbox(self) -> None:
        try:
            label_hint = self._ask_target(
                "  チェックボックスの目印(ラベルの表示文字)を入力してください"
                "(間違えた場合は「キャンセル」): "
            )
        except _ActionCancelled:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return
        checked = self._ask("  チェックを入れますか(N=外す)? (Y/n): ").strip().lower() != "n"

        try:
            self.browser.check_checkbox_by_text(label_hint, checked=checked)
            print(f"  → 実際に '{label_hint}' を {checked} にできました。")
            verify_cfg = self._ask_verification()
            retry_cfg = self._ask_retry()
            self.steps.append({
                "handler": "browser", "action": "check_checkbox_by_text",
                "params": {"label_hint": label_hint, "checked": checked},
                "verify": verify_cfg, "verify_skip": False,
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_click(self) -> None:
        try:
            candidates = self.browser.list_interactive_texts()
        except Exception:
            candidates = []
        if candidates:
            print("  今の画面でクリックできそうな表示文字(参考):")
            for t in candidates[:15]:
                print(f"    - {t}")

        try:
            text_hint = self._ask_target(
                "  押したいボタン/リンクに表示されている文字を入力してください"
                "(間違えた場合は「キャンセル」): "
            )
        except _ActionCancelled:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return

        try:
            self.browser.click_by_text(text_hint)
            print(f"  → 実際に '{text_hint}' をクリックして確認できました。")
            verify_cfg = self._ask_verification()
            retry_cfg = self._ask_retry()
            obstruction_wait = self._ask_obstruction_wait()
            self.steps.append({
                "handler": "browser", "action": "click_by_text",
                "params": {"text_hint": text_hint, "obstruction_wait_seconds": obstruction_wait},
                "verify": verify_cfg, "verify_skip": False,
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except ElementNotFoundError as e:
            print(f"  ⚠ {e}")
            selector = self._ask(
                "  見つかりませんでした。F12の開発者ツールで調べたCSSセレクタ"
                "(class/id等)があれば入力してください(不要ならそのままEnter): "
            )
            if selector:
                try:
                    self.browser.click_selector(selector)
                    print(f"  → CSSセレクタ '{selector}' でクリックを確認できました。")
                    verify_cfg = self._ask_verification()
                    retry_cfg = self._ask_retry()
                    obstruction_wait = self._ask_obstruction_wait()
                    self.steps.append({
                        "handler": "browser", "action": "click_selector",
                        "params": {"selector": selector, "obstruction_wait_seconds": obstruction_wait},
                        "verify": verify_cfg, "verify_skip": False,
                        "retry": retry_cfg,
                    })
                    print("  → 登録しました。\n")
                    return
                except Exception as e2:  # noqa: BLE001
                    print(f"  ⚠ CSSセレクタでも見つかりませんでした: {e2}")

            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                verify_cfg = self._ask_verification()
                retry_cfg = self._ask_retry()
                obstruction_wait = self._ask_obstruction_wait()
                self.steps.append({
                    "handler": "browser", "action": "click_by_text",
                    "params": {"text_hint": text_hint, "obstruction_wait_seconds": obstruction_wait},
                    "verify": verify_cfg, "verify_skip": False,
                    "retry": retry_cfg,
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。メニューに戻ります。\n")

    def _record_type(self) -> None:
        try:
            label_hint = self._ask_target(
                "  入力欄の目印(ラベル・プレースホルダーなど)を入力してください"
                "(間違えた場合は「キャンセル」): "
            )
        except _ActionCancelled:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return

        slot_name = self._ask(
            "  この値は実行するたびに変わりますか?"
            " 変わる場合はスロット名(例: order_no)を、固定値ならそのままEnter: "
        )

        if slot_name:
            if slot_name not in self.required_slots:
                self.required_slots.append(slot_name)
            test_value = self._ask(f"  動作確認用に '{slot_name}' に実際入力する値を入力してください: ")
            param_value = "{{" + slot_name + "}}"
        else:
            test_value = self._ask("  入力する固定値を入力してください: ")
            param_value = test_value

        press_enter = self._ask("  入力したあと、Enterキーで送信しますか? (y/N): ").lower() == "y"

        try:
            self.browser.type_by_text(label_hint, test_value, press_enter=press_enter)
            print(f"  → 実際に '{label_hint}' へ入力して確認できました。")
            verify_cfg = self._ask_verification()
            retry_cfg = self._ask_retry()
            self.steps.append({
                "handler": "browser", "action": "type_by_text",
                "params": {"label_hint": label_hint, "value": param_value, "press_enter": press_enter},
                "verify": verify_cfg, "verify_skip": False,
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except ElementNotFoundError as e:
            print(f"  ⚠ {e}")
            selector = self._ask(
                "  見つかりませんでした。F12の開発者ツールで調べたCSSセレクタ"
                "(class/id等)があれば入力してください(不要ならそのままEnter): "
            )
            if selector:
                try:
                    self.browser.type_by_selector(selector, test_value, press_enter=press_enter)
                    print(f"  → CSSセレクタ '{selector}' で入力を確認できました。")
                    verify_cfg = self._ask_verification()
                    retry_cfg = self._ask_retry()
                    self.steps.append({
                        "handler": "browser", "action": "type_by_selector",
                        "params": {"selector": selector, "value": param_value, "press_enter": press_enter},
                        "verify": verify_cfg, "verify_skip": False,
                        "retry": retry_cfg,
                    })
                    print("  → 登録しました。\n")
                    return
                except Exception as e2:  # noqa: BLE001
                    print(f"  ⚠ CSSセレクタでも見つかりませんでした: {e2}")

            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                verify_cfg = self._ask_verification()
                retry_cfg = self._ask_retry()
                self.steps.append({
                    "handler": "browser", "action": "type_by_text",
                    "params": {"label_hint": label_hint, "value": param_value, "press_enter": press_enter},
                    "verify": verify_cfg, "verify_skip": False,
                    "retry": retry_cfg,
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                if slot_name and slot_name in self.required_slots:
                    self.required_slots.remove(slot_name)
                print("  → 登録しませんでした。メニューに戻ります。\n")

    def _record_select(self) -> None:
        try:
            label_hint = self._ask_target(
                "  ドロップダウンの目印(ラベル名など)を入力してください"
                "(間違えた場合は「キャンセル」): "
            )
        except _ActionCancelled:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return

        option_text = self._ask("  選択したい選択肢の表示文字を入力してください: ")
        try:
            self.browser.select_by_text(label_hint, option_text)
            print(f"  → 実際に '{option_text}' を選択して確認できました。")
            verify_cfg = self._ask_verification()
            retry_cfg = self._ask_retry()
            self.steps.append({
                "handler": "browser", "action": "select_by_text",
                "params": {"label_hint": label_hint, "option_text": option_text},
                "verify": verify_cfg, "verify_skip": False,
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except ElementNotFoundError as e:
            print(f"  ⚠ {e}")
            selector = self._ask(
                "  見つかりませんでした。F12の開発者ツールで調べたCSSセレクタ"
                "(class/id等)があれば入力してください(不要ならそのままEnter): "
            )
            if selector:
                try:
                    self.browser.select_by_selector(selector, option_text)
                    print(f"  → CSSセレクタ '{selector}' で選択を確認できました。")
                    verify_cfg = self._ask_verification()
                    retry_cfg = self._ask_retry()
                    self.steps.append({
                        "handler": "browser", "action": "select_by_selector",
                        "params": {"selector": selector, "option_text": option_text},
                        "verify": verify_cfg, "verify_skip": False,
                        "retry": retry_cfg,
                    })
                    print("  → 登録しました。\n")
                except Exception as e2:  # noqa: BLE001
                    print(f"  ⚠ CSSセレクタでも見つかりませんでした: {e2}\n")

    def _record_save_page_pdf(self) -> None:
        result = self._ask_sluttable_value("保存先PDFのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        scale_raw = self._ask(
            "  倍率を指定しますか?(1ページに収めたい場合など。空Enterで既定値1.0): "
        )
        try:
            scale = float(scale_raw) if scale_raw else 1.0
        except ValueError:
            scale = 1.0

        landscape = self._ask("  横向きで保存しますか? (y/N): ").lower() == "y"

        try:
            self.browser.save_page_as_pdf(test_value, scale=scale, landscape=landscape)
            print(f"  → 実際にPDFとして保存できました: {test_value}")
            self.steps.append({
                "handler": "browser", "action": "save_page_as_pdf",
                "params": {"save_path": param_value, "scale": scale, "landscape": landscape},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_wait(self) -> None:
        raw = self._ask("  何秒待ちますか?(数字を入力、間違えた場合は「キャンセル」): ")
        if raw in CANCEL_WORDS:
            print("  → この操作の記録をキャンセルしました。メニューに戻ります。\n")
            return
        try:
            seconds = float(raw)
        except ValueError:
            print("  数字で入力してください。\n")
            return
        self.steps.append({
            "handler": "browser", "action": "wait_seconds",
            "params": {"seconds": seconds},
            "verify": {"type": "none"}, "verify_skip": False,
        })
        print(f"  → {seconds}秒の待機を登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    # ---------- エクスプローラー操作(ファイル/フォルダ) ----------

    def _record_explorer_menu(self) -> None:
        while True:
            print("エクスプローラーでは何をしますか?")
            print("  1) 指定したパスを開く")
            print("  2) 新しいフォルダを作成する")
            print("  3) ファイルを移動する")
            print("  4) ファイルをコピーする")
            print("  5) フォルダを移動する")
            print("  6) フォルダをコピーする")
            print("  7) ファイル名を変更する")
            print("  8) フォルダ名を変更する")
            print("  9) パスが存在するか確認する")
            print("  10) フォルダ内のファイル一覧を取得する")
            print("  11) ファイルを削除する")
            print("  12) フォルダを削除する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_explorer_open()
            elif choice == "2":
                self._record_explorer_create_folder()
            elif choice == "3":
                self._record_explorer_move(kind="file")
            elif choice == "4":
                self._record_explorer_copy(kind="file")
            elif choice == "5":
                self._record_explorer_move(kind="folder")
            elif choice == "6":
                self._record_explorer_copy(kind="folder")
            elif choice == "7":
                self._record_explorer_rename(kind="file")
            elif choice == "8":
                self._record_explorer_rename(kind="folder")
            elif choice == "9":
                self._record_explorer_path_exists()
            elif choice == "10":
                self._record_explorer_list_files()
            elif choice == "11":
                self._record_explorer_delete(kind="file")
            elif choice == "12":
                self._record_explorer_delete(kind="folder")
            elif choice == "0":
                return
            else:
                print("0〜12のいずれかを入力してください。\n")

    def _record_explorer_path_exists(self) -> None:
        result = self._ask_sluttable_value("確認するパス(ファイル or フォルダ)")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        try:
            exists = self.explorer.path_exists(test_value)
            print(f"  → 実際に確認できました: {exists}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, exists)
            step = {"handler": "explorer", "action": "path_exists", "params": {"path": param_value}}
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_list_files(self) -> None:
        result = self._ask_sluttable_value("一覧を取得するフォルダのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        pattern = self._ask("  絞り込みパターン(例: *.xlsx。空Enterですべて): ").strip() or "*"
        include_folders = self._ask("  サブフォルダも一覧に含めますか? (y/N): ").lower() == "y"

        try:
            files = self.explorer.list_files_in_folder(test_value, pattern=pattern, include_folders=include_folders)
            preview = files[:5]
            more = "..." if len(files) > 5 else ""
            print(f"  → 実際に取得できました({len(files)}件): {preview}{more}")
            store_as = self._ask_store_as()
            if store_as:
                self.record_variable(store_as, files)
            step = {
                "handler": "explorer", "action": "list_files_in_folder",
                "params": {"path": param_value, "pattern": pattern, "include_folders": include_folders},
            }
            if store_as:
                step["store_as"] = store_as
            self.steps.append(step)
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_delete(self, kind: str) -> None:
        label = "ファイル" if kind == "file" else "フォルダ"
        result = self._ask_sluttable_value(f"削除する{label}のパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        recursive = False
        if kind == "folder":
            recursive = self._ask(
                "  中にファイル/フォルダが残っていても中身ごと削除しますか?(通常はN) (y/N): "
            ).lower() == "y"

        confirm = self._ask(
            f"  ⚠ これから実際に '{test_value}' を削除して動作確認します"
            f"(ゴミ箱には移動されず、元に戻せません)。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 削除を中止しました。この手順は登録しません。\n")
            return

        action = "delete_file" if kind == "file" else "delete_folder"
        params: dict = {"path": param_value}
        if kind == "folder":
            params["recursive"] = recursive
        try:
            result_msg = getattr(self.explorer, action)(test_value, **({"recursive": recursive} if kind == "folder" else {}))
            print(f"  → 実際に削除できました: {result_msg}")
            self.steps.append({"handler": "explorer", "action": action, "params": params})
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_open(self) -> None:
        result = self._ask_sluttable_value("開くパス(フォルダ or ファイル)")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        try:
            self.explorer.open_path(test_value)
            print(f"  → 実際に開けました: {test_value}")
            self.steps.append({
                "handler": "explorer", "action": "open_path",
                "params": {"path": param_value},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_create_folder(self) -> None:
        result = self._ask_sluttable_value("作成するフォルダのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        exist_ok = self._ask("  既に存在する場合はそのまま使いますか?(通常はN) (y/N): ").lower() == "y"

        try:
            self.explorer.create_folder(test_value, exist_ok=exist_ok)
            print(f"  → 実際に作成できました: {test_value}")
            self.steps.append({
                "handler": "explorer", "action": "create_folder",
                "params": {"path": param_value, "exist_ok": exist_ok},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_move(self, kind: str) -> None:
        label = "ファイル" if kind == "file" else "フォルダ"
        src_result = self._ask_sluttable_value(f"移動元の{label}のパス")
        if src_result is None:
            print("  → キャンセルしました。\n")
            return
        src_test, src_param = src_result

        dest_result = self._ask_sluttable_value("移動先のパス")
        if dest_result is None:
            print("  → キャンセルしました。\n")
            return
        dest_test, dest_param = dest_result

        overwrite = self._ask(
            "  移動先に同名のものが既にある場合、上書きしますか?(通常はN) (y/N): "
        ).lower() == "y"

        confirm = self._ask(
            f"  ⚠ これから実際に '{src_test}' を '{dest_test}' へ移動して動作確認します"
            f"(元の場所からは無くなります)。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 移動を中止しました。この手順は登録しません。\n")
            return

        action = "move_file" if kind == "file" else "move_folder"
        try:
            result = getattr(self.explorer, action)(src_test, dest_test, overwrite=overwrite)
            print(f"  → 実際に移動できました: {result}")
            self.steps.append({
                "handler": "explorer", "action": action,
                "params": {
                    "source": src_param, "destination": dest_param, "overwrite": overwrite,
                },
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_copy(self, kind: str) -> None:
        label = "ファイル" if kind == "file" else "フォルダ"
        src_result = self._ask_sluttable_value(f"コピー元の{label}のパス")
        if src_result is None:
            print("  → キャンセルしました。\n")
            return
        src_test, src_param = src_result

        dest_result = self._ask_sluttable_value("コピー先のパス")
        if dest_result is None:
            print("  → キャンセルしました。\n")
            return
        dest_test, dest_param = dest_result

        overwrite = self._ask(
            "  コピー先に同名のものが既にある場合、上書きしますか?(通常はN) (y/N): "
        ).lower() == "y"

        action = "copy_file" if kind == "file" else "copy_folder"
        try:
            result = getattr(self.explorer, action)(src_test, dest_test, overwrite=overwrite)
            print(f"  → 実際にコピーできました: {result}")
            self.steps.append({
                "handler": "explorer", "action": action,
                "params": {
                    "source": src_param, "destination": dest_param, "overwrite": overwrite,
                },
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_explorer_rename(self, kind: str) -> None:
        label = "ファイル" if kind == "file" else "フォルダ"
        path_result = self._ask_sluttable_value(f"名前を変更する{label}のパス")
        if path_result is None:
            print("  → キャンセルしました。\n")
            return
        path_test, path_param = path_result

        name_result = self._ask_sluttable_value("変更後の名前(パスではなく名前だけ)")
        if name_result is None:
            print("  → キャンセルしました。\n")
            return
        name_test, name_param = name_result

        overwrite = self._ask(
            "  変更後の名前が既に使われている場合、上書きしますか?(通常はN) (y/N): "
        ).lower() == "y"

        confirm = self._ask(
            f"  ⚠ これから実際に '{path_test}' の名前を '{name_test}' に変更して動作確認します。"
            f" よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 名前変更を中止しました。この手順は登録しません。\n")
            return

        action = "rename_file" if kind == "file" else "rename_folder"
        try:
            result = getattr(self.explorer, action)(path_test, name_test, overwrite=overwrite)
            print(f"  → 実際に名前を変更できました: {result}")
            self.steps.append({
                "handler": "explorer", "action": action,
                "params": {
                    "path": path_param, "new_name": name_param, "overwrite": overwrite,
                },
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    # ---------- 実行ファイル(exe/py) ----------

    def _record_process_menu(self) -> None:
        while True:
            print("実行ファイル(exe/py)では何をしますか?")
            print("  1) 登録済み、または新しいexe/pyを実行する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_run_process()
            elif choice == "0":
                return
            else:
                print("0〜1のいずれかを入力してください。\n")

    def _choose_script(self) -> str | None:
        scripts = self.process.list_registered()
        if scripts:
            print("実行するスクリプト/実行ファイルを選んでください:")
            keys = list(scripts.keys())
            for i, key in enumerate(keys, start=1):
                info = scripts[key]
                print(f"  {i}. {key} ({info['kind']}: {info['path']})")
            print("  0. 新しく登録する")

            choice = self._ask("番号> ")
            if choice.strip() == "0":
                return self._register_new_script()
            try:
                idx = int(choice) - 1
                if idx < 0:
                    raise ValueError
                return keys[idx]
            except (ValueError, IndexError):
                print("  入力が正しくありません。もう一度選んでください。\n")
                return self._choose_script()
        else:
            print("  登録済みのスクリプトがまだありません。新しく登録します。")
            return self._register_new_script()

    def _register_new_script(self) -> str | None:
        run_key = self._ask("このスクリプト/実行ファイルの識別名を入力してください(半角英数字): ")
        if run_key in CANCEL_WORDS:
            print("  → キャンセルしました。\n")
            return None
        path = self._ask("ファイルのフルパスを入力してください(.py または .exe): ")
        kind_raw = self._ask("種類を入力してください(1: pyファイル / 2: exeファイル): ").strip()
        kind = "exe" if kind_raw == "2" else "python"
        try:
            self.process.register_script(run_key, path, kind)
            print(f"  → '{run_key}' を登録しました\n")
            return run_key
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return None

    def _record_run_process(self) -> None:
        run_key = self._choose_script()
        if run_key is None:
            return

        args_raw = self._ask("  実行時の引数があればスペース区切りで入力してください(不要ならそのままEnter): ")
        args = args_raw.split() if args_raw else []

        confirm = self._ask(
            f"  ⚠ これから実際に '{run_key}' を実行して動作確認します。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 実行確認をキャンセルしました。この手順は登録しません。\n")
            return

        try:
            output = self.process.run_registered(run_key, args=args, timeout=60)
            print(f"  → 実際に実行して確認できました。出力(先頭200文字): {output[:200]}")
            retry_cfg = self._ask_retry()
            self.steps.append({
                "handler": "process", "action": "run_registered",
                "params": {"run_key": run_key, "args": args},
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                self.steps.append({
                    "handler": "process", "action": "run_registered",
                    "params": {"run_key": run_key, "args": args},
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。\n")

    # ---------- デスクトップ操作(画面の画像でクリック等) ----------

    def _record_desktop_menu(self) -> None:
        print("  ※ 画面上の任意の位置をクリック・入力できてしまうため、対象画像は")
        print("     誤検出しにくい特徴的なものを用意してください。")
        while True:
            print("デスクトップでは何をしますか?")
            print("  1) 画面上の画像を探してクリックする")
            print("  2) 画面上の画像を探してマウスを移動する(クリックしない)")
            print("  3) 座標を直接指定してクリックする(解像度に依存するため非推奨)")
            print("  4) 文字列をキーボード入力する")
            print("  5) 特殊キーを送信する(例: enter, tab, ctrl+s)")
            print("  6) 今の画面のスクリーンショットを撮る(画像素材の準備用)")
            print("  7) 開いているウィンドウのタイトル一覧を見る(目印探し用)")
            print("  8) ウィンドウをアクティブにする(タイトル指定。Excel/PDF/エクスプローラー等)")
            print("  9) ウィンドウサイズを指定する(タイトル指定)")
            print("  10) ウィンドウ位置を指定する(タイトル指定)")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_desktop_click_image()
            elif choice == "2":
                self._record_desktop_move_image()
            elif choice == "3":
                self._record_desktop_click_coords()
            elif choice == "4":
                self._record_desktop_type()
            elif choice == "5":
                self._record_desktop_press_key()
            elif choice == "6":
                self._record_desktop_screenshot()
            elif choice == "7":
                self._record_desktop_list_titles()
            elif choice == "8":
                self._record_desktop_activate_window()
            elif choice == "9":
                self._record_desktop_set_window_size()
            elif choice == "10":
                self._record_desktop_set_window_position()
            elif choice == "0":
                return
            else:
                print("0〜10のいずれかを入力してください。\n")

    def _record_desktop_list_titles(self) -> None:
        try:
            titles = self.desktop.list_window_titles()
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return
        if not titles:
            print("  (今開いているウィンドウが見つかりませんでした)\n")
            return
        print("  今開いているウィンドウのタイトル:")
        for t in titles:
            print(f"    - {t}")
        print("  (この一覧取得自体はマクロの手順として登録されません)\n")

    def _record_desktop_activate_window(self) -> None:
        result = self._ask_sluttable_value("アクティブにしたいウィンドウのタイトル(部分一致)")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result
        try:
            self.desktop.activate_window_by_title(test_value)
            print(f"  → 実際にアクティブにできました: {test_value}")
            self.steps.append({
                "handler": "desktop", "action": "activate_window_by_title",
                "params": {"title_hint": param_value},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_desktop_set_window_size(self) -> None:
        result = self._ask_sluttable_value("対象ウィンドウのタイトル(部分一致。Excel/PDF/エクスプローラー等)")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        mode = self._ask("  割合(%)で指定しますか? (y/N): ").strip().lower()
        width = height = width_percent = height_percent = None
        if mode == "y":
            try:
                screen = self.desktop.get_screen_size()
                print(f"  (参考: 画面解像度 {screen['width']}x{screen['height']})")
            except Exception:  # noqa: BLE001
                pass
            w_raw = self._ask("  幅の割合(%。空欄で変更しない): ").strip()
            h_raw = self._ask("  高さの割合(%。空欄で変更しない): ").strip()
            width_percent = float(w_raw) if w_raw else None
            height_percent = float(h_raw) if h_raw else None
        else:
            w_raw = self._ask("  幅(ピクセル。空欄で変更しない): ").strip()
            h_raw = self._ask("  高さ(ピクセル。空欄で変更しない): ").strip()
            width = int(w_raw) if w_raw.isdigit() else None
            height = int(h_raw) if h_raw.isdigit() else None

        try:
            result_msg = self.desktop.set_window_size_by_title(
                test_value, width=width, height=height,
                width_percent=width_percent, height_percent=height_percent,
            )
            print(f"  → 実際に設定できました: {result_msg}")
            self.steps.append({
                "handler": "desktop", "action": "set_window_size_by_title",
                "params": {
                    "title_hint": param_value, "width": width, "height": height,
                    "width_percent": width_percent, "height_percent": height_percent,
                },
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_desktop_set_window_position(self) -> None:
        result = self._ask_sluttable_value("対象ウィンドウのタイトル(部分一致。Excel/PDF/エクスプローラー等)")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        print("  画面の左上を基準(0,0)として、右方向・下方向がプラスです。")
        x_raw = self._ask("  X座標: ")
        y_raw = self._ask("  Y座標: ")
        try:
            x, y = int(x_raw), int(y_raw)
        except ValueError:
            print("  数字で入力してください。\n")
            return

        try:
            result_msg = self.desktop.set_window_position_by_title(test_value, x, y)
            print(f"  → 実際に設定できました: {result_msg}")
            self.steps.append({
                "handler": "desktop", "action": "set_window_position_by_title",
                "params": {"title_hint": param_value, "x": x, "y": y},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _ask_desktop_verification(self, default_image_path: str, default_confidence: float) -> dict:
        print("  この操作が『本当に成功したか』をどう確認しますか?")
        print("    1) クリックした画像がこの後消える")
        print("    2) 別の画像がこの後新しく表示される")
        print("    3) 確認しない(後から個別に変更も可能)")
        choice = self._ask("  番号> ")
        if choice == "1":
            return {
                "type": "image_disappears",
                "value": f"{default_image_path}|{default_confidence}",
                "timeout": 10,
            }
        if choice == "2":
            img = self._ask("  表示されるはずの画像ファイルのパスを入力してください: ")
            if img:
                return {
                    "type": "image_appears",
                    "value": f"{img}|{default_confidence}",
                    "timeout": 10,
                }
        return {"type": "none"}

    def _ask_desktop_region(self) -> list[int] | None:
        """画像検索の対象を画面全体にするか、特定の矩形領域だけに絞るかを聞く。
        領域を絞ると、似たような画像が画面の他の場所にもある場合の誤検出を
        防げるほか、検索範囲が狭まる分わずかに速くなる。
        """
        try:
            screen = self.desktop.get_screen_size()
            print(f"  (参考: 画面解像度 {screen['width']}x{screen['height']})")
        except Exception:  # noqa: BLE001
            pass
        choice = self._ask(
            "  検索範囲は画面全体でよいですか? "
            "特定の領域だけに絞る場合は「領域」と入力してください(空Enterで画面全体): "
        ).strip()
        if choice not in ("領域", "region"):
            return None
        print("  領域の左上を基準(0,0)として、幅・高さをピクセルで指定してください。")
        try:
            left = int(self._ask("  領域の左端のX座標: "))
            top = int(self._ask("  領域の上端のY座標: "))
            width = int(self._ask("  領域の幅(ピクセル): "))
            height = int(self._ask("  領域の高さ(ピクセル): "))
        except ValueError:
            print("  数字で入力してください。画面全体を対象にします。\n")
            return None
        return [left, top, width, height]

    def _record_desktop_screenshot(self) -> None:
        result = self._ask_sluttable_value("保存先の画像パス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, _ = result
        region = self._ask_desktop_region()
        try:
            self.desktop.take_screenshot(test_value, region=region)
            print(f"  → スクリーンショットを保存しました: {test_value}")
            print("  → ここから対象のボタン/アイコン部分だけをトリミングして別ファイルに")
            print("     保存し、そのパスを「画像を探して...」の操作で指定してください。")
            print("  (このスクリーンショット取得自体はマクロの手順として登録されません)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_desktop_click_image(self) -> None:
        print("  ※ あらかじめ対象ボタン/アイコンを切り出した画像ファイルを用意してください")
        print("     (無ければ先に「6) スクリーンショットを撮る」で撮影し、トリミングできます)")
        result = self._ask_sluttable_value("クリックしたい対象が写っている画像ファイルのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        conf_raw = self._ask(
            "  一致の緩さ(confidence)を0.1〜1.0で指定してください(空Enterで既定値0.8): "
        )
        try:
            confidence = float(conf_raw) if conf_raw else 0.8
        except ValueError:
            confidence = 0.8
        region = self._ask_desktop_region()

        try:
            self.desktop.locate_and_click(test_value, confidence=confidence, timeout=10, region=region)
            print("  → 実際に画像を見つけてクリックできました。")
            verify_cfg = self._ask_desktop_verification(test_value, confidence)
            retry_cfg = self._ask_retry()
            self.steps.append({
                "handler": "desktop", "action": "locate_and_click",
                "params": {
                    "image_path": param_value, "confidence": confidence, "timeout": 10, "region": region,
                },
                "verify": verify_cfg, "verify_skip": False,
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() == "y":
                self.steps.append({
                    "handler": "desktop", "action": "locate_and_click",
                    "params": {
                        "image_path": param_value, "confidence": confidence, "timeout": 10, "region": region,
                    },
                })
                print("  → 未確認のまま登録しました。\n")
            else:
                print("  → 登録しませんでした。\n")

    def _record_desktop_move_image(self) -> None:
        result = self._ask_sluttable_value("マウスを移動したい対象が写っている画像ファイルのパス")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        conf_raw = self._ask(
            "  一致の緩さ(confidence)を0.1〜1.0で指定してください(空Enterで既定値0.8): "
        )
        try:
            confidence = float(conf_raw) if conf_raw else 0.8
        except ValueError:
            confidence = 0.8
        region = self._ask_desktop_region()

        try:
            self.desktop.move_to_image(test_value, confidence=confidence, timeout=10, region=region)
            print("  → 実際に画像を見つけてマウスを移動できました。")
            retry_cfg = self._ask_retry()
            self.steps.append({
                "handler": "desktop", "action": "move_to_image",
                "params": {
                    "image_path": param_value, "confidence": confidence, "timeout": 10, "region": region,
                },
                "retry": retry_cfg,
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_desktop_click_coords(self) -> None:
        x_raw = self._ask("  クリックしたいX座標を入力してください(間違えた場合は「キャンセル」): ")
        if x_raw in CANCEL_WORDS:
            print("  → キャンセルしました。\n")
            return
        y_raw = self._ask("  クリックしたいY座標を入力してください: ")
        try:
            x, y = int(x_raw), int(y_raw)
        except ValueError:
            print("  数字で入力してください。\n")
            return

        confirm = self._ask(
            f"  ⚠ 座標({x},{y})を実際にクリックして動作確認します"
            f"(画面解像度が変わると位置がずれます)。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 中止しました。この手順は登録しません。\n")
            return

        try:
            self.desktop.click_at(x, y)
            print(f"  → 実際にクリックできました: ({x},{y})")
            self.steps.append({
                "handler": "desktop", "action": "click_at",
                "params": {"x": x, "y": y},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_desktop_type(self) -> None:
        result = self._ask_sluttable_value("入力する文字列")
        if result is None:
            print("  → キャンセルしました。\n")
            return
        test_value, param_value = result

        confirm = self._ask(
            "  ⚠ 今フォーカスされている場所に実際にキー入力して動作確認します。よろしいですか? (y/N): "
        )
        if confirm.lower() != "y":
            print("  → 中止しました。この手順は登録しません。\n")
            return

        try:
            self.desktop.type_text(test_value)
            print("  → 実際に入力できました。")
            self.steps.append({
                "handler": "desktop", "action": "type_text",
                "params": {"text": param_value},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    def _record_desktop_press_key(self) -> None:
        key_raw = self._ask(
            "  送信するキーを入力してください(例: enter, tab, esc, ctrl+s。"
            "間違えた場合は「キャンセル」): "
        )
        if key_raw in CANCEL_WORDS:
            print("  → キャンセルしました。\n")
            return

        confirm = self._ask(f"  ⚠ '{key_raw}' を実際に送信して動作確認します。よろしいですか? (y/N): ")
        if confirm.lower() != "y":
            print("  → 中止しました。この手順は登録しません。\n")
            return

        try:
            self.desktop.press_key(key_raw)
            print(f"  → 実際に送信できました: {key_raw}")
            self.steps.append({
                "handler": "desktop", "action": "press_key",
                "params": {"key": key_raw},
            })
            print("  → 登録しました。(間違えていたら次のメニューで「12」から取り消せます)\n")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")

    # ---------- テキスト加工 ----------

    def _ask_text_input(self, label: str = "対象の文字列") -> tuple[str | None, str]:
        """テキスト加工の入力値を聞く。前の手順の結果を使う場合は
        {{変数名}} をそのまま含めて入力してもらう(値全体でも、
        "ID:{{no}}" のように他の文字と組み合わせても良い)。
        その場合、動作確認用に実際の値へ置き換えたものを別途聞く。
        戻り値: (動作確認に使う実際の文字列 or None, macros.jsonに書くparam値)
        """
        value = self._ask(
            f"  {label}を入力してください(前の手順の結果を使う場合はそのまま {{変数名}} と書けます): "
        )
        if "{{" in value and "}}" in value:
            test_value = self._ask(
                f"  動作確認用に、実際の値に置き換えたものを入力してください(例: {value}。"
                f"分からなければ空Enterで動作確認をスキップ): "
            )
            return (test_value or None), value
        return value, value

    def _record_text_menu(self) -> None:
        while True:
            print("テキスト加工では何をしますか?")
            print("  1) 文字を探して切り出す")
            print("  2) 文字を置換する")
            print("  3) 日付・時刻を取得する(yyyyMMdd_hhmmss等、ファイル名にも使える書式)")
            print("  4) 文字をつなげる/付加する")
            print("  5) クリップボードにコピーする")
            print("  6) クリップボードから取得する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_text_cut()
            elif choice == "2":
                self._record_text_replace()
            elif choice == "3":
                self._record_text_datetime()
            elif choice == "4":
                self._record_text_combine()
            elif choice == "5":
                self._record_text_copy_to_clipboard()
            elif choice == "6":
                self._record_text_get_from_clipboard()
            elif choice == "0":
                return
            else:
                print("0〜6のいずれかを入力してください。\n")

    def _record_text_cut(self) -> None:
        test_text, param_text = self._ask_text_input("切り出し対象の文字列")
        marker = self._ask("  探す文字(目印): ")
        if not marker:
            print("  → 目印が未入力のため、この手順は登録しませんでした。\n")
            return
        include = self._ask("  その文字を含めて切り出しますか? (y/N): ").lower() == "y"

        mode = self._ask(
            "  切り出し方法を選んでください\n"
            "    1) 文字数を指定する\n"
            "    2) 別の文字が出てくるまで\n"
            "    3) 目印から最後まですべて\n"
            "  番号> "
        )
        length = None
        end_marker = None
        if mode == "1":
            raw = self._ask("  切り出す文字数: ")
            try:
                length = int(raw)
            except ValueError:
                print("  ⚠ 数字で入力してください\n")
                return
        elif mode == "2":
            end_marker = self._ask("  終わりの目印となる文字: ")
            if not end_marker:
                print("  ⚠ 終わりの目印が未入力です\n")
                return

        store_as = self._ask_store_as()
        params = {"text": param_text, "marker": marker, "include_marker": include}
        if length is not None:
            params["length"] = length
        if end_marker:
            params["end_marker"] = end_marker

        result = None
        confirmed = False
        if test_text is not None:
            try:
                result = self.text.cut_from_marker(
                    test_text, marker, include_marker=include, length=length, end_marker=end_marker
                )
                print(f"  → 動作確認できました。結果: {result[:100]!r}")
                confirmed = True
            except Exception as e:  # noqa: BLE001
                print(f"  ⚠ {e}")
                if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() != "y":
                    print("  → 登録しませんでした。\n")
                    return
        else:
            print("  → 前の手順の結果を使うため、動作確認はスキップして登録します。")

        step = {"handler": "text", "action": "cut_from_marker", "params": params}
        if store_as:
            step["store_as"] = store_as
            if confirmed:
                self.record_variable(store_as, result)
        self.steps.append(step)
        print("  → 登録しました。\n")

    def _record_text_replace(self) -> None:
        test_text, param_text = self._ask_text_input("置換対象の文字列")
        search = self._ask("  探す文字: ")
        replace = self._ask("  置き換える文字: ")
        store_as = self._ask_store_as()

        if test_text is not None:
            result = self.text.replace_text(test_text, search, replace)
            print(f"  → 動作確認できました。結果: {result[:100]!r}")
        else:
            print("  → 前の手順の結果を使うため、動作確認はスキップして登録します。")

        step = {
            "handler": "text", "action": "replace_text",
            "params": {"text": param_text, "search": search, "replace": replace},
        }
        if store_as:
            step["store_as"] = store_as
            if test_text is not None:
                self.record_variable(store_as, result)
        self.steps.append(step)
        print("  → 登録しました。\n")

    def _record_text_datetime(self) -> None:
        print("  よく使う書式:")
        print("    1) yyyyMMdd_hhmmss  (例: 20260830_143022。ファイル名の一意化に)")
        print("    2) yyyyMMdd         (例: 20260830)")
        print("    3) hhmmss           (例: 143022)")
        print("    4) yyyy             (西暦年。例: 2026)")
        print("    5) YYYY             (年度。例: 2025)")
        print("    6) 自分で書式を入力する(yyyy/YYYY/MM/M/dd/d/hh/mm/ss を組み合わせ可)")
        choice = self._ask("  番号(空Enterで1)> ").strip() or "1"
        presets = {
            "1": "yyyyMMdd_hhmmss", "2": "yyyyMMdd", "3": "hhmmss", "4": "yyyy", "5": "YYYY",
        }
        if choice == "6":
            format_code = self._ask(
                "  書式コードを入力してください(例: yyyy年MM月dd日): "
            ).strip()
            if not format_code:
                print("  → 未入力のため、この手順は登録しませんでした。\n")
                return
        else:
            format_code = presets.get(choice, "yyyyMMdd_hhmmss")

        fy_month_raw = self._ask("  年度の開始月(YYYYを使わないなら無関係。空Enterで既定4月): ").strip()
        try:
            fy_month = int(fy_month_raw) if fy_month_raw else 4
        except ValueError:
            fy_month = 4

        try:
            result = self.text.format_now(format_code, fiscal_year_start_month=fy_month)
            print(f"  → 動作確認できました。結果: {result}")
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}\n")
            return

        store_as = self._ask_store_as()
        if store_as:
            self.record_variable(store_as, result)
        step = {
            "handler": "text", "action": "format_now",
            "params": {"format_code": format_code, "fiscal_year_start_month": fy_month},
        }
        if store_as:
            step["store_as"] = store_as
        self.steps.append(step)
        print("  → 登録しました。\n")

    def _record_text_combine(self) -> None:
        print("  つなげたい文字列を1つずつ入力します(空Enterで入力終了)")
        print("  前の手順の結果を使う場合は {{変数名}} の形で入力してください")
        parts: list[str] = []
        while True:
            part = self._ask(f"  {len(parts) + 1}件目(空Enterで入力終了): ")
            if not part:
                break
            parts.append(part)

        if len(parts) < 2:
            print("  → 2件以上入力してください。この手順は登録しませんでした。\n")
            return

        separator = self._ask("  区切り文字(空Enterでそのままつなげる): ")
        store_as = self._ask_store_as()

        has_variable = any("{{" in p and "}}" in p for p in parts)
        if not has_variable:
            result = self.text.combine_text(parts, separator=separator)
            print(f"  → 動作確認できました。結果: {result[:100]!r}")
        else:
            print("  → 前の手順の結果を使うため、動作確認はスキップして登録します。")

        step = {
            "handler": "text", "action": "combine_text",
            "params": {"parts": parts, "separator": separator},
        }
        if store_as:
            step["store_as"] = store_as
            if not has_variable:
                self.record_variable(store_as, result)
        self.steps.append(step)
        print("  → 登録しました。\n")

    def _record_text_copy_to_clipboard(self) -> None:
        test_text, param_text = self._ask_text_input("クリップボードにコピーする文字列")

        if test_text is not None:
            try:
                result = self.text.copy_to_clipboard(test_text)
                print(f"  → 実際にコピーできました: {result}")
            except Exception as e:  # noqa: BLE001
                print(f"  ⚠ {e}")
                if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() != "y":
                    print("  → 登録しませんでした。\n")
                    return
        else:
            print("  → 前の手順の結果を使うため、動作確認はスキップして登録します。")

        self.steps.append({"handler": "text", "action": "copy_to_clipboard", "params": {"text": param_text}})
        print("  → 登録しました。\n")

    def _record_text_get_from_clipboard(self) -> None:
        value = None
        confirmed = False
        try:
            value = self.text.get_from_clipboard()
            print(f"  → 実際に取得できました: {value[:100]!r}")
            confirmed = True
        except Exception as e:  # noqa: BLE001
            print(f"  ⚠ {e}")
            if self._ask("  それでもこの手順として登録しますか? (y/N): ").lower() != "y":
                print("  → 登録しませんでした。\n")
                return

        store_as = self._ask_store_as()
        if store_as and confirmed:
            self.record_variable(store_as, value)
        step = {"handler": "text", "action": "get_from_clipboard", "params": {}}
        if store_as:
            step["store_as"] = store_as
        self.steps.append(step)
        print("  → 登録しました。\n")

    # ---------- リスト(配列) ----------

    def _record_list_menu(self) -> None:
        while True:
            print("リストでは何をしますか?")
            print("  1) 空のリストを作成する")
            print("  2) リストに値を追加する")
            print("  3) リストの要素を取得する(0始まり)")
            print("  4) リストの長さを取得する")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_list_create_empty()
            elif choice == "2":
                self._record_list_append()
            elif choice == "3":
                self._record_list_get_item()
            elif choice == "4":
                self._record_list_length()
            elif choice == "0":
                return
            else:
                print("0〜4のいずれかを入力してください。\n")

    def _record_list_create_empty(self) -> None:
        print("  空のリストを作成します。後で「リストに追加する」で1件ずつ積み上げていけます")
        print("  (Excelのセル範囲をまとめて読み込みたいだけなら、Excelメニューの")
        print("   「セル範囲をリストとして取得する」の方が簡単です)。")
        store_as = self._ask("  作成するリストの変数名を入力してください(例: result): ").strip()
        if not store_as:
            print("  → 名前が無いと後で使えないため、この手順は登録しませんでした。\n")
            return
        self.steps.append({
            "handler": "list", "action": "create_empty", "params": {}, "store_as": store_as,
        })
        self.record_variable(store_as, [])
        print(f"  → 登録しました(変数名: {store_as})。\n")

    def _record_list_append(self) -> None:
        lst_var = self._ask("  追加先のリストが入っている変数名を入力してください(例: result): ").strip()
        if not lst_var:
            print("  → キャンセルしました。\n")
            return
        value_result = self._ask_sluttable_value("追加する値")
        if value_result is None:
            print("  → キャンセルしました。\n")
            return
        _, value_param = value_result
        store_as = self._ask(
            f"  結果を保存する変数名(空Enterで '{lst_var}' に上書き=そのリストへ追記していく形になります): "
        ).strip() or lst_var

        step = {
            "handler": "list", "action": "append",
            "params": {"lst": "{{" + lst_var + "}}", "value": value_param},
            "store_as": store_as,
        }
        # リストの中身は実行時にしか定まらないため、動作確認はせずそのまま登録する
        self.steps.append(step)
        print(f"  → 登録しました('{lst_var}' に値を追加し、'{store_as}' に保存します)。\n")

    def _record_list_get_item(self) -> None:
        lst_var = self._ask("  対象のリストが入っている変数名を入力してください: ").strip()
        if not lst_var:
            print("  → キャンセルしました。\n")
            return
        idx_result = self._ask_sluttable_value("取得したい位置(0始まり。例: 0、または {{i}})")
        if idx_result is None:
            print("  → キャンセルしました。\n")
            return
        _, idx_param = idx_result
        store_as = self._ask_store_as()
        step = {
            "handler": "list", "action": "get_item",
            "params": {"lst": "{{" + lst_var + "}}", "index": idx_param},
        }
        if store_as:
            step["store_as"] = store_as
        self.steps.append(step)
        print("  → 登録しました。(同じことは {{変数名[位置]}} の書き方でも直接できます)\n")

    def _record_list_length(self) -> None:
        lst_var = self._ask("  対象のリストが入っている変数名を入力してください: ").strip()
        if not lst_var:
            print("  → キャンセルしました。\n")
            return
        store_as = self._ask_store_as()
        step = {
            "handler": "list", "action": "length",
            "params": {"lst": "{{" + lst_var + "}}"},
        }
        if store_as:
            step["store_as"] = store_as
        self.steps.append(step)
        print("  → 登録しました。(同じことは {{変数名.length}} の書き方でも直接できます)\n")

    # ---------- 制御構文(For/If/Goto) ----------

    def _existing_labels(self) -> list[str]:
        return [
            s["params"]["name"] for s in self.steps
            if s.get("handler") == "control" and s.get("action") == "label"
        ]

    def _record_control_menu(self) -> None:
        while True:
            print("制御構文では何をしますか?(VBAのFor〜Next、If〜Then〜Else、Gotoに相当)")
            print("  1) ラベルを置く(ジャンプ先の目印)")
            print("  2) 指定したラベルへジャンプする(goto)")
            print("  3) 条件を満たしたらジャンプする(IF文)")
            print("  4) 変数に値を設定する(例: A = A + 1)")
            print("  5) 繰り返しを開始する(for)")
            print("  6) 繰り返しを終了する(next)")
            print("  7) 変数の型を変換する(文字列/整数/小数)")
            print("  0) 戻る")
            choice = self._ask("番号> ")
            print()

            if choice == "1":
                self._record_control_label()
            elif choice == "2":
                self._record_control_goto()
            elif choice == "3":
                self._record_control_if_goto()
            elif choice == "4":
                self._record_control_set_value()
            elif choice == "5":
                self._record_control_for_start()
            elif choice == "6":
                self._record_control_for_end()
            elif choice == "7":
                self._record_control_convert_type()
            elif choice == "0":
                return
            else:
                print("0〜7のいずれかを入力してください。\n")

    def _record_control_label(self) -> None:
        name = self._ask("  このラベルの名前を入力してください(例: Label1): ").strip()
        if not name:
            print("  → キャンセルしました。\n")
            return
        if name in self._existing_labels():
            print(f"  ⚠ 同じ名前のラベル '{name}' がこのマクロ内に既にあります。別の名前にしてください。\n")
            return
        self.steps.append({"handler": "control", "action": "label", "params": {"name": name}})
        print(f"  → ラベル '{name}' を置きました。(間違えていたら次のメニューで「12」から取り消せます)\n")

    def _record_control_goto(self) -> None:
        labels = self._existing_labels()
        if labels:
            print("  これまでに置いたラベル:", ", ".join(labels))
        name = self._ask(
            "  ジャンプ先のラベル名を入力してください"
            "(まだ置いていない場合は、この後の手順で置く予定の名前でも構いません): "
        ).strip()
        if not name:
            print("  → キャンセルしました。\n")
            return
        self.steps.append({"handler": "control", "action": "goto", "params": {"label": name}})
        print(f"  → '{name}' への無条件ジャンプを登録しました。\n")

    def _record_control_if_goto(self) -> None:
        print("  例:「IF A<>B THEN GOTO Label1」のような条件付きジャンプを登録します。")
        print("  (条件が不成立の場合はそのまま次の手順に進むため、続く手順が「ELSE」の")
        print("   代わりになります)")
        left = self._ask("  左辺(例: {{A}}): ").strip()
        print("  比較演算子を選んでください: 1)==  2)!=  3)<  4)<=  5)>  6)>=")
        op_choice = self._ask("  番号> ")
        op_map = {"1": "==", "2": "!=", "3": "<", "4": "<=", "5": ">", "6": ">="}
        op = op_map.get(op_choice, "==")
        right = self._ask("  右辺(例: {{B}}): ").strip()
        labels = self._existing_labels()
        if labels:
            print("  これまでに置いたラベル:", ", ".join(labels))
        label = self._ask("  条件が真のときにジャンプするラベル名: ").strip()
        if not left or not right or not label:
            print("  → 入力が不足しているため、この手順は登録しませんでした。\n")
            return
        self.steps.append({
            "handler": "control", "action": "if_goto",
            "params": {"left": left, "op": op, "right": right, "label": label},
        })
        print(f"  → 登録しました: IF {left} {op} {right} THEN GOTO {label}\n")

    def _record_control_set_value(self) -> None:
        var_name = self._ask("  値を設定する変数名を入力してください(例: A): ").strip()
        if not var_name:
            print("  → キャンセルしました。\n")
            return
        value = self._ask(
            f"  '{var_name}' に設定する値を入力してください"
            f"(例: {{{{{var_name}+1}}}} で現在値+1、固定値もそのまま入力可): "
        ).strip()
        if not value:
            print("  → キャンセルしました。\n")
            return
        self.steps.append({
            "handler": "control", "action": "set_value",
            "params": {"value": value}, "store_as": var_name,
        })
        print(f"  → 登録しました: {var_name} = {value}\n")

    def _record_control_convert_type(self) -> None:
        print("  Excelのセル値等は先頭が0の値(郵便番号等)を保持するため基本的に")
        print("  文字列として扱われます。計算に使いたい場合や、逆に確実に文字列")
        print("  として扱いたい場合に、変数の型を強制的に変換します。")
        var_name = self._ask("  変換する変数名を入力してください(例: A): ").strip()
        if not var_name:
            print("  → キャンセルしました。\n")
            return
        print("  変換先の型を選んでください:")
        print("    1) 文字列(str)")
        print("    2) 整数(int)")
        print("    3) 小数(float)")
        type_choice = self._ask("  番号> ")
        action = {"1": "to_str", "2": "to_int", "3": "to_float"}.get(type_choice)
        if action is None:
            print("  → 1〜3のいずれかを選んでください。この手順は登録しませんでした。\n")
            return
        value = "{{" + var_name + "}}"
        self.steps.append({
            "handler": "control", "action": action,
            "params": {"value": value}, "store_as": var_name,
        })
        print(f"  → 登録しました: {var_name} を{action}で変換します\n")

    def _record_control_for_start(self) -> None:
        var_name = self._ask("  ループカウンタの変数名(空Enterで既定 'i'): ").strip() or "i"
        start_raw = self._ask("  開始値(空Enterで0): ").strip() or "0"
        end_raw = self._ask(
            "  終了値(例: 10、またはリスト全件処理なら {{リスト変数名.length-1}} ): "
        ).strip()
        if not end_raw:
            print("  → 終了値が未入力のため、この手順は登録しませんでした。\n")
            return

        def _as_int_or_str(raw: str):
            try:
                return int(raw)
            except ValueError:
                return raw

        start_val = _as_int_or_str(start_raw)
        end_val = _as_int_or_str(end_raw)

        self.steps.append({
            "handler": "control", "action": "for_start",
            "params": {"var": var_name, "start": start_val, "end": end_val},
        })
        print(f"  → 登録しました: for {var_name} = {start_val} to {end_val}")
        print("  この後に繰り返したい手順を登録し、最後に「6) 繰り返しを終了する」を")
        print("  忘れずに登録してください。\n")

    def _record_control_for_end(self) -> None:
        # 対応する(まだ閉じていない)直近のfor_startを探す
        open_stack: list[str] = []
        for s in self.steps:
            if s.get("handler") != "control":
                continue
            if s.get("action") == "for_start":
                open_stack.append(s["params"].get("var", "i"))
            elif s.get("action") == "for_end":
                if open_stack:
                    open_stack.pop()
        if not open_stack:
            print("  ⚠ 対応する「繰り返しを開始する(for)」がまだ登録されていません。\n")
            return
        var_name = open_stack[-1]
        self.steps.append({"handler": "control", "action": "for_end", "params": {"var": var_name}})
        print(f"  → 繰り返し終了(for {var_name})を登録しました。\n")

    # ---------- 保存 ----------

    def _finish(self) -> str:
        if self._site_opened:
            self.browser.close()
            self.steps.append({"handler": "browser", "action": "close", "params": {}})

        print("最後に、このマクロの保存情報を教えてください。")
        macro_name = self._ask("保存時に使う名前(半角英数字。例: monthly_report)はどうしますか?: ")
        description = self._ask("この操作の説明を一言でお願いします(例: 月次レポートを作成する): ")
        keywords_raw = self._ask(
            "この操作を呼び出すときに使いそうな言葉をカンマ区切りで教えてください(例: 月次,レポート): "
        )
        keywords = [k.strip() for k in keywords_raw.split(",") if k.strip()]

        macro_def = {
            "description": description,
            "required_slots": self.required_slots,
            "steps": self.steps,
        }
        if any(s.get("handler") == "browser" for s in self.steps):
            # このマクロを記録したときに使っていたブラウザ(chrome/edge)をそのまま
            # 保存しておく。実行時、このマクロだけ自動でそのブラウザに切り替わる
            # (社内で作業によってEdge指定されている場合など、マクロごとに
            # 使うブラウザが違っても、同じセッション内で混在して実行できる)。
            macro_def["browser"] = self.browser.browser
        self._save_macro(macro_name, macro_def)
        self._save_intent(macro_name, description, keywords)

        print(f"\n✅ マクロ '{macro_name}' を保存しました。")
        print(f"   次回から「{description}」のような言葉で呼び出せます。")
        return macro_name

    def _save_macro(self, macro_name: str, macro_def: dict) -> None:
        path = self.config_dir / "macros.json"
        backup_file(path, self.config_dir / "backups")
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        data.setdefault("macros", {})[macro_name] = macro_def
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _save_intent(self, macro_name: str, description: str, keywords: list[str]) -> None:
        path = self.config_dir / "intents.json"
        backup_file(path, self.config_dir / "backups")
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        intents = data.setdefault("intents", [])
        # 同名IDが既にあれば置き換える
        intents = [i for i in intents if i.get("id") != macro_name]
        intents.append({
            "id": macro_name,
            "description": description,
            "keywords": keywords,
            "patterns": [],
            "macro": macro_name,
        })
        data["intents"] = intents
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
