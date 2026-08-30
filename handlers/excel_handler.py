"""
ExcelHandler: 基本はopenpyxlのみで完結するExcel操作ハンドラ。
.xlsm(マクロ有効ブック)はVBAを保持したまま読み書きできる(keep_vba)。

複数ブックの同時オープンとアクティブ切り替え:
  load_workbook() は毎回「エイリアス」を付けてメモリ上に保持し、最後に
  読み込んだ(または switch_workbook() で切り替えた)ブックが「アクティブ」
  になる。read_sheet_to_records / write_cells / save_workbook_as 等は
  常にアクティブなブックに対して行われる。これにより、2つ以上のExcelを
  開いたまま行き来しながら操作できる(例: A.xlsxを読み込む → B.xlsxを
  読み込む → switch_workbookでAに戻す → Aに書き込む → Bに切り替えて書き込む)。

RPAでよく使うセル/シート操作:
  get_cell_value(特定セル1つの読み込み)、get_last_row(最終行取得)、
  copy_cell_range(セルのコピー&貼り付け。値のみ/数式のみ/すべて(書式含む)
  から貼り付け方式を選べる)、copy_sheet(シートのコピー)、
  create_sheet(新規シート追加)、rename_sheet(既存シート名の変更)、
  get_sheet_names(シート名一覧)、set_print_area(印刷範囲の指定) に対応。

ただし以下の2つの操作だけは openpyxl では実現できない
(openpyxlはファイル形式としての読み書きのみで、実際にExcelアプリケーションを
描画・実行する機能を持たないため)。この2つに限り、実際のExcelアプリケーションを
COM経由で操作する(**Windows + Excelインストール環境専用**、pywin32が必要)。
これらは毎回パスを直接指定するため、複数ブックの「アクティブ切り替え」とは
独立して、どのファイルに対しても実行できる:

- save_as_pdf: あらかじめ設定された印刷範囲・ページ設定のままPDFへ書き出す
- run_excel_macro: .xlsmに既に組み込まれているVBAマクロを実行する

いずれもネットワークアクセスは一切行いません。
"""
from __future__ import annotations

import csv
import logging
from copy import copy as _copy_style
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

logger = logging.getLogger("rpa_local_ai.excel")

_MACRO_ENABLED_EXTENSIONS = {".xlsm", ".xltm"}


def _import_win32com():
    try:
        import win32com.client  # type: ignore
        return win32com.client
    except ImportError as e:
        raise RuntimeError(
            "この操作にはWindows上のExcel本体 + pywin32が必要です"
            "(pip install pywin32)。Excelがインストールされていない環境や"
            "Windows以外の環境では実行できません。"
        ) from e


class ExcelHandler:
    def __init__(self):
        # alias -> {"wb": Workbook, "path": Path, "last_records": list}
        self._workbooks: dict[str, dict[str, Any]] = {}
        self._active_alias: str | None = None

    def _active(self) -> dict[str, Any]:
        if self._active_alias is None or self._active_alias not in self._workbooks:
            raise RuntimeError("先に load_workbook を実行してください")
        return self._workbooks[self._active_alias]

    def list_open_workbooks(self) -> list[str]:
        return list(self._workbooks.keys())

    def load_workbook(self, path: str, alias: str | None = None) -> str:
        """Excelファイルを開いてメモリ上に保持し、アクティブにする。
        alias を省略した場合はファイル名(拡張子抜き)を自動的に使う。
        同じaliasで再度読み込むと、既存の内容を上書きして開き直す。
        """
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"Excelファイルが見つかりません: {p}")
        keep_vba = p.suffix.lower() in _MACRO_ENABLED_EXTENSIONS
        wb = load_workbook(filename=p, data_only=True, keep_vba=keep_vba)

        used_alias = alias or p.stem
        self._workbooks[used_alias] = {"wb": wb, "path": p, "last_records": []}
        self._active_alias = used_alias

        logger.info(
            "Excelを読み込みました: %s (alias=%s, シート: %s, keep_vba=%s)",
            p, used_alias, wb.sheetnames, keep_vba,
        )
        return f"loaded: {p} (alias={used_alias})"

    def switch_workbook(self, alias: str) -> str:
        """既に開いている複数のExcelの中から、以降の操作対象(アクティブ)を切り替える。"""
        if alias not in self._workbooks:
            raise KeyError(
                f"開いていないエイリアスです: {alias}"
                f"(現在開いているもの: {list(self._workbooks.keys())})"
            )
        self._active_alias = alias
        logger.info("アクティブなExcelを切り替えました: %s", alias)
        return f"switched to: {alias}"

    def close_workbook(self, alias: str | None = None) -> str:
        """開いているExcelをメモリ上から閉じる(ファイル自体には影響しない)。
        alias省略時はアクティブなものを閉じる。
        """
        target = alias or self._active_alias
        if target is None or target not in self._workbooks:
            raise KeyError(f"開いていないエイリアスです: {target}")
        del self._workbooks[target]
        if self._active_alias == target:
            self._active_alias = next(iter(self._workbooks), None)
        return f"closed: {target}"

    def read_sheet_to_records(self, sheet_name: str) -> list[dict[str, Any]]:
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name} (存在: {wb.sheetnames})")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            entry["last_records"] = []
            return []
        header, *body = rows
        records = [dict(zip(header, row)) for row in body]
        entry["last_records"] = records
        return records

    def write_records_to_csv(self, output_path: str) -> str:
        entry = self._active()
        records = entry["last_records"]
        if not records:
            raise RuntimeError("書き出す対象のレコードがありません(read_sheet_to_records未実行)")
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = list(records[0].keys())
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
        logger.info("CSVを出力しました: %s (%d行)", out, len(records))
        return str(out)

    def write_cells(self, sheet_name: str, cell_values: dict[str, Any]) -> str:
        """cell_values 例: {"B2": "GAO", "C2": 12345}"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        for cell_ref, value in cell_values.items():
            ws[cell_ref] = value
        return f"{len(cell_values)} セルを書き込みました"

    def get_cell_value(self, sheet_name: str, cell_ref: str) -> str:
        """特定シート・特定セル1つの値を文字列として読み込む。
        テキスト加工パイプライン(textハンドラ)に渡す入力としてもよく使う。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name} (存在: {wb.sheetnames})")
        value = wb[sheet_name][cell_ref].value
        return "" if value is None else str(value)

    def get_last_row(self, sheet_name: str, column: str | None = None) -> int:
        """最終行番号を取得する。columnを指定すると、その列で値が入っている
        最後の行番号を返す(Excelで言う Ctrl+↓ に相当)。省略時はシート全体の
        使用範囲の最終行(ws.max_row)を返す。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        if column is None:
            return ws.max_row

        col_idx = int(column) if str(column).isdigit() else column_index_from_string(column)
        last_row = 0
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=col_idx).value is not None:
                last_row = row
        return last_row

    def get_last_column(self, sheet_name: str, row: str | int | None = None) -> str:
        """最終列を列文字("A","B",...)で取得する。rowを指定すると、その行で
        値が入っている最後の列を返す(Excelで言う行方向のCtrl+→に相当)。
        省略時はシート全体の使用範囲の最終列(ws.max_column)を返す。

        戻り値は列文字のため、そのまま {{変数名}} でセル参照の一部として
        埋め込める。さらに {{変数名+1}} / {{変数名-1}} と書くと、その1つ右/左の
        列文字が得られる(表の右隣の列に続けて書き込みたい場合等に使う)。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        if row is None:
            return get_column_letter(ws.max_column)

        row_idx = int(row)
        last_col = 0
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row_idx, column=col).value is not None:
                last_col = col
        # 該当行に値が1つも無い場合は「A列より前」を表す空文字列を返す
        # (get_last_rowが値0を返すのと同じ考え方。{{last_col+1}}とすれば
        # 正しく"A"が得られる)
        return get_column_letter(last_col) if last_col else ""

    def get_sheet_names(self) -> list[str]:
        """アクティブなブックのシート名一覧を取得する。"""
        entry = self._active()
        return list(entry["wb"].sheetnames)

    def get_range_as_list(self, sheet_name: str, cell_range: str) -> list[Any]:
        """指定範囲(1列または1行)の値をリストとして取得する。
        例: "A1:A10"(1列10件) または "A1:J1"(1行10件)。単一セルの指定も可
        (要素数1のリストになる)。2次元(複数行かつ複数列)の範囲は指定できない。
        戻り値のリストは、他のRPAツールと同様Pythonのリストと同じ0始まりの
        番号でアクセスできる({{変数名[0]}} が先頭要素)。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]

        cells = ws[cell_range]
        if not isinstance(cells, tuple):
            # 単一セル
            return [cells.value]

        if cells and isinstance(cells[0], tuple):
            rows = cells
            if len(rows[0]) == 1:
                # 1列(複数行 x 1列)
                return [row[0].value for row in rows]
            if len(rows) == 1:
                # 1行(1行 x 複数列)
                return [c.value for c in rows[0]]
            raise ValueError(
                f"1列または1行の範囲のみ指定できます"
                f"(現在: {len(rows)}行 x {len(rows[0])}列): {cell_range}"
            )

        # 1行だけの範囲(cellsがセルのタプル1段のみ)
        return [c.value for c in cells]

    def find_cell(self, sheet_name: str, value: Any, column: str | None = None) -> str:
        """指定した値と一致する最初のセルのアドレス(例: "B5")を検索する。
        column を指定すると、その列だけを上から順に検索する
        (省略時はシート全体を行優先で検索する)。値は文字列化して比較する
        (Excel側の見た目の値と一致させるため)。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        target = str(value)

        if column is not None:
            col_idx = int(column) if str(column).isdigit() else column_index_from_string(column)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None and str(cell.value) == target:
                    return cell.coordinate
        else:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value) == target:
                        return cell.coordinate

        raise ValueError(f"値 '{value}' に一致するセルが見つかりませんでした(シート: {sheet_name})")

    def append_row(self, sheet_name: str, values: list[Any], start_column: str = "A") -> str:
        """シートの最終行の次の行に、values を横並びで書き込む(Excelでよく使う
        「表の末尾に1行追記する」操作)。start_column から右方向に書き込む。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        start_col_idx = column_index_from_string(start_column)
        target_row = ws.max_row + 1
        if ws.max_row == 1 and all(c.value is None for c in ws[1]):
            target_row = 1
        for offset, value in enumerate(values):
            ws.cell(row=target_row, column=start_col_idx + offset, value=value)
        return f"{target_row}行目に{len(values)}件追記しました"

    def insert_rows(self, sheet_name: str, row: int, count: int = 1) -> str:
        """row で指定した行の直前に、count行分の空行を挿入する(既存の行は下にずれる)。"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        if count < 1:
            raise ValueError("countは1以上を指定してください")
        wb[sheet_name].insert_rows(int(row), amount=int(count))
        return f"{row}行目の前に{count}行挿入しました"

    def delete_rows(self, sheet_name: str, row: int, count: int = 1) -> str:
        """row で指定した行から count行分を削除する(以降の行は上に詰まる)。"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        if count < 1:
            raise ValueError("countは1以上を指定してください")
        wb[sheet_name].delete_rows(int(row), amount=int(count))
        return f"{row}行目から{count}行削除しました"

    def clear_range(self, sheet_name: str, cell_range: str) -> str:
        """指定範囲のセルの値をすべて空にする(書式は変更しない)。"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        cells = ws[cell_range]
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        elif cells and not isinstance(cells[0], tuple):
            cells = (cells,)
        count = 0
        for row in cells:
            for cell in row:
                cell.value = None
                count += 1
        return f"{count} セルを空にしました: {cell_range}"

    def delete_sheet(self, sheet_name: str) -> str:
        """既存シートを削除する(ブック内で唯一のシートは削除できない)。"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name} (存在: {wb.sheetnames})")
        if len(wb.sheetnames) <= 1:
            raise ValueError("ブック内で唯一のシートは削除できません")
        wb.remove(wb[sheet_name])
        logger.info("シートを削除しました: %s", sheet_name)
        return f"deleted sheet: {sheet_name}"

    def copy_cell_range(
        self, sheet_name: str, source_range: str, dest_cell: str, paste_type: str = "values"
    ) -> str:
        """source_range(例: "A1:B3")を、dest_cellを左上として貼り付ける。

        paste_type:
          "values"   既定。計算済みの値のみを貼り付ける(数式は貼り付けない)。
          "formulas" 数式をそのまま貼り付ける(セル参照は自動調整されない。
                     元のファイルを読み直して数式文字列を取得するため、
                     load_workbook後にこのブック上でまだ保存していない変更が
                     元セルにある場合、その変更は数式の取得には反映されない)。
          "all"      値/数式に加えて、数値の表示形式・フォント・塗りつぶし・
                     罫線・配置もあわせて貼り付ける(列幅/行高・条件付き書式・
                     コメント・入力規則は対象外)。
        """
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]

        if paste_type not in ("values", "formulas", "all"):
            raise ValueError(
                f"paste_typeは 'values' / 'formulas' / 'all' のいずれかを指定してください: {paste_type}"
            )

        cells = ws[source_range]
        # 単一セルの場合はタプルのタプルにならないため揃える
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        elif cells and not isinstance(cells[0], tuple):
            cells = (cells,)

        dest_col_letters = "".join(c for c in dest_cell if c.isalpha())
        dest_row_digits = "".join(c for c in dest_cell if c.isdigit())
        if not dest_col_letters or not dest_row_digits:
            raise ValueError(f"貼り付け先セルの指定が不正です: {dest_cell}")
        dest_row = int(dest_row_digits)
        dest_col = column_index_from_string(dest_col_letters)

        formula_ws = None
        if paste_type in ("formulas", "all"):
            # 数式文字列は data_only=True のブックからは取得できないため、
            # 同じファイルを data_only=False で読み直して数式だけを参照する
            formula_wb = load_workbook(filename=entry["path"], data_only=False)
            if sheet_name not in formula_wb.sheetnames:
                raise KeyError(f"シートが見つかりません(数式読込用): {sheet_name}")
            formula_ws = formula_wb[sheet_name]

        count = 0
        for r_offset, row in enumerate(cells):
            for c_offset, cell in enumerate(row):
                target = ws.cell(row=dest_row + r_offset, column=dest_col + c_offset)

                if paste_type == "values":
                    target.value = cell.value
                else:
                    src_formula_cell = formula_ws.cell(row=cell.row, column=cell.column)
                    target.value = src_formula_cell.value

                if paste_type == "all":
                    target.number_format = cell.number_format
                    target.font = _copy_style(cell.font)
                    target.fill = _copy_style(cell.fill)
                    target.border = _copy_style(cell.border)
                    target.alignment = _copy_style(cell.alignment)

                count += 1
        return f"{count} セルをコピーしました({paste_type}): {source_range} -> {dest_cell}"

    def create_sheet(self, sheet_name: str, index: int | None = None) -> str:
        """新しいシートを追加する。index省略時は末尾に追加する(0始まりで
        位置を指定可能。例: 0で先頭に追加)。"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name in wb.sheetnames:
            raise ValueError(f"シート名 '{sheet_name}' は既に存在します")
        wb.create_sheet(title=sheet_name, index=index)
        logger.info("シートを追加しました: %s (index=%s)", sheet_name, index)
        return sheet_name

    def rename_sheet(self, old_name: str, new_name: str) -> str:
        """既存シートの名前を変更する。"""
        entry = self._active()
        wb = entry["wb"]
        if old_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {old_name} (存在: {wb.sheetnames})")
        if new_name in wb.sheetnames:
            raise ValueError(f"シート名 '{new_name}' は既に存在します")
        ws = wb[old_name]
        ws.title = new_name
        logger.info("シート名を変更しました: %s -> %s", old_name, new_name)
        return new_name

    def copy_sheet(self, source_sheet_name: str, new_sheet_name: str) -> str:
        """シートを丸ごとコピーする(同一ブック内)。"""
        entry = self._active()
        wb = entry["wb"]
        if source_sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {source_sheet_name} (存在: {wb.sheetnames})")
        if new_sheet_name in wb.sheetnames:
            raise ValueError(f"シート名 '{new_sheet_name}' は既に存在します")
        src = wb[source_sheet_name]
        new_ws = wb.copy_worksheet(src)
        new_ws.title = new_sheet_name
        return new_sheet_name

    def set_print_area(self, sheet_name: str, cell_range: str) -> str:
        """指定範囲(例: "A1:H30")をそのシートの印刷範囲に設定する。"""
        entry = self._active()
        wb = entry["wb"]
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: {sheet_name}")
        ws = wb[sheet_name]
        ws.print_area = cell_range
        return f"印刷範囲を設定しました: {sheet_name}!{cell_range}"

    def save_workbook_as(self, output_path: str) -> str:
        entry = self._active()
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        # .xlsm/.xltmとして保存する場合、load_workbook時にkeep_vba=Trueで
        # 読み込んでいれば、既存のVBAはそのまま保持される
        entry["wb"].save(out)
        logger.info("Excelを保存しました: %s", out)
        return str(out)

    def save_as_pdf(self, path: str, output_path: str, sheet_name: str | None = None) -> str:
        """あらかじめファイル側で設定済みの印刷範囲・ページ設定のままPDFへ
        書き出す(**Windows + Excelインストール環境専用**)。
        sheet_name を指定しない場合はブック全体(全シート)を書き出す。
        印刷範囲やページ設定(用紙サイズ・拡大縮小・余白等)は、この関数では
        一切変更しない。事前にExcel側で設定を済ませておくこと。
        アクティブブックの状態とは独立して、指定したpathに対して直接動作する。
        """
        win32com = _import_win32com()

        p = Path(path).resolve()
        if not p.exists():
            raise FileNotFoundError(f"Excelファイルが見つかりません: {p}")
        out = Path(output_path).resolve()
        out.parent.mkdir(parents=True, exist_ok=True)

        excel = win32com.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(p))
            try:
                target = wb.Sheets(sheet_name) if sheet_name else wb
                # xlTypePDF = 0
                target.ExportAsFixedFormat(0, str(out))
            finally:
                wb.Close(SaveChanges=False)
        finally:
            excel.Quit()

        logger.info("ExcelをPDFに書き出しました: %s -> %s", p, out)
        return str(out)

    def run_excel_macro(self, path: str, macro_name: str, save_after: bool = True) -> str:
        """.xlsm/.xltmに既に組み込まれているVBAマクロを実行する
        (**Windows + Excelインストール環境専用**)。openpyxlはVBAコードの
        実行を一切行えないため、実際のExcelアプリケーションを操作する。
        アクティブブックの状態とは独立して、指定したpathに対して直接動作する。
        """
        win32com = _import_win32com()

        p = Path(path).resolve()
        if not p.exists():
            raise FileNotFoundError(f"Excelファイルが見つかりません: {p}")

        excel = win32com.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(p))
            try:
                try:
                    excel.Run(macro_name)
                except Exception as e:  # noqa: BLE001
                    raise RuntimeError(f"マクロ '{macro_name}' の実行に失敗しました: {e}") from e
                if save_after:
                    wb.Save()
            finally:
                wb.Close(SaveChanges=save_after)
        finally:
            excel.Quit()

        logger.info("Excelマクロを実行しました: %s (%s)", macro_name, p)
        return f"macro executed: {macro_name}"
