"""
gui_recorder.py — 操作の登録をボタンクリック中心のGUIで行うためのツール。

コマンドライン版(main.pyの「操作を登録」)と全く同じ土台 —
MacroRecorderが内部に持つExcel/PDF/Web/エクスプローラー/実行ファイル/
デスクトップの各ハンドラ、および保存ロジック(_save_macro/_save_intent、
自動バックアップ)— をそのまま再利用し、対話部分だけをTkinterのフォーム・
ボタンに置き換えたもの。cmdでのメニュー番号入力の代わりに、ボタンを押す・
入力欄に文字を入れて送信する・画像はクリップボードから貼り付けて送信する、
ファイル/フォルダはエクスプローラーからドラッグ&ドロップする、という操作で
マクロを組み立てられる。

ドラッグ&ドロップには任意の依存 `tkinterdnd2` を使う(pip install tkinterdnd2)。
インストールされていない場合はD&Dが使えないだけで、他の機能(参照...ボタン、
クリップボード貼り付け等)には影響しない。

使い方:
    python gui_recorder.py
"""
from __future__ import annotations

import sys
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter

_NO_VALUE = object()  # register_stepでvalue未指定を表す番人値(Noneも正当な値のため)

BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

from engine.recorder import MacroRecorder  # noqa: E402
from handlers.browser_handler import (  # noqa: E402
    BrowserHandler,
    ElementNotFoundError,
    SiteNotWhitelistedError,
)
from handlers.desktop_handler import ImageNotFoundError  # noqa: E402
from handlers.explorer_handler import PathConflictError  # noqa: E402
from handlers.process_handler import ScriptNotWhitelistedError  # noqa: E402

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False

CONFIG_DIR = BASE_DIR / "config"
CLIP_IMAGE_DIR = BASE_DIR / "workdir" / "clip_images"

DOMAIN_LABELS = {
    "excel": "Excel",
    "pdf": "PDF",
    "web": "Webサイト",
    "explorer": "エクスプローラー",
    "process": "実行ファイル(exe/py)",
    "desktop": "デスクトップ",
    "text": "テキスト加工",
    "list": "リスト",
    "control": "制御構文",
}

DOMAIN_ACTIONS = {
    "excel": [
        "ファイルを開く", "シートを読み込む", "CSVへ書き出す", "セルに書き込む",
        "別名で保存する", "PDFとして保存する", "VBAマクロを実行する",
        "開いている別のExcelに切り替える", "セル1つの値を読み込む", "最終行を取得する",
        "セルをコピーして貼り付ける", "シートをコピーする", "シート名の一覧を取得する",
        "印刷範囲を指定する", "新規シートを追加する", "シート名を変更する",
        "セル範囲(1列/1行)をリストとして取得する",
        "値でセルを検索する", "表の末尾に1行追記する", "行を挿入する", "行を削除する",
        "セル範囲の値を空にする", "シートを削除する", "最終列を取得する",
    ],
    "pdf": [
        "テキストを抽出する(ページ全体)", "複数PDFを結合する", "PDFを分割する",
        "PDFを回転する", "OCRでテキスト化する(ページ全体)",
        "範囲を指定してテキストを取得する", "ページ範囲を1ファイルに抜き出す",
        "ページ数を取得する", "表(罫線あり)をCSVとして抽出する", "埋め込み画像を抽出する",
    ],
    "web": [
        "サイトを開く/切り替える", "クリックする", "入力する", "選択する",
        "待機する", "画面をPDF保存する",
        "画面から文字を読み取る", "画面から属性値を読み取る", "画面から文字のリストを読み取る",
        "チェックボックスをON/OFFする",
        "ウィンドウサイズを指定する", "ウィンドウ位置を指定する", "表示倍率(ズーム)を指定する",
    ],
    "explorer": [
        "パスを開く", "フォルダを作成する", "ファイルを移動する", "ファイルをコピーする",
        "フォルダを移動する", "フォルダをコピーする", "ファイル名を変更する", "フォルダ名を変更する",
        "パスが存在するか確認する", "フォルダ内のファイル一覧を取得する",
        "ファイルを削除する", "フォルダを削除する",
    ],
    "process": ["exe/pyを実行する"],
    "desktop": [
        "画像を探してクリックする", "画像を探してマウス移動する", "座標をクリックする",
        "文字列を入力する", "特殊キーを送信する", "スクリーンショットを撮る",
        "開いているウィンドウのタイトル一覧を見る", "ウィンドウをアクティブにする",
        "ウィンドウサイズを指定する(タイトル指定)", "ウィンドウ位置を指定する(タイトル指定)",
    ],
    "text": [
        "文字を探して切り出す", "文字を置換する", "日付・時刻を取得する",
        "文字をつなげる/付加する", "クリップボードにコピーする", "クリップボードから取得する",
    ],
    "list": [
        "空のリストを作成する", "リストに値を追加する",
        "リストの要素を取得する", "リストの長さを取得する",
    ],
    "control": [
        "ラベルを置く", "指定したラベルへジャンプする(goto)",
        "条件を満たしたらジャンプする(IF文)", "変数に値を設定する",
        "繰り返しを開始する(for)", "繰り返しを終了する(next)",
        "変数の型を変換する(文字列/整数/小数)",
    ],
}


def parse_dnd_paths(data: str) -> list[str]:
    """tkinterdnd2のevent.dataをパースしてパスのリストにする。
    スペースを含むパスは "{C:/path with spaces/file.xlsx}" のように波括弧で
    囲まれて渡ってくる(tkdndの標準仕様)。複数ファイルはスペース区切り。
    """
    paths: list[str] = []
    current = ""
    in_brace = False
    for ch in data:
        if ch == "{":
            in_brace = True
            current = ""
        elif ch == "}":
            in_brace = False
            paths.append(current)
            current = ""
        elif ch == " " and not in_brace:
            if current:
                paths.append(current)
                current = ""
        else:
            current += ch
    if current:
        paths.append(current)
    return paths


def enable_path_drop(widget, on_path) -> None:
    """widget(Entry/Listbox等)にファイル/フォルダのドラッグ&ドロップを
    有効にする。tkinterdnd2が無い環境では何もしない(他の入力手段は使える)。
    ドロップされた最初のパスを on_path(path: str) に渡す。
    """
    if not DND_AVAILABLE:
        return
    widget.drop_target_register(DND_FILES)

    def _on_drop(event):
        paths = parse_dnd_paths(event.data)
        if paths:
            on_path(paths[0])

    widget.dnd_bind("<<Drop>>", _on_drop)


def _offset_template(var_name: str, offset: int) -> str:
    """{{var_name}} に対するオフセット付きテンプレート文字列を組み立てる。
    offset=0なら "{{var_name}}"、それ以外は "{{var_name+N}}"/"{{var_name-N}}"。
    """
    if offset == 0:
        return "{{" + var_name + "}}"
    sign = "+" if offset > 0 else "-"
    return "{{" + var_name + sign + str(abs(offset)) + "}}"


def grab_clipboard_image(save_dir: Path) -> str | None:
    """クリップボードの画像を取得し、PNGとして保存してパスを返す。
    画像がクリップボードに無い/取得できない場合はNoneを返し、詳しい理由は例外で伝える。
    """
    from PIL import Image, ImageGrab

    try:
        data = ImageGrab.grabclipboard()
    except NotImplementedError as e:
        raise RuntimeError(
            "このOSではクリップボード画像の取得に対応していません"
            "(Linuxではxclipまたはwl-pasteのインストールが必要です): " + str(e)
        ) from e

    if data is None:
        raise RuntimeError("クリップボードに画像が見つかりませんでした。先に画面を範囲コピーしてください。")

    if isinstance(data, list):
        # 一部の環境では画像そのものではなく「コピーされたファイルパス一覧」が返る
        if len(data) == 1 and Path(data[0]).exists():
            try:
                Image.open(data[0]).verify()
            except Exception as e:  # noqa: BLE001
                raise RuntimeError(f"クリップボードのファイルを画像として開けませんでした: {e}") from e
            return str(data[0])
        raise RuntimeError("クリップボードの内容を単一の画像として扱えませんでした。")

    save_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    out_path = save_dir / f"clip_{ts}.png"
    data.save(out_path)
    return str(out_path)


class ValueSlotField(ttk.Frame):
    """値の入力欄 + 『スロット名(空欄なら固定値)』の組み合わせウィジェット。"""

    def __init__(self, parent, label: str, width: int = 40):
        super().__init__(parent)
        ttk.Label(self, text=label, width=22, anchor="w").grid(row=0, column=0, sticky="w")
        self.value_var = tk.StringVar()
        self.entry = ttk.Entry(self, textvariable=self.value_var, width=width)
        self.entry.grid(row=0, column=1, padx=4, sticky="we")
        ttk.Label(self, text="スロット名(空欄で固定値):").grid(row=0, column=2, padx=(10, 2))
        self.slot_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.slot_var, width=14).grid(row=0, column=3)
        self._extra_col = 4
        self.columnconfigure(1, weight=1)

    def add_button(self, text: str, command) -> ttk.Button:
        btn = ttk.Button(self, text=text, command=command)
        btn.grid(row=0, column=self._extra_col, padx=3)
        self._extra_col += 1
        if text == "参照...":
            # 参照ボタンが付くフィールド=パスを入力する欄、という判断で
            # 併せてドラッグ&ドロップも有効にする
            self.enable_path_drop()
        return btn

    def enable_path_drop(self) -> None:
        enable_path_drop(self.entry, self.value_var.set)

    def browse_file(self) -> None:
        path = filedialog.askopenfilename()
        if path:
            self.value_var.set(path)

    def browse_dir(self) -> None:
        path = filedialog.askdirectory()
        if path:
            self.value_var.set(path)

    def browse_save_file(self, defaultext: str = "") -> None:
        path = filedialog.asksaveasfilename(defaultextension=defaultext)
        if path:
            self.value_var.set(path)

    def get(self) -> tuple[str, str, str | None]:
        """戻り値: (動作確認に使う実際の値, macros.jsonに書くparam値, スロット名 or None)"""
        value = self.value_var.get()
        slot = self.slot_var.get().strip()
        if slot:
            return value, "{{" + slot + "}}", slot
        return value, value, None


class PlainField(ttk.Frame):
    """スロット化しない単純な入力欄(数値・オプション文字列等)。
    is_path=True にすると、参照ボタンとドラッグ&ドロップの両方を有効にする。
    """

    def __init__(self, parent, label: str, width: int = 20, default: str = "", is_path: bool = False):
        super().__init__(parent)
        ttk.Label(self, text=label, width=30, anchor="w").grid(row=0, column=0, sticky="w")
        self.var = tk.StringVar(value=default)
        self.entry = ttk.Entry(self, textvariable=self.var, width=width)
        self.entry.grid(row=0, column=1, sticky="w")
        if is_path:
            ttk.Button(self, text="参照...", command=self.browse_file).grid(row=0, column=2, padx=3)
            enable_path_drop(self.entry, self.var.set)

    def browse_file(self) -> None:
        path = filedialog.askopenfilename()
        if path:
            self.var.set(path)

    def get(self) -> str:
        return self.var.get()


class BoolField(ttk.Frame):
    def __init__(self, parent, label: str, default: bool = False):
        super().__init__(parent)
        self.var = tk.BooleanVar(value=default)
        ttk.Checkbutton(self, text=label, variable=self.var).grid(row=0, column=0, sticky="w")

    def get(self) -> bool:
        return self.var.get()


class ImagePasteField(ttk.Frame):
    """画像パス欄 + 参照 + クリップボードから貼り付け + ドラッグ&ドロップ + プレビュー表示。"""

    def __init__(self, parent, label: str):
        super().__init__(parent)
        ttk.Label(self, text=label, width=22, anchor="w").grid(row=0, column=0, sticky="w")
        self.value_var = tk.StringVar()
        self.entry = ttk.Entry(self, textvariable=self.value_var, width=40)
        self.entry.grid(row=0, column=1, padx=4)
        ttk.Button(self, text="参照...", command=self._browse).grid(row=0, column=2, padx=3)
        ttk.Button(self, text="クリップボードから貼り付け", command=self._paste).grid(row=0, column=3, padx=3)
        self.preview_label = ttk.Label(self, text="(画像プレビュー。ここへファイルをドラッグ&ドロップも可)")
        self.preview_label.grid(row=1, column=0, columnspan=4, pady=(4, 0), sticky="w")
        self._preview_img = None  # 参照保持(GC対策)

        def _on_drop_path(path: str) -> None:
            self.value_var.set(path)
            self._update_preview(path)

        enable_path_drop(self.entry, _on_drop_path)
        enable_path_drop(self.preview_label, _on_drop_path)

    def _browse(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("画像ファイル", "*.png *.jpg *.jpeg *.bmp"), ("すべてのファイル", "*.*")]
        )
        if path:
            self.value_var.set(path)
            self._update_preview(path)

    def _paste(self) -> None:
        try:
            path = grab_clipboard_image(CLIP_IMAGE_DIR)
        except RuntimeError as e:
            messagebox.showwarning("貼り付けできません", str(e))
            return
        self.value_var.set(path)
        self._update_preview(path)

    def _update_preview(self, path: str) -> None:
        try:
            from PIL import Image, ImageTk

            img = Image.open(path)
            img.thumbnail((160, 120))
            self._preview_img = ImageTk.PhotoImage(img)
            self.preview_label.configure(image=self._preview_img, text="")
        except Exception:  # noqa: BLE001
            self.preview_label.configure(text="(プレビュー表示できませんでした)")

    def get(self) -> str:
        return self.value_var.get()


class PairsField(ttk.Frame):
    """『セル参照: 値』のような対応表を追加していく入力ウィジェット(Excelのセル書込用)。"""

    def __init__(self, parent, key_label: str, value_label: str):
        super().__init__(parent)
        self.key_label = key_label
        self.value_label = value_label
        self.rows: list[tuple[tk.StringVar, tk.StringVar]] = []

        self.rows_frame = ttk.Frame(self)
        self.rows_frame.pack(fill="x")
        ttk.Button(self, text="+ 行を追加", command=self.add_row).pack(anchor="w", pady=(2, 0))
        self.add_row()
        self.add_row()

    def add_row(self) -> None:
        row_idx = len(self.rows)
        k_var, v_var = tk.StringVar(), tk.StringVar()
        row = ttk.Frame(self.rows_frame)
        row.pack(fill="x", pady=1)
        ttk.Label(row, text=self.key_label, width=14).pack(side="left")
        ttk.Entry(row, textvariable=k_var, width=10).pack(side="left", padx=3)
        ttk.Label(row, text=self.value_label, width=10).pack(side="left")
        ttk.Entry(row, textvariable=v_var, width=24).pack(side="left", padx=3)
        self.rows.append((k_var, v_var))

    def get(self) -> dict[str, str]:
        result = {}
        for k_var, v_var in self.rows:
            k = k_var.get().strip()
            if k:
                result[k] = v_var.get()
        return result


class PathListField(ttk.Frame):
    """複数ファイルパスを追加していくウィジェット(PDF結合用)。
    ドラッグ&ドロップで複数ファイルを一度に追加することもできる。
    """

    def __init__(self, parent, label: str):
        super().__init__(parent)
        ttk.Label(self, text=label + "(ここへドラッグ&ドロップも可)").pack(anchor="w")
        self.listbox = tk.Listbox(self, height=4, width=60)
        self.listbox.pack(side="left", fill="x", expand=True)
        btns = ttk.Frame(self)
        btns.pack(side="left", padx=4)
        ttk.Button(btns, text="追加...", command=self._add).pack(fill="x")
        ttk.Button(btns, text="削除", command=self._remove).pack(fill="x", pady=(2, 0))

        if DND_AVAILABLE:
            self.listbox.drop_target_register(DND_FILES)

            def _on_drop(event):
                for p in parse_dnd_paths(event.data):
                    self.listbox.insert("end", p)

            self.listbox.dnd_bind("<<Drop>>", _on_drop)

    def _add(self) -> None:
        paths = filedialog.askopenfilenames(filetypes=[("PDFファイル", "*.pdf")])
        for p in paths:
            self.listbox.insert("end", p)

    def _remove(self) -> None:
        sel = list(self.listbox.curselection())
        for idx in reversed(sel):
            self.listbox.delete(idx)

    def get(self) -> list[str]:
        return list(self.listbox.get(0, "end"))


_AppBase = TkinterDnD.Tk if DND_AVAILABLE else tk.Tk


class RecorderApp(_AppBase):
    def __init__(self, browser: str = "chrome"):
        super().__init__()
        self.title("疑似ローカルAI — 操作の登録 (GUI)")
        self.geometry("1140x760")

        self.recorder = MacroRecorder(CONFIG_DIR, browser=browser)
        self.base_step_count = 0

        self._build_layout()
        self._on_domain_changed()

        if not DND_AVAILABLE:
            self.log(
                "ヒント: pip install tkinterdnd2 を行うと、Excel/PDF/フォルダ/exe等の"
                "パス欄にファイルをドラッグ&ドロップできるようになります。"
            )

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---------- レイアウト ----------

    def _build_layout(self) -> None:
        top = ttk.Frame(self, padding=8)
        top.pack(fill="x")

        ttk.Label(top, text="領域:").grid(row=0, column=0, sticky="w")
        self.domain_var = tk.StringVar(value="excel")
        domain_combo = ttk.Combobox(
            top, textvariable=self.domain_var, state="readonly",
            values=list(DOMAIN_LABELS.keys()), width=14,
        )
        domain_combo["values"] = list(DOMAIN_LABELS.keys())
        # 表示はラベル、内部値はキーにするため専用マッピングを使う
        self.domain_combo = ttk.Combobox(
            top, state="readonly", width=20,
            values=[DOMAIN_LABELS[k] for k in DOMAIN_LABELS],
        )
        self.domain_combo.current(0)
        self.domain_combo.grid(row=0, column=1, padx=(4, 16))
        self.domain_combo.bind("<<ComboboxSelected>>", lambda e: self._on_domain_changed())

        ttk.Label(top, text="操作:").grid(row=0, column=2, sticky="w")
        self.action_combo = ttk.Combobox(top, state="readonly", width=28)
        self.action_combo.grid(row=0, column=3, padx=4)
        self.action_combo.bind("<<ComboboxSelected>>", lambda e: self._on_action_changed())

        ttk.Label(top, text="ブラウザ:").grid(row=0, column=4, sticky="w", padx=(16, 0))
        browser_labels = {"chrome": "Chrome", "edge": "Edge"}
        self._browser_label_to_key = {v: k for k, v in browser_labels.items()}
        self.browser_combo = ttk.Combobox(
            top, state="readonly", width=10, values=list(browser_labels.values()),
        )
        self.browser_combo.set(browser_labels.get(self.recorder.browser.browser, "Chrome"))
        self.browser_combo.grid(row=0, column=5, padx=4)
        self.browser_combo.bind("<<ComboboxSelected>>", lambda e: self._on_browser_changed())

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True)

        left = ttk.Frame(body)
        left.pack(side="left", fill="both", expand=True, padx=(8, 4), pady=4)

        right = ttk.Frame(body, width=260)
        right.pack(side="right", fill="y", padx=(4, 8), pady=4)
        right.pack_propagate(False)

        self.form_frame = ttk.LabelFrame(left, text="入力", padding=10)
        self.form_frame.pack(fill="x")

        log_frame = ttk.LabelFrame(left, text="ログ(実行結果・エラー)", padding=4)
        log_frame.pack(fill="both", expand=False, pady=4)
        self.log_text = scrolledtext.ScrolledText(log_frame, height=7, state="disabled")
        self.log_text.pack(fill="both", expand=True)

        steps_frame = ttk.LabelFrame(left, text="記録済みの手順", padding=4)
        steps_frame.pack(fill="both", expand=True, pady=4)
        self.steps_listbox = tk.Listbox(steps_frame)
        self.steps_listbox.pack(fill="both", expand=True)

        bottom = ttk.Frame(left)
        bottom.pack(fill="x", pady=(4, 0))
        ttk.Button(bottom, text="元に戻す(直前の操作を取り消す)", command=self._undo).pack(side="left")
        ttk.Button(bottom, text="保存して終了", command=self._finish).pack(side="right")
        ttk.Button(bottom, text="中止(保存しない)", command=self._cancel_all).pack(side="right", padx=6)

        var_frame = ttk.LabelFrame(right, text="変数一覧(記録時点の値)", padding=4)
        var_frame.pack(fill="both", expand=True)
        ttk.Label(
            var_frame,
            text="store_asで保存した変数・リストです。実行のたびに\n値が変わる場合があります(Excelの最終行等)。",
            foreground="#557", justify="left", wraplength=230,
        ).pack(anchor="w", pady=(0, 4))
        columns = ("name", "value")
        self.variables_tree = ttk.Treeview(var_frame, columns=columns, show="headings", height=14)
        self.variables_tree.heading("name", text="変数名(型)")
        self.variables_tree.heading("value", text="値")
        self.variables_tree.column("name", width=90, anchor="w")
        self.variables_tree.column("value", width=140, anchor="w")
        self.variables_tree.pack(fill="both", expand=True)

    def log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def refresh_steps(self) -> None:
        self.steps_listbox.delete(0, "end")
        for i, step in enumerate(self.recorder.steps, start=1):
            self.steps_listbox.insert(
                "end", f"{i}. {step['handler']}.{step['action']}  {step.get('params', {})}"
            )

    def register_step(self, step: dict, value: Any = _NO_VALUE) -> None:
        """手順を登録する。value を渡すと(store_asが設定されている場合)、
        記録時点で確認できた実値として変数一覧パネルにも反映する。
        """
        self.recorder.steps.append(step)
        store_as = step.get("store_as")
        if store_as and value is not _NO_VALUE:
            self.recorder.record_variable(store_as, value)
            self.refresh_variables()
        self.refresh_steps()
        self.log(f"✅ 登録しました: {step['handler']}.{step['action']}")

    def refresh_variables(self) -> None:
        self.variables_tree.delete(*self.variables_tree.get_children())
        for name, value in self.recorder.variables.items():
            preview = repr(value)
            if len(preview) > 200:
                preview = preview[:200] + "..."
            self.variables_tree.insert("", "end", values=(f"{name} ({type(value).__name__})", preview))

    # ---------- 領域/操作の切り替え ----------

    def _on_domain_changed(self) -> None:
        label = self.domain_combo.get()
        domain_key = next(k for k, v in DOMAIN_LABELS.items() if v == label)
        self.current_domain = domain_key
        self.action_combo["values"] = DOMAIN_ACTIONS[domain_key]
        self.action_combo.current(0)
        self._on_action_changed()

    def _on_action_changed(self) -> None:
        for child in self.form_frame.winfo_children():
            child.destroy()
        domain = self.current_domain
        action_label = self.action_combo.get()
        builder = getattr(self, f"_build_{domain}", None)
        if builder is None:
            ttk.Label(self.form_frame, text="(未対応)").pack()
            return
        builder(action_label)

    def _on_browser_changed(self) -> None:
        label = self.browser_combo.get()
        new_browser = self._browser_label_to_key.get(label, "chrome")
        if new_browser == self.recorder.browser.browser:
            return

        # 既にブラウザが起動済み(サイトを開いている等)の場合、切り替えると
        # 今開いている画面は失われるため、先に確認する。
        if self.recorder.browser._driver is not None:
            if not self._confirm(
                f"ブラウザを{label}に切り替えます。今開いているブラウザ画面は閉じられます。"
                "よろしいですか?"
            ):
                # 取り消し: プルダウンの表示を元のブラウザに戻す
                current_label = {"chrome": "Chrome", "edge": "Edge"}.get(
                    self.recorder.browser.browser, "Chrome"
                )
                self.browser_combo.set(current_label)
                return
            try:
                self.recorder.browser.close()
            except Exception:  # noqa: BLE001
                pass

        self.recorder.browser = BrowserHandler(
            self.recorder.config_dir / "whitelist_urls.json",
            headless=False,
            browser=new_browser,
        )
        self.recorder._site_opened = False
        self.log(f"→ ブラウザを{label}に切り替えました")

    # ---------- 共通ヘルパー ----------

    def _submit_button(self, text: str = "動作確認して登録") -> ttk.Button:
        return ttk.Button(self.form_frame, text=text)

    def _ask_verify(self, image_or_text_default: str | None = None, is_image: bool = False) -> dict:
        """簡易な確認方法ダイアログ。「なし」「表示文字/画像が出る」「消える」「URL変化」を選ばせる。"""
        win = tk.Toplevel(self)
        win.title("実行後の確認方法")
        win.grab_set()
        result: dict = {"type": "none"}

        ttk.Label(win, text="この操作が成功したことをどう確認しますか?", padding=8).pack()
        choice_var = tk.StringVar(value="none")
        options = [("確認しない", "none")]
        if is_image:
            options += [("対象画像が消える", "image_disappears"), ("別の画像が新しく表示される", "image_appears")]
        else:
            options += [("特定の文字が新しく表示される", "text_appears"), ("URLが変わる", "url_changes")]
        for text, val in options:
            ttk.Radiobutton(win, text=text, variable=choice_var, value=val).pack(anchor="w", padx=16)

        detail_var = tk.StringVar()
        detail_entry = ttk.Entry(win, textvariable=detail_var, width=40)
        ttk.Label(win, text="(表示される文字 / 画像パスを指定する場合はここに)").pack(pady=(8, 0))
        detail_entry.pack()

        def on_ok():
            vtype = choice_var.get()
            if vtype == "none":
                result["type"] = "none"
            elif vtype in ("text_appears",):
                result.update({"type": vtype, "value": detail_var.get(), "timeout": 10})
            elif vtype == "url_changes":
                result.update({"type": vtype, "timeout": 10})
            elif vtype == "image_disappears":
                val = detail_var.get() or (image_or_text_default or "")
                result.update({"type": vtype, "value": val, "timeout": 10})
            elif vtype == "image_appears":
                result.update({"type": vtype, "value": detail_var.get(), "timeout": 10})
            win.destroy()

        ttk.Button(win, text="OK", command=on_ok).pack(pady=10)
        self.wait_window(win)
        return result

    def _ask_retry(self) -> dict:
        raw = simpledialog.askstring(
            "自動リトライ",
            "失敗時の自動リトライ回数(不要なら0のままでOK):",
            initialvalue="0",
            parent=self,
        )
        try:
            count = max(int(raw), 0) if raw else 0
        except ValueError:
            count = 0
        return {"count": count, "interval_seconds": 2}

    def _confirm(self, message: str) -> bool:
        return messagebox.askyesno("確認", message)

    def _ask_store_as(self) -> str | None:
        """この手順の結果を後で使うための変数名を、ダイアログで聞く。
        空欄なら使わない。使う場合は {{変数名}} の形で他のテキスト欄に入力すれば参照できる。
        """
        name = simpledialog.askstring(
            "変数として保存",
            "この結果に名前を付けて後の手順で使いますか?\n"
            "(空欄なら使わない。使う場合の書き方: {{変数名}})",
            parent=self,
        )
        name = (name or "").strip()
        return name or None

    @staticmethod
    def _has_template(value: str) -> bool:
        """値の中に {{変数名}} が含まれているか(前の手順の結果を埋め込む
        テンプレートかどうか)を判定する。含まれる場合、実行時にしか
        値が定まらないため、その場での動作確認はできない。
        """
        return "{{" in value and "}}" in value

    # ---------- Excel ----------

    def _build_excel(self, action: str) -> None:
        f = self.form_frame
        if action == "ファイルを開く":
            field = ValueSlotField(f, "開くExcelファイルのパス")
            field.add_button("参照...", field.browse_file)
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                try:
                    self.recorder.excel.load_workbook(test_v)
                    self.log(f"→ 開けました: {test_v}")
                    self.register_step({"handler": "excel", "action": "load_workbook",
                                         "params": {"path": param_v}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "シートを読み込む":
            field = ValueSlotField(f, "読み込むシート名")
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                try:
                    records = self.recorder.excel.read_sheet_to_records(test_v)
                    self.log(f"→ {len(records)}件読み込めました")
                    self.register_step({"handler": "excel", "action": "read_sheet_to_records",
                                         "params": {"sheet_name": param_v}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "CSVへ書き出す":
            field = ValueSlotField(f, "出力するCSVファイルのパス(例: report_{{last_row}}.csv も可)")
            field.add_button("参照...", lambda: field.browse_save_file(".csv"))
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                step = {"handler": "excel", "action": "write_records_to_csv", "params": {"output_path": param_v}}
                if self._has_template(test_v):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.excel.write_records_to_csv(test_v)
                    self.log(f"→ 書き出せました: {test_v}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "セルに書き込む":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)

            ttk.Label(f, text="セル参照の指定方法:").pack(anchor="w", pady=(6, 0))
            basis_var = tk.StringVar(value="direct")
            basis_frame = ttk.Frame(f)
            basis_frame.pack(anchor="w", pady=2)
            ttk.Radiobutton(basis_frame, text="直接指定", variable=basis_var, value="direct").pack(side="left")
            ttk.Radiobutton(
                basis_frame, text="最終行の続き(列だけ指定)", variable=basis_var, value="last_row"
            ).pack(side="left")
            ttk.Radiobutton(
                basis_frame, text="最終列の続き(行だけ指定)", variable=basis_var, value="last_col"
            ).pack(side="left")

            sub_frame = ttk.Frame(f)
            sub_frame.pack(fill="x", pady=4)
            sub_widgets: dict = {}

            def rebuild_sub(*_args):
                for child in sub_frame.winfo_children():
                    child.destroy()
                basis = basis_var.get()
                if basis == "last_row":
                    sub_widgets["scope"] = PlainField(sub_frame, "最終行を判定する基準列(空欄でシート全体)")
                    sub_widgets["scope"].pack(fill="x", pady=2)
                    sub_widgets["offset"] = PlainField(sub_frame, "最終行から何行後に書き込むか(空欄で1)")
                    sub_widgets["offset"].pack(fill="x", pady=2)
                    sub_widgets["pairs"] = PairsField(sub_frame, "列(例:B)", "値")
                    sub_widgets["pairs"].pack(fill="x")
                elif basis == "last_col":
                    sub_widgets["scope"] = PlainField(sub_frame, "最終列を判定する基準行(行番号)")
                    sub_widgets["scope"].pack(fill="x", pady=2)
                    sub_widgets["offset"] = PlainField(sub_frame, "最終列から何列後に書き込むか(空欄で1)")
                    sub_widgets["offset"].pack(fill="x", pady=2)
                    sub_widgets["pairs"] = PairsField(sub_frame, "行(例:5)", "値")
                    sub_widgets["pairs"].pack(fill="x")
                else:
                    sub_widgets["pairs"] = PairsField(sub_frame, "セル(例:B2)", "値")
                    sub_widgets["pairs"].pack(fill="x")

            basis_var.trace_add("write", rebuild_sub)
            rebuild_sub()

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                basis = basis_var.get()
                raw_pairs = sub_widgets["pairs"].get()
                if not raw_pairs:
                    self.log("⚠ セルの値が1つも入力されていません")
                    return

                prereq_step = None
                prereq_value = None
                if basis == "last_row":
                    scope = sub_widgets["scope"].get().strip()
                    try:
                        offset = int(sub_widgets["offset"].get().strip() or "1")
                    except ValueError:
                        offset = 1
                    try:
                        current_last_row = self.recorder.excel.get_last_row(sheet_test, column=scope or None)
                    except Exception as e:  # noqa: BLE001
                        self.log(f"⚠ {e}")
                        return
                    var_name = self.recorder._next_auto_var("last_row")
                    prereq_step = {
                        "handler": "excel", "action": "get_last_row",
                        "params": {"sheet_name": sheet_param, "column": scope or None},
                        "store_as": var_name,
                    }
                    prereq_value = current_last_row
                    row_test = current_last_row + offset
                    row_part = _offset_template(var_name, offset)
                    cell_values_test = {f"{col}{row_test}": v for col, v in raw_pairs.items()}
                    cell_values_param = {f"{col}{row_part}": v for col, v in raw_pairs.items()}
                    self.log(f"→ 今の時点の最終行は{current_last_row}行目なので、動作確認では{row_test}行目に書き込みます")
                elif basis == "last_col":
                    scope = sub_widgets["scope"].get().strip()
                    if not scope.isdigit():
                        self.log("⚠ 基準行は数字で入力してください")
                        return
                    try:
                        offset = int(sub_widgets["offset"].get().strip() or "1")
                    except ValueError:
                        offset = 1
                    try:
                        current_last_col = self.recorder.excel.get_last_column(sheet_test, row=scope)
                    except Exception as e:  # noqa: BLE001
                        self.log(f"⚠ {e}")
                        return
                    col_idx_base = column_index_from_string(current_last_col) if current_last_col else 0
                    if col_idx_base + offset < 1:
                        self.log("⚠ 指定したオフセットでは列がA列より前になります")
                        return
                    var_name = self.recorder._next_auto_var("last_col")
                    prereq_step = {
                        "handler": "excel", "action": "get_last_column",
                        "params": {"sheet_name": sheet_param, "row": int(scope)},
                        "store_as": var_name,
                    }
                    prereq_value = current_last_col
                    col_test = get_column_letter(col_idx_base + offset)
                    col_part = _offset_template(var_name, offset)
                    cell_values_test = {f"{col_test}{row}": v for row, v in raw_pairs.items()}
                    cell_values_param = {f"{col_part}{row}": v for row, v in raw_pairs.items()}
                    self.log(
                        f"→ 今の時点の最終列は{current_last_col or '(無し)'}なので、"
                        f"動作確認では{col_test}列に書き込みます"
                    )
                else:
                    cell_values_test = raw_pairs
                    cell_values_param = raw_pairs

                try:
                    self.recorder.excel.write_cells(sheet_test, cell_values_test)
                    self.log("→ セルへの書き込みを確認できました")
                    if prereq_step:
                        self.register_step(prereq_step, prereq_value)
                    self.register_step({
                        "handler": "excel", "action": "write_cells",
                        "params": {"sheet_name": sheet_param, "cell_values": cell_values_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "別名で保存する":
            field = ValueSlotField(f, "保存先のパス(例: backup_{{last_row}}.xlsx も可)")
            field.add_button("参照...", lambda: field.browse_save_file(".xlsx"))
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                step = {"handler": "excel", "action": "save_workbook_as", "params": {"output_path": param_v}}
                if self._has_template(test_v):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.excel.save_workbook_as(test_v)
                    self.log(f"→ 保存できました: {test_v}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "PDFとして保存する":
            ttk.Label(f, text="※ Windows + Excel + pywin32 が必要です。印刷範囲・ページ設定は変更しません。",
                      foreground="#a55").pack(anchor="w", pady=(0, 4))
            path_field = ValueSlotField(f, "元のExcelファイルのパス")
            path_field.add_button("参照...", path_field.browse_file)
            path_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力PDFファイルのパス")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".pdf"))
            out_field.pack(fill="x", pady=4)
            sheet_field = PlainField(f, "シート名(空欄でブック全体)")
            sheet_field.pack(fill="x", pady=4)

            def on_submit():
                path_test, path_param, _ = path_field.get()
                out_test, out_param, _ = out_field.get()
                sheet_name = sheet_field.get() or None
                try:
                    self.recorder.excel.save_as_pdf(path_test, out_test, sheet_name=sheet_name)
                    self.log(f"→ PDFへ書き出せました: {out_test}")
                    self.register_step({
                        "handler": "excel", "action": "save_as_pdf",
                        "params": {"path": path_param, "output_path": out_param, "sheet_name": sheet_name},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    if self._confirm("未確認のままこの手順を登録しますか?"):
                        self.register_step({
                            "handler": "excel", "action": "save_as_pdf",
                            "params": {"path": path_param, "output_path": out_param, "sheet_name": sheet_name},
                        })

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "VBAマクロを実行する":
            ttk.Label(f, text="※ Windows + Excel + pywin32 が必要です。", foreground="#a55").pack(
                anchor="w", pady=(0, 4)
            )
            path_field = ValueSlotField(f, ".xlsmファイルのパス")
            path_field.add_button("参照...", path_field.browse_file)
            path_field.pack(fill="x", pady=4)
            macro_field = ValueSlotField(f, "マクロ名(Sub名)")
            macro_field.pack(fill="x", pady=4)

            def on_submit():
                path_test, path_param, _ = path_field.get()
                macro_test, macro_param, _ = macro_field.get()
                if not self._confirm(f"実際に '{macro_test}' を実行して動作確認します。よろしいですか?"):
                    return
                try:
                    self.recorder.excel.run_excel_macro(path_test, macro_test)
                    self.log(f"→ マクロ '{macro_test}' を実行できました")
                    self.register_step({
                        "handler": "excel", "action": "run_excel_macro",
                        "params": {"path": path_param, "macro_name": macro_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    if self._confirm("未確認のままこの手順を登録しますか?"):
                        self.register_step({
                            "handler": "excel", "action": "run_excel_macro",
                            "params": {"path": path_param, "macro_name": macro_param},
                        })

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "開いている別のExcelに切り替える":
            aliases = self.recorder.excel.list_open_workbooks()
            if len(aliases) < 2:
                ttk.Label(f, text="切り替え可能なExcelがまだ2つ以上開かれていません。").pack(pady=10)
                return
            ttk.Label(f, text="切り替え先のExcel:").pack(anchor="w")
            combo = ttk.Combobox(f, state="readonly", values=aliases)
            combo.current(0)
            combo.pack(fill="x", pady=4)

            def on_submit():
                alias = combo.get()
                try:
                    self.recorder.excel.switch_workbook(alias)
                    self.log(f"→ '{alias}' に切り替えました")
                    self.register_step({"handler": "excel", "action": "switch_workbook",
                                         "params": {"alias": alias}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="切り替えて登録", command=on_submit).pack(pady=6)

        elif action == "セル1つの値を読み込む":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            cell_field = ValueSlotField(f, "セル参照(例: B2)")
            cell_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                cell_test, cell_param, _ = cell_field.get()
                try:
                    value = self.recorder.excel.get_cell_value(sheet_test, cell_test)
                    self.log(f"→ 読み込めました: {value!r}")
                    store_as = self._ask_store_as()
                    step = {"handler": "excel", "action": "get_cell_value",
                            "params": {"sheet_name": sheet_param, "cell_ref": cell_param}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, value)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "最終行を取得する":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            col_field = PlainField(f, "対象列(例: A。空欄でシート全体)")
            col_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                column = col_field.get().strip() or None
                try:
                    last_row = self.recorder.excel.get_last_row(sheet_test, column=column)
                    self.log(f"→ 取得できました: {last_row}行目")
                    store_as = self._ask_store_as()
                    step = {"handler": "excel", "action": "get_last_row",
                            "params": {"sheet_name": sheet_param, "column": column}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, last_row)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "最終列を取得する":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            row_field = PlainField(f, "対象行番号(空欄でシート全体)")
            row_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                row = row_field.get().strip() or None
                try:
                    last_col = self.recorder.excel.get_last_column(sheet_test, row=row)
                    self.log(f"→ 取得できました: {last_col or '(該当する列なし)'}列目")
                    store_as = self._ask_store_as()
                    step = {"handler": "excel", "action": "get_last_column",
                            "params": {"sheet_name": sheet_param, "row": row}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, last_col)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "セルをコピーして貼り付ける":
            ttk.Label(
                f,
                text="コピー元の範囲欄には {{last_row}} のように前の手順の結果を埋め込めます\n"
                     "(例: A1:B{{last_row}} 。変数名の後ろに+数値/-数値も書けます)",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            src_field = ValueSlotField(f, "コピー元の範囲(例: A1:B3 または A1:B{{last_row}})")
            src_field.pack(fill="x", pady=4)

            ttk.Label(f, text="貼り付け先の指定方法:").pack(anchor="w", pady=(6, 0))
            dest_basis_var = tk.StringVar(value="direct")
            dest_basis_frame = ttk.Frame(f)
            dest_basis_frame.pack(anchor="w", pady=2)
            ttk.Radiobutton(
                dest_basis_frame, text="直接指定", variable=dest_basis_var, value="direct"
            ).pack(side="left")
            ttk.Radiobutton(
                dest_basis_frame, text="最終行の続き(列を指定)", variable=dest_basis_var, value="last_row"
            ).pack(side="left")
            ttk.Radiobutton(
                dest_basis_frame, text="最終列の続き(行を指定)", variable=dest_basis_var, value="last_col"
            ).pack(side="left")

            dest_sub_frame = ttk.Frame(f)
            dest_sub_frame.pack(fill="x", pady=4)
            dest_widgets: dict = {}

            def rebuild_dest_sub(*_args):
                for child in dest_sub_frame.winfo_children():
                    child.destroy()
                basis = dest_basis_var.get()
                if basis == "last_row":
                    dest_widgets["col"] = PlainField(dest_sub_frame, "貼り付け先の列(例: D)")
                    dest_widgets["col"].pack(fill="x", pady=2)
                    dest_widgets["scope"] = PlainField(dest_sub_frame, "最終行を判定する基準列(空欄でシート全体)")
                    dest_widgets["scope"].pack(fill="x", pady=2)
                    dest_widgets["offset"] = PlainField(dest_sub_frame, "最終行から何行後に貼り付けるか(空欄で1)")
                    dest_widgets["offset"].pack(fill="x", pady=2)
                elif basis == "last_col":
                    dest_widgets["row"] = PlainField(dest_sub_frame, "貼り付け先の行番号(例: 5)")
                    dest_widgets["row"].pack(fill="x", pady=2)
                    dest_widgets["offset"] = PlainField(dest_sub_frame, "最終列から何列後に貼り付けるか(空欄で1)")
                    dest_widgets["offset"].pack(fill="x", pady=2)
                else:
                    dest_widgets["cell"] = ValueSlotField(dest_sub_frame, "貼り付け先の左上セル(例: D1)")
                    dest_widgets["cell"].pack(fill="x")

            dest_basis_var.trace_add("write", rebuild_dest_sub)
            rebuild_dest_sub()

            ttk.Label(f, text="貼り付け方法:").pack(anchor="w", pady=(6, 0))
            paste_type_var = tk.StringVar(value="values")
            paste_frame = ttk.Frame(f)
            paste_frame.pack(anchor="w", pady=2)
            ttk.Radiobutton(paste_frame, text="値のみ", variable=paste_type_var, value="values").pack(side="left")
            ttk.Radiobutton(paste_frame, text="数式", variable=paste_type_var, value="formulas").pack(side="left")
            ttk.Radiobutton(paste_frame, text="すべて(値/数式+書式)", variable=paste_type_var, value="all").pack(side="left")

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                src_test, src_param, _ = src_field.get()
                paste_type = paste_type_var.get()
                dest_basis = dest_basis_var.get()

                prereq_step = None
                prereq_value = None
                if dest_basis == "last_row":
                    col_raw = dest_widgets["col"].get().strip()
                    try:
                        col_idx = int(col_raw) if col_raw.isdigit() else column_index_from_string(col_raw.upper())
                        col_letter = get_column_letter(col_idx)
                    except ValueError:
                        self.log("⚠ 貼り付け先の列の指定が不正です")
                        return
                    scope = dest_widgets["scope"].get().strip()
                    try:
                        offset = int(dest_widgets["offset"].get().strip() or "1")
                    except ValueError:
                        offset = 1
                    try:
                        current_last_row = self.recorder.excel.get_last_row(sheet_test, column=scope or None)
                    except Exception as e:  # noqa: BLE001
                        self.log(f"⚠ {e}")
                        return
                    var_name = self.recorder._next_auto_var("last_row")
                    prereq_step = {
                        "handler": "excel", "action": "get_last_row",
                        "params": {"sheet_name": sheet_param, "column": scope or None},
                        "store_as": var_name,
                    }
                    prereq_value = current_last_row
                    dest_row_test = current_last_row + offset
                    dest_test = f"{col_letter}{dest_row_test}"
                    dest_param = f"{col_letter}{_offset_template(var_name, offset)}"
                    self.log(f"→ 今の時点の最終行は{current_last_row}行目なので、動作確認では{dest_test}に貼り付けます")
                elif dest_basis == "last_col":
                    row_raw = dest_widgets["row"].get().strip()
                    if not row_raw.isdigit():
                        self.log("⚠ 貼り付け先の行番号は数字で入力してください")
                        return
                    try:
                        offset = int(dest_widgets["offset"].get().strip() or "1")
                    except ValueError:
                        offset = 1
                    try:
                        current_last_col = self.recorder.excel.get_last_column(sheet_test, row=row_raw)
                    except Exception as e:  # noqa: BLE001
                        self.log(f"⚠ {e}")
                        return
                    col_idx_base = column_index_from_string(current_last_col) if current_last_col else 0
                    if col_idx_base + offset < 1:
                        self.log("⚠ 指定したオフセットでは列がA列より前になります")
                        return
                    var_name = self.recorder._next_auto_var("last_col")
                    prereq_step = {
                        "handler": "excel", "action": "get_last_column",
                        "params": {"sheet_name": sheet_param, "row": int(row_raw)},
                        "store_as": var_name,
                    }
                    prereq_value = current_last_col
                    dest_col_test = get_column_letter(col_idx_base + offset)
                    dest_test = f"{dest_col_test}{row_raw}"
                    dest_param = f"{_offset_template(var_name, offset)}{row_raw}"
                    self.log(
                        f"→ 今の時点の最終列は{current_last_col or '(無し)'}なので、"
                        f"動作確認では{dest_test}に貼り付けます"
                    )
                else:
                    dest_test, dest_param, _ = dest_widgets["cell"].get()

                step = {
                    "handler": "excel", "action": "copy_cell_range",
                    "params": {
                        "sheet_name": sheet_param, "source_range": src_param,
                        "dest_cell": dest_param, "paste_type": paste_type,
                    },
                }
                if self._has_template(src_test) or self._has_template(sheet_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    if prereq_step:
                        self.register_step(prereq_step, prereq_value)
                    self.register_step(step)
                    return
                try:
                    result = self.recorder.excel.copy_cell_range(
                        sheet_test, src_test, dest_test, paste_type=paste_type
                    )
                    self.log(f"→ {result}")
                    if prereq_step:
                        self.register_step(prereq_step, prereq_value)
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "シートをコピーする":
            src_field = ValueSlotField(f, "コピー元のシート名")
            src_field.pack(fill="x", pady=4)
            new_field = ValueSlotField(f, "新しいシート名")
            new_field.pack(fill="x", pady=4)

            def on_submit():
                src_test, src_param, _ = src_field.get()
                new_test, new_param, _ = new_field.get()
                try:
                    result = self.recorder.excel.copy_sheet(src_test, new_test)
                    self.log(f"→ シートをコピーできました: {result}")
                    self.register_step({
                        "handler": "excel", "action": "copy_sheet",
                        "params": {"source_sheet_name": src_param, "new_sheet_name": new_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "シート名の一覧を取得する":
            def on_submit():
                try:
                    names = self.recorder.excel.get_sheet_names()
                    self.log(f"→ 取得できました: {names}")
                    store_as = self._ask_store_as()
                    step = {"handler": "excel", "action": "get_sheet_names", "params": {}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, names)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "印刷範囲を指定する":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            range_field = ValueSlotField(f, "印刷範囲(例: A1:H30)")
            range_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                range_test, range_param, _ = range_field.get()
                try:
                    result = self.recorder.excel.set_print_area(sheet_test, range_test)
                    self.log(f"→ {result}")
                    self.register_step({
                        "handler": "excel", "action": "set_print_area",
                        "params": {"sheet_name": sheet_param, "cell_range": range_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "新規シートを追加する":
            name_field = ValueSlotField(f, "追加する新規シート名")
            name_field.pack(fill="x", pady=4)
            index_field = PlainField(f, "挿入位置(0で先頭、空欄で末尾に追加)")
            index_field.pack(fill="x", pady=4)

            def on_submit():
                name_test, name_param, _ = name_field.get()
                idx_raw = index_field.get().strip()
                index = None
                if idx_raw:
                    try:
                        index = int(idx_raw)
                    except ValueError:
                        self.log("⚠ 挿入位置は数字で入力してください")
                        return
                step = {"handler": "excel", "action": "create_sheet",
                        "params": {"sheet_name": name_param, "index": index}}
                if self._has_template(name_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    result = self.recorder.excel.create_sheet(name_test, index=index)
                    self.log(f"→ 追加できました: {result}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "シート名を変更する":
            old_field = ValueSlotField(f, "名前を変更する既存シート名")
            old_field.pack(fill="x", pady=4)
            new_field = ValueSlotField(f, "変更後の新しいシート名")
            new_field.pack(fill="x", pady=4)

            def on_submit():
                old_test, old_param, _ = old_field.get()
                new_test, new_param, _ = new_field.get()
                step = {"handler": "excel", "action": "rename_sheet",
                        "params": {"old_name": old_param, "new_name": new_param}}
                if self._has_template(old_test) or self._has_template(new_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    result = self.recorder.excel.rename_sheet(old_test, new_test)
                    self.log(f"→ 変更できました: {result}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "セル範囲(1列/1行)をリストとして取得する":
            ttk.Label(
                f, text="1列(例: A1:A10)または1行(例: A1:J1)の範囲を指定してください。",
                foreground="#557",
            ).pack(anchor="w", pady=(0, 6))
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            range_field = ValueSlotField(f, "範囲(1列または1行)")
            range_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                range_test, range_param, _ = range_field.get()
                try:
                    result = self.recorder.excel.get_range_as_list(sheet_test, range_test)
                    self.log(f"→ 取得できました({len(result)}件): {result[:5]}")
                    store_as = self._ask_store_as()
                    step = {"handler": "excel", "action": "get_range_as_list",
                            "params": {"sheet_name": sheet_param, "cell_range": range_param}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, result)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "値でセルを検索する":
            sheet_field = ValueSlotField(f, "検索対象のシート名")
            sheet_field.pack(fill="x", pady=4)
            value_field = ValueSlotField(f, "検索する値")
            value_field.pack(fill="x", pady=4)
            col_field = PlainField(f, "検索する列(例: A。空欄でシート全体)")
            col_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                value_test, value_param, _ = value_field.get()
                column = col_field.get().strip() or None
                try:
                    address = self.recorder.excel.find_cell(sheet_test, value_test, column=column)
                    self.log(f"→ 見つかりました: {address}")
                    store_as = self._ask_store_as()
                    step = {"handler": "excel", "action": "find_cell",
                            "params": {"sheet_name": sheet_param, "value": value_param, "column": column}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, address)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "表の末尾に1行追記する":
            sheet_field = ValueSlotField(f, "追記先のシート名")
            sheet_field.pack(fill="x", pady=4)
            pairs = PairsField(f, "列(例:1)", "値")
            ttk.Label(
                f, text="左の列番号は「開始列からの何番目か」(1始まり)。例: 開始列A・列2 なら実際はB列。",
                foreground="#557",
            ).pack(anchor="w")
            pairs.pack(fill="x", pady=4)
            start_col_field = PlainField(f, "開始列(空欄でA)", default="A")
            start_col_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                raw_pairs = pairs.get()
                if not raw_pairs:
                    self.log("⚠ 値が1つも入力されていません")
                    return
                try:
                    ordered = sorted(raw_pairs.items(), key=lambda kv: int(kv[0]))
                    values = [v for _, v in ordered]
                except ValueError:
                    self.log("⚠ 列は数字(1, 2, 3...)で入力してください")
                    return
                start_column = start_col_field.get().strip() or "A"
                try:
                    result = self.recorder.excel.append_row(sheet_test, values, start_column=start_column)
                    self.log(f"→ 追記できました: {result}")
                    self.register_step({
                        "handler": "excel", "action": "append_row",
                        "params": {"sheet_name": sheet_param, "values": values, "start_column": start_column},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "行を挿入する":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            row_field = PlainField(f, "何行目の前に挿入するか(行番号)")
            row_field.pack(fill="x", pady=4)
            count_field = PlainField(f, "挿入する行数(空欄で1)", default="1")
            count_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                row_raw = row_field.get().strip()
                if not row_raw.isdigit():
                    self.log("⚠ 行番号は数字で入力してください")
                    return
                count = int(count_field.get().strip() or "1")
                try:
                    result = self.recorder.excel.insert_rows(sheet_test, int(row_raw), count=count)
                    self.log(f"→ 挿入できました: {result}")
                    self.register_step({
                        "handler": "excel", "action": "insert_rows",
                        "params": {"sheet_name": sheet_param, "row": int(row_raw), "count": count},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "行を削除する":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            row_field = PlainField(f, "何行目から削除するか(行番号)")
            row_field.pack(fill="x", pady=4)
            count_field = PlainField(f, "削除する行数(空欄で1)", default="1")
            count_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                row_raw = row_field.get().strip()
                if not row_raw.isdigit():
                    self.log("⚠ 行番号は数字で入力してください")
                    return
                count = int(count_field.get().strip() or "1")
                if not self._confirm(
                    f"実際にシート '{sheet_test}' の{row_raw}行目から{count}行を削除して"
                    f"動作確認します(元に戻せません)。よろしいですか?"
                ):
                    return
                try:
                    result = self.recorder.excel.delete_rows(sheet_test, int(row_raw), count=count)
                    self.log(f"→ 削除できました: {result}")
                    self.register_step({
                        "handler": "excel", "action": "delete_rows",
                        "params": {"sheet_name": sheet_param, "row": int(row_raw), "count": count},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "セル範囲の値を空にする":
            sheet_field = ValueSlotField(f, "対象シート名")
            sheet_field.pack(fill="x", pady=4)
            range_field = ValueSlotField(f, "空にする範囲(例: A2:C10)")
            range_field.pack(fill="x", pady=4)

            def on_submit():
                sheet_test, sheet_param, _ = sheet_field.get()
                range_test, range_param, _ = range_field.get()
                try:
                    result = self.recorder.excel.clear_range(sheet_test, range_test)
                    self.log(f"→ {result}")
                    self.register_step({
                        "handler": "excel", "action": "clear_range",
                        "params": {"sheet_name": sheet_param, "cell_range": range_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "シートを削除する":
            name_field = ValueSlotField(f, "削除するシート名")
            name_field.pack(fill="x", pady=4)

            def on_submit():
                name_test, name_param, _ = name_field.get()
                if not self._confirm(
                    f"実際にシート '{name_test}' を削除して動作確認します(元に戻せません)。よろしいですか?"
                ):
                    return
                try:
                    result = self.recorder.excel.delete_sheet(name_test)
                    self.log(f"→ 削除できました: {result}")
                    self.register_step({
                        "handler": "excel", "action": "delete_sheet",
                        "params": {"sheet_name": name_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

    # ---------- PDF ----------

    def _build_pdf(self, action: str) -> None:
        f = self.form_frame
        if action == "テキストを抽出する(ページ全体)":
            in_field = ValueSlotField(f, "抽出元PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力テキストファイルのパス")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".txt"))
            out_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                step = {"handler": "pdf", "action": "extract_text",
                        "params": {"input_path": in_param, "output_path": out_param}}
                if self._has_template(in_test) or self._has_template(out_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.pdf.extract_text(in_test, out_test)
                    self.log(f"→ 抽出できました: {out_test}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "複数PDFを結合する":
            paths_field = PathListField(f, "結合するPDF(2件以上)")
            paths_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力先のパス(例: merged_{{last_row}}.pdf も可)")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".pdf"))
            out_field.pack(fill="x", pady=4)

            def on_submit():
                paths = paths_field.get()
                if len(paths) < 2:
                    self.log("⚠ 2件以上のPDFを追加してください")
                    return
                out_test, out_param, _ = out_field.get()
                step = {"handler": "pdf", "action": "merge_pdfs",
                        "params": {"input_paths": paths, "output_path": out_param}}
                if self._has_template(out_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.pdf.merge_pdfs(paths, out_test)
                    self.log(f"→ 結合できました: {out_test}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "PDFを分割する":
            in_field = ValueSlotField(f, "分割元PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力先フォルダ")
            out_field.add_button("参照...", out_field.browse_dir)
            out_field.pack(fill="x", pady=4)
            pages_field = PlainField(f, "何ページごとに分割するか", default="1")
            pages_field.pack(fill="x", pady=4)
            range_field = PlainField(f, "対象ページ範囲(例: 3,10。空欄で全ページ)")
            range_field.pack(fill="x", pady=4)
            pattern_field = PlainField(
                f, "出力ファイル名ルール(空欄で既定 {stem}_part{part})", width=30
            )
            pattern_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                try:
                    pages_per_file = int(pages_field.get() or "1")
                except ValueError:
                    pages_per_file = 1
                start_page = end_page = None
                range_raw = range_field.get().strip()
                if range_raw:
                    try:
                        s, e_ = [x.strip() for x in range_raw.split(",")]
                        start_page, end_page = int(s), int(e_)
                    except ValueError:
                        self.log("⚠ ページ範囲の形式が正しくないため、全ページを対象にします")
                pattern = pattern_field.get().strip() or None
                step = {
                    "handler": "pdf", "action": "split_pdf",
                    "params": {
                        "input_path": in_param, "output_dir": out_param,
                        "pages_per_file": pages_per_file,
                        "start_page": start_page, "end_page": end_page,
                        "filename_pattern": pattern,
                    },
                }
                if self._has_template(in_test) or self._has_template(out_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    outputs = self.recorder.pdf.split_pdf(
                        in_test, out_test, pages_per_file,
                        start_page=start_page, end_page=end_page, filename_pattern=pattern,
                    )
                    self.log(f"→ {len(outputs)}件に分割できました(例: {outputs[0]})")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "PDFを回転する":
            in_field = ValueSlotField(f, "回転元PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力先のパス")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".pdf"))
            out_field.pack(fill="x", pady=4)
            degrees_field = PlainField(f, "回転角度(90/180/270)", default="90")
            degrees_field.pack(fill="x", pady=4)
            pages_field = PlainField(f, "対象ページ(例:1,3。空欄で全ページ)")
            pages_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                try:
                    degrees = int(degrees_field.get() or "90")
                except ValueError:
                    degrees = 90
                pages_raw = pages_field.get().strip()
                pages = None
                if pages_raw:
                    try:
                        pages = [int(x.strip()) for x in pages_raw.split(",") if x.strip()]
                    except ValueError:
                        pages = None
                step = {
                    "handler": "pdf", "action": "rotate_pdf",
                    "params": {"input_path": in_param, "output_path": out_param,
                               "degrees": degrees, "pages": pages},
                }
                if self._has_template(in_test) or self._has_template(out_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.pdf.rotate_pdf(in_test, out_test, degrees, pages)
                    self.log(f"→ 回転できました: {out_test}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "OCRでテキスト化する(ページ全体)":
            ttk.Label(f, text="※ pytesseract/pdf2image + Tesseract OCR/Poppler が必要です。",
                      foreground="#a55").pack(anchor="w", pady=(0, 4))
            in_field = ValueSlotField(f, "OCR対象PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力テキストファイルのパス")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".txt"))
            out_field.pack(fill="x", pady=4)
            lang_field = PlainField(f, "OCR言語", default="jpn+eng")
            lang_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                language = lang_field.get() or "jpn+eng"
                step = {
                    "handler": "pdf", "action": "ocr_pdf_to_text",
                    "params": {"input_path": in_param, "output_path": out_param, "language": language},
                }
                if self._has_template(in_test) or self._has_template(out_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.pdf.ocr_pdf_to_text(in_test, out_test, language)
                    self.log(f"→ OCR抽出できました: {out_test}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    if self._confirm("未確認のままこの手順を登録しますか?"):
                        self.register_step(step)

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "範囲を指定してテキストを取得する":
            ttk.Label(
                f, text="※ 基準点はページ左上(0,0)、x/yはページ幅・高さに対する%(0〜100)で指定します。",
                foreground="#557",
            ).pack(anchor="w", pady=(0, 4))
            in_field = ValueSlotField(f, "対象PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            xl_field = PlainField(f, "x_left(0〜100)", default="0")
            xl_field.pack(fill="x", pady=2)
            xr_field = PlainField(f, "x_right(0〜100)", default="100")
            xr_field.pack(fill="x", pady=2)
            yu_field = PlainField(f, "y_upper(0〜100)", default="0")
            yu_field.pack(fill="x", pady=2)
            yl_field = PlainField(f, "y_lower(0〜100)", default="100")
            yl_field.pack(fill="x", pady=2)
            page_field = PlainField(f, "対象ページ番号", default="1")
            page_field.pack(fill="x", pady=4)
            ocr_field = BoolField(f, "文字が選択できないPDF(OCRを使う)")
            ocr_field.pack(anchor="w", pady=2)
            lang_field = PlainField(f, "OCR言語(OCR使用時)", default="jpn+eng")
            lang_field.pack(fill="x", pady=2)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                try:
                    x_left = float(xl_field.get())
                    x_right = float(xr_field.get())
                    y_upper = float(yu_field.get())
                    y_lower = float(yl_field.get())
                    page_number = int(page_field.get() or "1")
                except ValueError:
                    self.log("⚠ 数値を正しく入力してください")
                    return
                ocr = ocr_field.get()
                language = lang_field.get() or "jpn+eng"
                try:
                    result = self.recorder.pdf.extract_text_in_area(
                        in_test, x_left, x_right, y_upper, y_lower,
                        page_number=page_number, ocr=ocr, ocr_language=language,
                    )
                    self.log(f"→ 取得できました: {result[:150]!r}")
                    store_as = self._ask_store_as()
                    step = {
                        "handler": "pdf", "action": "extract_text_in_area",
                        "params": {
                            "input_path": in_param, "x_left": x_left, "x_right": x_right,
                            "y_upper": y_upper, "y_lower": y_lower, "page_number": page_number,
                            "ocr": ocr, "ocr_language": language,
                        },
                    }
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, result)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "ページ範囲を1ファイルに抜き出す":
            in_field = ValueSlotField(f, "対象PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力先のパス")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".pdf"))
            out_field.pack(fill="x", pady=4)
            start_field = PlainField(f, "開始ページ番号", default="1")
            start_field.pack(fill="x", pady=2)
            end_field = PlainField(f, "終了ページ番号", default="1")
            end_field.pack(fill="x", pady=2)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                try:
                    start_page, end_page = int(start_field.get()), int(end_field.get())
                except ValueError:
                    self.log("⚠ ページ番号は数字で入力してください")
                    return
                step = {
                    "handler": "pdf", "action": "extract_page_range",
                    "params": {"input_path": in_param, "output_path": out_param,
                               "start_page": start_page, "end_page": end_page},
                }
                if self._has_template(in_test) or self._has_template(out_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                try:
                    self.recorder.pdf.extract_page_range(in_test, out_test, start_page, end_page)
                    self.log(f"→ {start_page}〜{end_page}ページを抜き出せました: {out_test}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "ページ数を取得する":
            in_field = ValueSlotField(f, "対象PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                try:
                    count = self.recorder.pdf.get_page_count(in_test)
                    self.log(f"→ 取得できました: {count}ページ")
                    store_as = self._ask_store_as()
                    step = {"handler": "pdf", "action": "get_page_count", "params": {"input_path": in_param}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, count)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "表(罫線あり)をCSVとして抽出する":
            in_field = ValueSlotField(f, "対象PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力CSVファイルのパス")
            out_field.add_button("参照...", lambda: out_field.browse_save_file(".csv"))
            out_field.pack(fill="x", pady=4)
            page_field = PlainField(f, "対象ページ番号(空欄で全ページ)")
            page_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                page_raw = page_field.get().strip()
                page_number = int(page_raw) if page_raw.isdigit() else None
                try:
                    self.recorder.pdf.extract_tables(in_test, out_test, page_number=page_number)
                    self.log(f"→ 抽出できました: {out_test}")
                    self.register_step({
                        "handler": "pdf", "action": "extract_tables",
                        "params": {"input_path": in_param, "output_path": out_param, "page_number": page_number},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "埋め込み画像を抽出する":
            in_field = ValueSlotField(f, "対象PDFのパス")
            in_field.add_button("参照...", in_field.browse_file)
            in_field.pack(fill="x", pady=4)
            out_field = ValueSlotField(f, "出力先フォルダのパス")
            out_field.add_button("参照...", out_field.browse_dir)
            out_field.pack(fill="x", pady=4)
            page_field = PlainField(f, "対象ページ番号(空欄で全ページ)")
            page_field.pack(fill="x", pady=4)

            def on_submit():
                in_test, in_param, _ = in_field.get()
                out_test, out_param, _ = out_field.get()
                page_raw = page_field.get().strip()
                page_number = int(page_raw) if page_raw.isdigit() else None
                try:
                    outputs = self.recorder.pdf.extract_images(in_test, out_test, page_number=page_number)
                    self.log(f"→ 抽出できました({len(outputs)}件)")
                    self.register_step({
                        "handler": "pdf", "action": "extract_images",
                        "params": {"input_path": in_param, "output_dir": out_param, "page_number": page_number},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

    # ---------- Web ----------

    def _current_site_names(self) -> list[str]:
        return list(self.recorder.browser._sites.keys())

    def _build_web(self, action: str) -> None:
        f = self.form_frame

        if action == "サイトを開く/切り替える":
            ttk.Label(f, text="登録済みサイト:").pack(anchor="w")
            site_combo = ttk.Combobox(f, state="readonly", values=self._current_site_names())
            if self._current_site_names():
                site_combo.current(0)
            site_combo.pack(fill="x", pady=4)

            def open_existing():
                key = site_combo.get()
                if not key:
                    self.log("⚠ サイトを選択してください")
                    return
                try:
                    self.recorder.browser.open_registered_site(key)
                    self.recorder._site_opened = True
                    self.log(f"→ '{key}' を開きました")
                    self.register_step({"handler": "browser", "action": "open_registered_site",
                                         "params": {"site_key": key}})
                except SiteNotWhitelistedError as e:
                    self.log(f"⚠ {e}")
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ ブラウザを開けませんでした: {e}")

            ttk.Button(f, text="このサイトを開いて登録", command=open_existing).pack(pady=4)

            ttk.Separator(f).pack(fill="x", pady=8)
            ttk.Label(f, text="新しいサイトを登録:").pack(anchor="w")
            key_field = PlainField(f, "サイトの識別名(半角英数字)")
            key_field.pack(fill="x", pady=2)
            url_field = PlainField(f, "サイトのURL", width=50)
            url_field.pack(fill="x", pady=2)

            def register_new():
                key = key_field.get().strip()
                url = url_field.get().strip()
                if not key or not url:
                    self.log("⚠ 識別名とURLの両方を入力してください")
                    return
                try:
                    self.recorder.browser.register_site(key, url)
                    self.recorder.browser.open_registered_site(key)
                    self.recorder._site_opened = True
                    self.log(f"→ '{key}' を登録して開きました")
                    self.register_step({"handler": "browser", "action": "open_registered_site",
                                         "params": {"site_key": key}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="登録して開く", command=register_new).pack(pady=4)
            return

        if not self.recorder._site_opened:
            ttk.Label(f, text="先に「サイトを開く/切り替える」でサイトを開いてください。",
                      foreground="#a55").pack(pady=10)
            return

        if action == "クリックする":
            ttk.Button(f, text="今の画面で押せそうな文字を見る", command=self._show_web_candidates).pack(
                anchor="w"
            )
            field = ValueSlotField(f, "押したいボタン/リンクの表示文字")
            field.pack(fill="x", pady=4)

            def on_submit():
                text_test, text_param, _ = field.get()
                try:
                    self.recorder.browser.click_by_text(text_test)
                    self.log(f"→ '{text_test}' をクリックできました")
                    verify_cfg = self._ask_verify()
                    retry_cfg = self._ask_retry()
                    self.register_step({
                        "handler": "browser", "action": "click_by_text",
                        "params": {"text_hint": text_param},
                        "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
                    })
                except ElementNotFoundError as e:
                    self.log(f"⚠ {e}")
                    selector = simpledialog.askstring(
                        "CSSセレクタ", "F12で調べたCSSセレクタがあれば入力(無ければ空欄でOK):", parent=self
                    )
                    if selector:
                        try:
                            self.recorder.browser.click_selector(selector)
                            self.log(f"→ セレクタ '{selector}' でクリックできました")
                            verify_cfg = self._ask_verify()
                            retry_cfg = self._ask_retry()
                            self.register_step({
                                "handler": "browser", "action": "click_selector",
                                "params": {"selector": selector},
                                "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
                            })
                        except Exception as e2:  # noqa: BLE001
                            self.log(f"⚠ セレクタでも失敗: {e2}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "入力する":
            ttk.Button(f, text="今の画面で押せそうな文字を見る", command=self._show_web_candidates).pack(
                anchor="w"
            )
            label_field = ValueSlotField(f, "入力欄の目印(ラベル等)")
            label_field.pack(fill="x", pady=4)
            value_field = ValueSlotField(f, "入力する値")
            value_field.pack(fill="x", pady=4)
            enter_field = BoolField(f, "入力後にEnterキーで送信する")
            enter_field.pack(anchor="w", pady=2)

            def on_submit():
                label_test, label_param, _ = label_field.get()
                value_test, value_param, _ = value_field.get()
                press_enter = enter_field.get()
                try:
                    self.recorder.browser.type_by_text(label_test, value_test, press_enter=press_enter)
                    self.log(f"→ '{label_test}' へ入力できました")
                    verify_cfg = self._ask_verify()
                    retry_cfg = self._ask_retry()
                    self.register_step({
                        "handler": "browser", "action": "type_by_text",
                        "params": {"label_hint": label_param, "value": value_param, "press_enter": press_enter},
                        "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
                    })
                except ElementNotFoundError as e:
                    self.log(f"⚠ {e}")
                    selector = simpledialog.askstring(
                        "CSSセレクタ", "F12で調べたCSSセレクタがあれば入力(無ければ空欄でOK):", parent=self
                    )
                    if selector:
                        try:
                            self.recorder.browser.type_by_selector(selector, value_test, press_enter=press_enter)
                            self.log(f"→ セレクタ '{selector}' で入力できました")
                            verify_cfg = self._ask_verify()
                            retry_cfg = self._ask_retry()
                            self.register_step({
                                "handler": "browser", "action": "type_by_selector",
                                "params": {"selector": selector, "value": value_param, "press_enter": press_enter},
                                "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
                            })
                        except Exception as e2:  # noqa: BLE001
                            self.log(f"⚠ セレクタでも失敗: {e2}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "選択する":
            label_field = ValueSlotField(f, "ドロップダウンの目印")
            label_field.pack(fill="x", pady=4)
            option_field = PlainField(f, "選択したい選択肢の表示文字", width=30)
            option_field.pack(fill="x", pady=4)

            def on_submit():
                label_test, label_param, _ = label_field.get()
                option_text = option_field.get()
                try:
                    self.recorder.browser.select_by_text(label_test, option_text)
                    self.log(f"→ '{option_text}' を選択できました")
                    verify_cfg = self._ask_verify()
                    retry_cfg = self._ask_retry()
                    self.register_step({
                        "handler": "browser", "action": "select_by_text",
                        "params": {"label_hint": label_param, "option_text": option_text},
                        "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
                    })
                except ElementNotFoundError as e:
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "待機する":
            seconds_field = PlainField(f, "待機する秒数", default="2")
            seconds_field.pack(fill="x", pady=4)

            def on_submit():
                try:
                    seconds = float(seconds_field.get())
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                self.register_step({
                    "handler": "browser", "action": "wait_seconds",
                    "params": {"seconds": seconds},
                    "verify": {"type": "none"}, "verify_skip": False,
                })
                self.log(f"→ {seconds}秒の待機を登録しました")

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "画面をPDF保存する":
            field = ValueSlotField(f, "保存先PDFのパス")
            field.add_button("参照...", lambda: field.browse_save_file(".pdf"))
            field.pack(fill="x", pady=4)
            scale_field = PlainField(f, "倍率(1ページに収めたい場合等)", default="1.0")
            scale_field.pack(fill="x", pady=4)
            landscape_field = BoolField(f, "横向きで保存する")
            landscape_field.pack(anchor="w", pady=2)

            def on_submit():
                test_v, param_v, _ = field.get()
                try:
                    scale = float(scale_field.get() or "1.0")
                except ValueError:
                    scale = 1.0
                landscape = landscape_field.get()
                try:
                    self.recorder.browser.save_page_as_pdf(test_v, scale=scale, landscape=landscape)
                    self.log(f"→ PDF保存できました: {test_v}")
                    self.register_step({
                        "handler": "browser", "action": "save_page_as_pdf",
                        "params": {"save_path": param_v, "scale": scale, "landscape": landscape},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "画面から文字を読み取る":
            ttk.Label(f, text="F12の開発者ツールで調べたCSSセレクタ(class/id等)を指定してください。",
                      foreground="#557").pack(anchor="w", pady=(0, 4))
            selector_field = PlainField(f, "CSSセレクタ", width=40)
            selector_field.pack(fill="x", pady=4)

            def on_submit():
                selector = selector_field.get().strip()
                try:
                    text = self.recorder.browser.get_text_by_selector(selector)
                    self.log(f"→ 読み取れました: {text!r}")
                    store_as = self._ask_store_as()
                    step = {"handler": "browser", "action": "get_text_by_selector", "params": {"selector": selector}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, text)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "画面から属性値を読み取る":
            ttk.Label(f, text="F12の開発者ツールで調べたCSSセレクタ(class/id等)を指定してください。",
                      foreground="#557").pack(anchor="w", pady=(0, 4))
            selector_field = PlainField(f, "CSSセレクタ", width=40)
            selector_field.pack(fill="x", pady=4)
            attr_field = PlainField(f, "属性名(例: href, value, data-id)")
            attr_field.pack(fill="x", pady=4)

            def on_submit():
                selector = selector_field.get().strip()
                attribute = attr_field.get().strip()
                try:
                    value = self.recorder.browser.get_attribute_by_selector(selector, attribute)
                    self.log(f"→ 読み取れました: {value!r}")
                    store_as = self._ask_store_as()
                    step = {
                        "handler": "browser", "action": "get_attribute_by_selector",
                        "params": {"selector": selector, "attribute": attribute},
                    }
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, value)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "画面から文字のリストを読み取る":
            ttk.Label(
                f, text="一致するすべての要素の文字をリストとして読み取ります"
                        "(例: 表の1列全体なら \"table tr td:nth-child(2)\")。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 4))
            selector_field = PlainField(f, "CSSセレクタ", width=40)
            selector_field.pack(fill="x", pady=4)

            def on_submit():
                selector = selector_field.get().strip()
                try:
                    values = self.recorder.browser.get_text_list_by_selector(selector)
                    self.log(f"→ 読み取れました({len(values)}件): {values[:5]}")
                    store_as = self._ask_store_as()
                    step = {
                        "handler": "browser", "action": "get_text_list_by_selector",
                        "params": {"selector": selector},
                    }
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, values)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "チェックボックスをON/OFFする":
            label_field = ValueSlotField(f, "チェックボックスの目印(ラベルの表示文字)")
            label_field.pack(fill="x", pady=4)
            checked_field = BoolField(f, "チェックを入れる(オフでチェックを外す)", default=True)
            checked_field.pack(anchor="w", pady=2)

            def on_submit():
                label_test, label_param, _ = label_field.get()
                checked = checked_field.get()
                try:
                    self.recorder.browser.check_checkbox_by_text(label_test, checked=checked)
                    self.log(f"→ '{label_test}' を {checked} にできました")
                    verify_cfg = self._ask_verify()
                    retry_cfg = self._ask_retry()
                    self.register_step({
                        "handler": "browser", "action": "check_checkbox_by_text",
                        "params": {"label_hint": label_param, "checked": checked},
                        "verify": verify_cfg, "verify_skip": False, "retry": retry_cfg,
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "ウィンドウサイズを指定する":
            ttk.Label(
                f, text="pyautoguiと併用する場合、ウィンドウサイズ・位置・表示倍率を"
                        "固定しておくと座標がずれにくくなります。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            percent_field = BoolField(f, "画面全体に対する割合(%)で指定する(オフ=ピクセル指定)")
            percent_field.pack(anchor="w", pady=2)
            width_field = PlainField(f, "幅(空欄で変更しない)")
            width_field.pack(fill="x", pady=4)
            height_field = PlainField(f, "高さ(空欄で変更しない)")
            height_field.pack(fill="x", pady=4)

            def on_submit():
                use_percent = percent_field.get()
                w_raw = width_field.get().strip()
                h_raw = height_field.get().strip()
                width = height = width_percent = height_percent = None
                try:
                    if use_percent:
                        width_percent = float(w_raw) if w_raw else None
                        height_percent = float(h_raw) if h_raw else None
                    else:
                        width = int(w_raw) if w_raw else None
                        height = int(h_raw) if h_raw else None
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                try:
                    result = self.recorder.browser.set_window_size(
                        width=width, height=height,
                        width_percent=width_percent, height_percent=height_percent,
                    )
                    self.log(f"→ 設定できました: {result}")
                    self.register_step({
                        "handler": "browser", "action": "set_window_size",
                        "params": {
                            "width": width, "height": height,
                            "width_percent": width_percent, "height_percent": height_percent,
                        },
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "ウィンドウ位置を指定する":
            ttk.Label(f, text="画面の左上を基準(0,0)として、右方向・下方向がプラスです。",
                      foreground="#557").pack(anchor="w", pady=(0, 6))
            x_field = PlainField(f, "X座標")
            x_field.pack(fill="x", pady=4)
            y_field = PlainField(f, "Y座標")
            y_field.pack(fill="x", pady=4)

            def on_submit():
                try:
                    x, y = int(x_field.get()), int(y_field.get())
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                try:
                    result = self.recorder.browser.set_window_position(x, y)
                    self.log(f"→ 設定できました: {result}")
                    self.register_step({
                        "handler": "browser", "action": "set_window_position",
                        "params": {"x": x, "y": y},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "表示倍率(ズーム)を指定する":
            ttk.Label(
                f, text="※ 別のページに遷移すると設定はリセットされます"
                        "(維持したい場合は遷移のたびに登録し直してください)。",
                foreground="#a55",
            ).pack(anchor="w", pady=(0, 6))
            percent_field = PlainField(f, "表示倍率(%。100=等倍)", default="100")
            percent_field.pack(fill="x", pady=4)

            def on_submit():
                try:
                    percent = float(percent_field.get() or "100")
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                try:
                    result = self.recorder.browser.set_zoom(percent)
                    self.log(f"→ 設定できました: {result}")
                    self.register_step({
                        "handler": "browser", "action": "set_zoom",
                        "params": {"percent": percent},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

    def _show_web_candidates(self) -> None:
        try:
            candidates = self.recorder.browser.list_interactive_texts()
        except Exception as e:  # noqa: BLE001
            self.log(f"⚠ 候補の取得に失敗しました: {e}")
            return
        if not candidates:
            self.log("(候補は見つかりませんでした)")
            return
        self.log("今の画面で押せそうな表示文字: " + " / ".join(candidates[:15]))

    # ---------- エクスプローラー ----------

    def _build_explorer(self, action: str) -> None:
        f = self.form_frame

        if action == "パスを開く":
            field = ValueSlotField(f, "開くパス")
            field.add_button("参照...", field.browse_file)
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                try:
                    self.recorder.explorer.open_path(test_v)
                    self.log(f"→ 開けました: {test_v}")
                    self.register_step({"handler": "explorer", "action": "open_path",
                                         "params": {"path": param_v}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "フォルダを作成する":
            field = ValueSlotField(f, "作成するフォルダのパス")
            field.add_button("参照...", field.browse_dir)
            field.pack(fill="x", pady=4)
            exist_ok_field = BoolField(f, "既に存在する場合はそのまま使う")
            exist_ok_field.pack(anchor="w", pady=2)

            def on_submit():
                test_v, param_v, _ = field.get()
                exist_ok = exist_ok_field.get()
                try:
                    self.recorder.explorer.create_folder(test_v, exist_ok=exist_ok)
                    self.log(f"→ 作成できました: {test_v}")
                    self.register_step({"handler": "explorer", "action": "create_folder",
                                         "params": {"path": param_v, "exist_ok": exist_ok}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action in ("ファイルを移動する", "フォルダを移動する", "ファイルをコピーする", "フォルダをコピーする"):
            is_folder = "フォルダ" in action
            is_move = "移動" in action
            src_label = f"{'移動' if is_move else 'コピー'}元の{'フォルダ' if is_folder else 'ファイル'}のパス"
            src_field = ValueSlotField(f, src_label)
            src_field.add_button("参照...", src_field.browse_dir if is_folder else src_field.browse_file)
            src_field.pack(fill="x", pady=4)
            dest_field = ValueSlotField(f, f"{'移動' if is_move else 'コピー'}先のパス")
            dest_field.add_button("参照...", dest_field.browse_dir)
            dest_field.pack(fill="x", pady=4)
            overwrite_field = BoolField(f, "同名のものがあれば上書きする")
            overwrite_field.pack(anchor="w", pady=2)

            action_name = {
                ("ファイルを移動する"): "move_file",
                ("フォルダを移動する"): "move_folder",
                ("ファイルをコピーする"): "copy_file",
                ("フォルダをコピーする"): "copy_folder",
            }[action]

            def on_submit():
                src_test, src_param, _ = src_field.get()
                dest_test, dest_param, _ = dest_field.get()
                overwrite = overwrite_field.get()
                step = {
                    "handler": "explorer", "action": action_name,
                    "params": {"source": src_param, "destination": dest_param, "overwrite": overwrite},
                }
                if self._has_template(src_test) or self._has_template(dest_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                if is_move:
                    if not self._confirm(
                        f"実際に '{src_test}' を '{dest_test}' へ移動します(元の場所から無くなります)。よろしいですか?"
                    ):
                        return
                try:
                    result = getattr(self.recorder.explorer, action_name)(src_test, dest_test, overwrite=overwrite)
                    self.log(f"→ 完了しました: {result}")
                    self.register_step(step)
                except (PathConflictError, Exception) as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action in ("ファイル名を変更する", "フォルダ名を変更する"):
            is_folder = "フォルダ" in action
            path_field = ValueSlotField(f, f"名前を変更する{'フォルダ' if is_folder else 'ファイル'}のパス")
            path_field.add_button("参照...", path_field.browse_dir if is_folder else path_field.browse_file)
            path_field.pack(fill="x", pady=4)
            name_field = ValueSlotField(f, "変更後の名前(名前のみ。例: report_{{last_row}}.xlsx も可)")
            name_field.pack(fill="x", pady=4)
            overwrite_field = BoolField(f, "同名のものがあれば上書きする")
            overwrite_field.pack(anchor="w", pady=2)

            action_name = "rename_folder" if is_folder else "rename_file"

            def on_submit():
                path_test, path_param, _ = path_field.get()
                name_test, name_param, _ = name_field.get()
                overwrite = overwrite_field.get()
                step = {
                    "handler": "explorer", "action": action_name,
                    "params": {"path": path_param, "new_name": name_param, "overwrite": overwrite},
                }
                if self._has_template(path_test) or self._has_template(name_test):
                    self.log("→ 変数参照が含まれるため動作確認をスキップして登録します")
                    self.register_step(step)
                    return
                if not self._confirm(f"実際に '{path_test}' の名前を '{name_test}' に変更します。よろしいですか?"):
                    return
                try:
                    result = getattr(self.recorder.explorer, action_name)(path_test, name_test, overwrite=overwrite)
                    self.log(f"→ 名前を変更できました: {result}")
                    self.register_step(step)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "パスが存在するか確認する":
            field = ValueSlotField(f, "確認するパス(ファイル or フォルダ)")
            field.add_button("参照...", field.browse_file)
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                try:
                    exists = self.recorder.explorer.path_exists(test_v)
                    self.log(f"→ 確認できました: {exists}")
                    store_as = self._ask_store_as()
                    step = {"handler": "explorer", "action": "path_exists", "params": {"path": param_v}}
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, exists)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "フォルダ内のファイル一覧を取得する":
            field = ValueSlotField(f, "対象フォルダのパス")
            field.add_button("参照...", field.browse_dir)
            field.pack(fill="x", pady=4)
            pattern_field = PlainField(f, "絞り込みパターン(例: *.xlsx)", default="*")
            pattern_field.pack(fill="x", pady=4)
            include_folders_field = BoolField(f, "サブフォルダも一覧に含める")
            include_folders_field.pack(anchor="w", pady=2)

            def on_submit():
                test_v, param_v, _ = field.get()
                pattern = pattern_field.get().strip() or "*"
                include_folders = include_folders_field.get()
                try:
                    files = self.recorder.explorer.list_files_in_folder(
                        test_v, pattern=pattern, include_folders=include_folders
                    )
                    self.log(f"→ 取得できました({len(files)}件): {files[:5]}")
                    store_as = self._ask_store_as()
                    step = {
                        "handler": "explorer", "action": "list_files_in_folder",
                        "params": {"path": param_v, "pattern": pattern, "include_folders": include_folders},
                    }
                    if store_as:
                        step["store_as"] = store_as
                    self.register_step(step, files)
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action in ("ファイルを削除する", "フォルダを削除する"):
            is_folder = action == "フォルダを削除する"
            path_field = ValueSlotField(f, f"削除する{'フォルダ' if is_folder else 'ファイル'}のパス")
            path_field.add_button("参照...", path_field.browse_dir if is_folder else path_field.browse_file)
            path_field.pack(fill="x", pady=4)
            recursive_field = None
            if is_folder:
                recursive_field = BoolField(f, "中身が残っていても丸ごと削除する(通常はオフ)")
                recursive_field.pack(anchor="w", pady=2)

            def on_submit():
                path_test, path_param, _ = path_field.get()
                recursive = recursive_field.get() if recursive_field else False
                if not self._confirm(
                    f"実際に '{path_test}' を削除して動作確認します"
                    f"(ゴミ箱には移動されず、元に戻せません)。よろしいですか?"
                ):
                    return
                try:
                    if is_folder:
                        result = self.recorder.explorer.delete_folder(path_test, recursive=recursive)
                        params = {"path": path_param, "recursive": recursive}
                        action_name = "delete_folder"
                    else:
                        result = self.recorder.explorer.delete_file(path_test)
                        params = {"path": path_param}
                        action_name = "delete_file"
                    self.log(f"→ 削除できました: {result}")
                    self.register_step({"handler": "explorer", "action": action_name, "params": params})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

    # ---------- 実行ファイル ----------

    def _build_process(self, action: str) -> None:
        f = self.form_frame
        if action != "exe/pyを実行する":
            return

        scripts = self.recorder.process.list_registered()
        ttk.Label(f, text="登録済みスクリプト:").pack(anchor="w")
        combo = ttk.Combobox(f, state="readonly", values=list(scripts.keys()))
        if scripts:
            combo.current(0)
        combo.pack(fill="x", pady=4)
        args_field = PlainField(f, "実行時の引数(スペース区切り、任意)", width=40)
        args_field.pack(fill="x", pady=4)

        def run_existing():
            key = combo.get()
            if not key:
                self.log("⚠ スクリプトを選択してください")
                return
            args = args_field.get().split() if args_field.get() else []
            if not self._confirm(f"'{key}' を実際に実行して動作確認します。よろしいですか?"):
                return
            try:
                output = self.recorder.process.run_registered(key, args=args, timeout=60)
                self.log(f"→ 実行結果(先頭200文字): {output[:200]}")
                retry_cfg = self._ask_retry()
                self.register_step({
                    "handler": "process", "action": "run_registered",
                    "params": {"run_key": key, "args": args}, "retry": retry_cfg,
                })
            except ScriptNotWhitelistedError as e:
                self.log(f"⚠ {e}")
            except Exception as e:  # noqa: BLE001
                self.log(f"⚠ {e}")

        ttk.Button(f, text="このスクリプトを実行して登録", command=run_existing).pack(pady=4)

        ttk.Separator(f).pack(fill="x", pady=8)
        ttk.Label(f, text="新しいexe/pyを登録:").pack(anchor="w")
        key_field = PlainField(f, "識別名(半角英数字)")
        key_field.pack(fill="x", pady=2)
        path_field = PlainField(f, "ファイルのフルパス(.py または .exe)(ここへドラッグ&ドロップも可)", width=45)
        path_field.pack(fill="x", pady=2)
        ttk.Button(f, text="参照...", command=lambda: path_field.var.set(
            filedialog.askopenfilename(filetypes=[("実行可能ファイル", "*.py *.exe"), ("すべて", "*.*")]) or path_field.var.get()
        )).pack(anchor="w")
        enable_path_drop(path_field.entry, path_field.var.set)
        kind_field = tk.StringVar(value="python")
        kind_frame = ttk.Frame(f)
        kind_frame.pack(anchor="w", pady=4)
        ttk.Radiobutton(kind_frame, text="pyファイル", variable=kind_field, value="python").pack(side="left")
        ttk.Radiobutton(kind_frame, text="exeファイル", variable=kind_field, value="exe").pack(side="left")
        new_args_field = PlainField(f, "実行時の引数(任意)", width=40)
        new_args_field.pack(fill="x", pady=4)

        def register_new():
            key = key_field.get().strip()
            path = path_field.get().strip()
            if not key or not path:
                self.log("⚠ 識別名とパスの両方を入力してください")
                return
            try:
                self.recorder.process.register_script(key, path, kind_field.get())
            except Exception as e:  # noqa: BLE001
                self.log(f"⚠ {e}")
                return
            args = new_args_field.get().split() if new_args_field.get() else []
            if not self._confirm(f"登録した '{key}' を実際に実行して動作確認します。よろしいですか?"):
                return
            try:
                output = self.recorder.process.run_registered(key, args=args, timeout=60)
                self.log(f"→ 実行結果(先頭200文字): {output[:200]}")
                retry_cfg = self._ask_retry()
                self.register_step({
                    "handler": "process", "action": "run_registered",
                    "params": {"run_key": key, "args": args}, "retry": retry_cfg,
                })
            except Exception as e:  # noqa: BLE001
                self.log(f"⚠ {e}")

        ttk.Button(f, text="登録して実行", command=register_new).pack(pady=4)

    # ---------- デスクトップ ----------

    def _build_desktop(self, action: str) -> None:
        f = self.form_frame
        ttk.Label(
            f, text="※ 画面上のどこでもクリック・入力できるため、実行前に必ず確認が入ります。",
            foreground="#a55",
        ).pack(anchor="w", pady=(0, 4))

        if action in ("画像を探してクリックする", "画像を探してマウス移動する"):
            is_click = action == "画像を探してクリックする"
            img_field = ImagePasteField(f, "対象が写っている画像")
            img_field.pack(fill="x", pady=4)
            conf_field = PlainField(f, "一致の緩さ(confidence, 0.1〜1.0)", default="0.8")
            conf_field.pack(fill="x", pady=4)
            slot_field = PlainField(f, "画像パスをスロットにする場合のスロット名(任意)")
            slot_field.pack(fill="x", pady=4)

            region_enabled_field = BoolField(f, "検索範囲を画面全体ではなく特定の領域に絞る")
            region_enabled_field.pack(anchor="w", pady=(8, 2))
            region_left_field = PlainField(f, "領域の左端X(ピクセル)")
            region_left_field.pack(fill="x", pady=2)
            region_top_field = PlainField(f, "領域の上端Y(ピクセル)")
            region_top_field.pack(fill="x", pady=2)
            region_width_field = PlainField(f, "領域の幅(ピクセル)")
            region_width_field.pack(fill="x", pady=2)
            region_height_field = PlainField(f, "領域の高さ(ピクセル)")
            region_height_field.pack(fill="x", pady=2)

            def on_submit():
                image_path = img_field.get()
                if not image_path:
                    self.log("⚠ 画像を指定してください(貼り付け または 参照)")
                    return
                try:
                    confidence = float(conf_field.get() or "0.8")
                except ValueError:
                    confidence = 0.8
                slot_name = slot_field.get().strip()
                param_path = "{{" + slot_name + "}}" if slot_name else image_path

                region = None
                if region_enabled_field.get():
                    try:
                        region = [
                            int(region_left_field.get()), int(region_top_field.get()),
                            int(region_width_field.get()), int(region_height_field.get()),
                        ]
                    except ValueError:
                        self.log("⚠ 領域は数字で入力してください")
                        return

                action_name = "locate_and_click" if is_click else "move_to_image"
                try:
                    getattr(self.recorder.desktop, action_name)(
                        image_path, confidence=confidence, timeout=10, region=region
                    )
                    self.log("→ 画像を見つけて実行できました")
                    step = {
                        "handler": "desktop", "action": action_name,
                        "params": {
                            "image_path": param_path, "confidence": confidence,
                            "timeout": 10, "region": region,
                        },
                    }
                    if is_click:
                        verify_cfg = self._ask_verify(image_or_text_default=f"{image_path}|{confidence}", is_image=True)
                        step["verify"] = verify_cfg
                        step["verify_skip"] = False
                    retry_cfg = self._ask_retry()
                    step["retry"] = retry_cfg
                    self.register_step(step)
                except (ImageNotFoundError, Exception) as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    if self._confirm("未確認のままこの手順を登録しますか?"):
                        self.register_step({
                            "handler": "desktop", "action": action_name,
                            "params": {
                                "image_path": param_path, "confidence": confidence,
                                "timeout": 10, "region": region,
                            },
                        })

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "座標をクリックする":
            x_field = PlainField(f, "X座標", default="0")
            x_field.pack(fill="x", pady=2)
            y_field = PlainField(f, "Y座標", default="0")
            y_field.pack(fill="x", pady=2)

            def on_submit():
                try:
                    x, y = int(x_field.get()), int(y_field.get())
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                if not self._confirm(f"座標({x},{y})を実際にクリックします。よろしいですか?"):
                    return
                try:
                    self.recorder.desktop.click_at(x, y)
                    self.log(f"→ クリックできました: ({x},{y})")
                    self.register_step({"handler": "desktop", "action": "click_at", "params": {"x": x, "y": y}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "文字列を入力する":
            field = ValueSlotField(f, "入力する文字列")
            field.pack(fill="x", pady=4)

            def on_submit():
                test_v, param_v, _ = field.get()
                if not self._confirm("今フォーカスされている場所に実際にキー入力します。よろしいですか?"):
                    return
                try:
                    self.recorder.desktop.type_text(test_v)
                    self.log("→ 入力できました")
                    self.register_step({"handler": "desktop", "action": "type_text",
                                         "params": {"text": param_v}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "特殊キーを送信する":
            field = PlainField(f, "送信するキー(例: enter, tab, ctrl+s)", width=20)
            field.pack(fill="x", pady=4)

            def on_submit():
                key = field.get().strip()
                if not key:
                    self.log("⚠ キーを入力してください")
                    return
                if not self._confirm(f"'{key}' を実際に送信します。よろしいですか?"):
                    return
                try:
                    self.recorder.desktop.press_key(key)
                    self.log(f"→ 送信できました: {key}")
                    self.register_step({"handler": "desktop", "action": "press_key", "params": {"key": key}})
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "スクリーンショットを撮る":
            ttk.Label(
                f, text="ここから対象のボタン/アイコン部分だけをトリミングして別ファイルに保存し、"
                        "そのパスを「画像を探して...」の操作で指定してください。\n"
                        "(このスクリーンショット取得自体はマクロの手順として登録されません)",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            path_field = ValueSlotField(f, "保存先の画像パス")
            path_field.add_button("参照...", lambda: path_field.browse_save_file(".png"))
            path_field.pack(fill="x", pady=4)

            region_enabled_field = BoolField(f, "画面全体ではなく特定の領域だけを撮影する")
            region_enabled_field.pack(anchor="w", pady=(8, 2))
            region_left_field = PlainField(f, "領域の左端X(ピクセル)")
            region_left_field.pack(fill="x", pady=2)
            region_top_field = PlainField(f, "領域の上端Y(ピクセル)")
            region_top_field.pack(fill="x", pady=2)
            region_width_field = PlainField(f, "領域の幅(ピクセル)")
            region_width_field.pack(fill="x", pady=2)
            region_height_field = PlainField(f, "領域の高さ(ピクセル)")
            region_height_field.pack(fill="x", pady=2)

            def on_submit():
                test_v, _, _ = path_field.get()
                region = None
                if region_enabled_field.get():
                    try:
                        region = [
                            int(region_left_field.get()), int(region_top_field.get()),
                            int(region_width_field.get()), int(region_height_field.get()),
                        ]
                    except ValueError:
                        self.log("⚠ 領域は数字で入力してください")
                        return
                try:
                    self.recorder.desktop.take_screenshot(test_v, region=region)
                    self.log(f"→ 保存できました: {test_v}")
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="撮影する(手順としては登録されません)", command=on_submit).pack(pady=6)

        elif action == "開いているウィンドウのタイトル一覧を見る":
            ttk.Label(
                f, text="ウィンドウをアクティブにする/サイズ・位置を指定するときの"
                        "目印探しに使います(この一覧取得自体は手順として登録されません)。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))

            def on_submit():
                try:
                    titles = self.recorder.desktop.list_window_titles()
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    return
                if not titles:
                    self.log("(今開いているウィンドウが見つかりませんでした)")
                    return
                for t in titles:
                    self.log(f"  - {t}")

            ttk.Button(f, text="一覧を表示する", command=on_submit).pack(pady=6)

        elif action == "ウィンドウをアクティブにする":
            title_field = ValueSlotField(f, "対象ウィンドウのタイトル(部分一致)")
            title_field.pack(fill="x", pady=4)

            def on_submit():
                title_test, title_param, _ = title_field.get()
                try:
                    self.recorder.desktop.activate_window_by_title(title_test)
                    self.log(f"→ アクティブにできました: {title_test}")
                    self.register_step({
                        "handler": "desktop", "action": "activate_window_by_title",
                        "params": {"title_hint": title_param},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "ウィンドウサイズを指定する(タイトル指定)":
            ttk.Label(
                f, text="Excel/PDFビューア/エクスプローラー等、どのアプリケーションの"
                        "ウィンドウでも部分一致するタイトルで指定できます。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            title_field = ValueSlotField(f, "対象ウィンドウのタイトル(部分一致)")
            title_field.pack(fill="x", pady=4)
            percent_field = BoolField(f, "画面全体に対する割合(%)で指定する(オフ=ピクセル指定)")
            percent_field.pack(anchor="w", pady=2)
            width_field = PlainField(f, "幅(空欄で変更しない)")
            width_field.pack(fill="x", pady=4)
            height_field = PlainField(f, "高さ(空欄で変更しない)")
            height_field.pack(fill="x", pady=4)

            def on_submit():
                title_test, title_param, _ = title_field.get()
                use_percent = percent_field.get()
                w_raw = width_field.get().strip()
                h_raw = height_field.get().strip()
                width = height = width_percent = height_percent = None
                try:
                    if use_percent:
                        width_percent = float(w_raw) if w_raw else None
                        height_percent = float(h_raw) if h_raw else None
                    else:
                        width = int(w_raw) if w_raw else None
                        height = int(h_raw) if h_raw else None
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                try:
                    result = self.recorder.desktop.set_window_size_by_title(
                        title_test, width=width, height=height,
                        width_percent=width_percent, height_percent=height_percent,
                    )
                    self.log(f"→ 設定できました: {result}")
                    self.register_step({
                        "handler": "desktop", "action": "set_window_size_by_title",
                        "params": {
                            "title_hint": title_param, "width": width, "height": height,
                            "width_percent": width_percent, "height_percent": height_percent,
                        },
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "ウィンドウ位置を指定する(タイトル指定)":
            ttk.Label(f, text="画面の左上を基準(0,0)として、右方向・下方向がプラスです。",
                      foreground="#557").pack(anchor="w", pady=(0, 6))
            title_field = ValueSlotField(f, "対象ウィンドウのタイトル(部分一致)")
            title_field.pack(fill="x", pady=4)
            x_field = PlainField(f, "X座標")
            x_field.pack(fill="x", pady=4)
            y_field = PlainField(f, "Y座標")
            y_field.pack(fill="x", pady=4)

            def on_submit():
                title_test, title_param, _ = title_field.get()
                try:
                    x, y = int(x_field.get()), int(y_field.get())
                except ValueError:
                    self.log("⚠ 数字で入力してください")
                    return
                try:
                    result = self.recorder.desktop.set_window_position_by_title(title_test, x, y)
                    self.log(f"→ 設定できました: {result}")
                    self.register_step({
                        "handler": "desktop", "action": "set_window_position_by_title",
                        "params": {"title_hint": title_param, "x": x, "y": y},
                    })
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

    # ---------- テキスト加工 ----------

    def _build_text(self, action: str) -> None:
        f = self.form_frame
        ttk.Label(
            f,
            text="前の手順の結果を使う場合は {{変数名}} とそのまま入力してください\n"
                 "(その場合は動作確認をスキップして登録します)",
            foreground="#557", justify="left",
        ).pack(anchor="w", pady=(0, 6))

        if action == "文字を探して切り出す":
            text_field = PlainField(f, "対象の文字列(または {{変数名}})", width=45)
            text_field.pack(fill="x", pady=4)
            marker_field = PlainField(f, "探す文字(目印)", width=30)
            marker_field.pack(fill="x", pady=4)
            include_field = BoolField(f, "その文字を含めて切り出す")
            include_field.pack(anchor="w", pady=2)
            mode_var = tk.StringVar(value="length")
            mode_frame = ttk.Frame(f)
            mode_frame.pack(anchor="w", pady=4)
            ttk.Radiobutton(mode_frame, text="文字数指定", variable=mode_var, value="length").pack(side="left")
            ttk.Radiobutton(mode_frame, text="別の文字が出るまで", variable=mode_var, value="end_marker").pack(side="left")
            ttk.Radiobutton(mode_frame, text="最後まで", variable=mode_var, value="rest").pack(side="left")
            length_field = PlainField(f, "文字数(文字数指定の場合)", default="10")
            length_field.pack(fill="x", pady=2)
            end_marker_field = PlainField(f, "終わりの目印(別の文字が出るまで、の場合)", width=30)
            end_marker_field.pack(fill="x", pady=2)

            def on_submit():
                text_val = text_field.get()
                marker = marker_field.get()
                include = include_field.get()
                mode = mode_var.get()
                length = None
                end_marker = None
                if mode == "length":
                    try:
                        length = int(length_field.get())
                    except ValueError:
                        self.log("⚠ 文字数は数字で入力してください")
                        return
                elif mode == "end_marker":
                    end_marker = end_marker_field.get()

                is_var = "{{" in text_val and "}}" in text_val
                if not is_var:
                    try:
                        result = self.recorder.text.cut_from_marker(
                            text_val, marker, include_marker=include, length=length, end_marker=end_marker
                        )
                        self.log(f"→ 動作確認できました: {result[:100]!r}")
                    except Exception as e:  # noqa: BLE001
                        self.log(f"⚠ {e}")
                        return
                else:
                    self.log("→ 変数参照のため動作確認をスキップします")

                store_as = self._ask_store_as()
                params = {"text": text_val, "marker": marker, "include_marker": include}
                if length is not None:
                    params["length"] = length
                if end_marker:
                    params["end_marker"] = end_marker
                step = {"handler": "text", "action": "cut_from_marker", "params": params}
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step, result if not is_var else _NO_VALUE)

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "文字を置換する":
            text_field = PlainField(f, "対象の文字列(または {{変数名}})", width=45)
            text_field.pack(fill="x", pady=4)
            search_field = PlainField(f, "探す文字", width=30)
            search_field.pack(fill="x", pady=4)
            replace_field = PlainField(f, "置き換える文字", width=30)
            replace_field.pack(fill="x", pady=4)

            def on_submit():
                text_val = text_field.get()
                search = search_field.get()
                replace = replace_field.get()
                is_var = "{{" in text_val and "}}" in text_val
                if not is_var:
                    result = self.recorder.text.replace_text(text_val, search, replace)
                    self.log(f"→ 動作確認できました: {result[:100]!r}")
                else:
                    self.log("→ 変数参照のため動作確認をスキップします")
                store_as = self._ask_store_as()
                step = {"handler": "text", "action": "replace_text",
                        "params": {"text": text_val, "search": search, "replace": replace}}
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step, result if not is_var else _NO_VALUE)

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "日付・時刻を取得する":
            ttk.Label(
                f, text="トークン: yyyy(西暦) YYYY(年度) MM/M(月) dd/d(日) hh mm ss(時分秒)\n"
                        "トークン以外の文字はそのまま残ります(例: yyyy年MM月dd日)",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))

            presets = {
                "yyyyMMdd_hhmmss (ファイル名の一意化等)": "yyyyMMdd_hhmmss",
                "yyyyMMdd": "yyyyMMdd",
                "hhmmss": "hhmmss",
                "yyyy (西暦年)": "yyyy",
                "YYYY (年度)": "YYYY",
                "自分で入力する": "",
            }
            ttk.Label(f, text="よく使う書式:").pack(anchor="w")
            preset_combo = ttk.Combobox(f, state="readonly", values=list(presets.keys()), width=40)
            preset_combo.current(0)
            preset_combo.pack(fill="x", pady=4)

            custom_field = PlainField(f, "書式コード(「自分で入力する」を選んだ場合)", width=30)
            custom_field.pack(fill="x", pady=4)
            fy_field = PlainField(f, "年度の開始月(YYYY使用時。空欄で既定4月)", default="4")
            fy_field.pack(fill="x", pady=4)

            def on_submit():
                preset_key = preset_combo.get()
                format_code = presets.get(preset_key) or custom_field.get().strip()
                if not format_code:
                    self.log("⚠ 書式コードを入力してください")
                    return
                try:
                    fy_month = int(fy_field.get() or "4")
                except ValueError:
                    fy_month = 4
                try:
                    result = self.recorder.text.format_now(format_code, fiscal_year_start_month=fy_month)
                    self.log(f"→ 動作確認できました: {result}")
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    return
                store_as = self._ask_store_as()
                step = {"handler": "text", "action": "format_now",
                        "params": {"format_code": format_code, "fiscal_year_start_month": fy_month}}
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step, result)

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "文字をつなげる/付加する":
            ttk.Label(f, text="つなげたい文字列を1つずつ入力してください(空欄の行は無視されます)").pack(anchor="w")
            part_vars: list[tk.StringVar] = []
            parts_frame = ttk.Frame(f)
            parts_frame.pack(fill="x", pady=4)

            def add_part_row():
                v = tk.StringVar()
                row = ttk.Frame(parts_frame)
                row.pack(fill="x", pady=1)
                ttk.Entry(row, textvariable=v, width=45).pack(side="left")
                part_vars.append(v)

            for _ in range(3):
                add_part_row()
            ttk.Button(f, text="+ 行を追加", command=add_part_row).pack(anchor="w")
            sep_field = PlainField(f, "区切り文字(空欄でそのままつなげる)", width=10)
            sep_field.pack(fill="x", pady=4)

            def on_submit():
                parts = [v.get() for v in part_vars if v.get()]
                if len(parts) < 2:
                    self.log("⚠ 2件以上入力してください")
                    return
                separator = sep_field.get()
                has_var = any("{{" in p and "}}" in p for p in parts)
                if not has_var:
                    result = self.recorder.text.combine_text(parts, separator=separator)
                    self.log(f"→ 動作確認できました: {result[:100]!r}")
                else:
                    self.log("→ 変数参照を含むため動作確認をスキップします")
                store_as = self._ask_store_as()
                step = {"handler": "text", "action": "combine_text",
                        "params": {"parts": parts, "separator": separator}}
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step, result if not has_var else _NO_VALUE)

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "クリップボードにコピーする":
            ttk.Label(
                f, text="値の欄に {{変数名}} を含めると、そこは前の手順の結果に置き換わります。",
                foreground="#557",
            ).pack(anchor="w", pady=(0, 6))
            text_field = PlainField(f, "コピーする文字列(または {{変数名}})", width=45)
            text_field.pack(fill="x", pady=4)

            def on_submit():
                text_val = text_field.get()
                has_var = "{{" in text_val and "}}" in text_val
                if not has_var:
                    try:
                        result = self.recorder.text.copy_to_clipboard(text_val)
                        self.log(f"→ {result}")
                    except Exception as e:  # noqa: BLE001
                        self.log(f"⚠ {e}")
                        if not self._confirm("未確認のままこの手順を登録しますか?"):
                            return
                else:
                    self.log("→ 変数参照のため動作確認をスキップします")
                self.register_step({"handler": "text", "action": "copy_to_clipboard",
                                     "params": {"text": text_val}})

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

        elif action == "クリップボードから取得する":
            def on_submit():
                value = None
                confirmed = False
                try:
                    value = self.recorder.text.get_from_clipboard()
                    self.log(f"→ 取得できました: {value[:100]!r}")
                    confirmed = True
                except Exception as e:  # noqa: BLE001
                    self.log(f"⚠ {e}")
                    if not self._confirm("未確認のままこの手順を登録しますか?"):
                        return
                store_as = self._ask_store_as()
                step = {"handler": "text", "action": "get_from_clipboard", "params": {}}
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step, value if confirmed else _NO_VALUE)

            ttk.Button(f, text="動作確認して登録", command=on_submit).pack(pady=6)

    # ---------- リスト(配列) ----------

    def _build_list(self, action: str) -> None:
        f = self.form_frame
        if action == "空のリストを作成する":
            ttk.Label(
                f, text="Excelのセル範囲をまとめて読み込みたいだけなら、Excel領域の\n"
                        "「セル範囲(1列/1行)をリストとして取得する」の方が簡単です。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            name_field = PlainField(f, "作成するリストの変数名(例: result)", width=20)
            name_field.pack(fill="x", pady=4)

            def on_submit():
                store_as = name_field.get().strip()
                if not store_as:
                    self.log("⚠ 変数名を入力してください")
                    return
                self.register_step({
                    "handler": "list", "action": "create_empty", "params": {}, "store_as": store_as,
                }, [])

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "リストに値を追加する":
            lst_field = PlainField(f, "追加先のリストが入っている変数名", width=20)
            lst_field.pack(fill="x", pady=4)
            value_field = PlainField(f, "追加する値(または {{変数名}})", width=30)
            value_field.pack(fill="x", pady=4)
            store_field = PlainField(f, "結果を保存する変数名(空欄で追加先に上書き)", width=20)
            store_field.pack(fill="x", pady=4)

            def on_submit():
                lst_var = lst_field.get().strip()
                if not lst_var:
                    self.log("⚠ リストの変数名を入力してください")
                    return
                value = value_field.get()
                store_as = store_field.get().strip() or lst_var
                self.register_step({
                    "handler": "list", "action": "append",
                    "params": {"lst": "{{" + lst_var + "}}", "value": value},
                    "store_as": store_as,
                })

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "リストの要素を取得する":
            lst_field = PlainField(f, "対象のリストが入っている変数名", width=20)
            lst_field.pack(fill="x", pady=4)
            idx_field = PlainField(f, "取得したい位置(0始まり。例: 0、または {{i}})", width=15)
            idx_field.pack(fill="x", pady=4)

            def on_submit():
                lst_var = lst_field.get().strip()
                if not lst_var:
                    self.log("⚠ リストの変数名を入力してください")
                    return
                index = idx_field.get()
                step = {
                    "handler": "list", "action": "get_item",
                    "params": {"lst": "{{" + lst_var + "}}", "index": index},
                }
                store_as = self._ask_store_as()
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step)
                self.log("(同じことは {{変数名[位置]}} の書き方でも直接できます)")

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "リストの長さを取得する":
            lst_field = PlainField(f, "対象のリストが入っている変数名", width=20)
            lst_field.pack(fill="x", pady=4)

            def on_submit():
                lst_var = lst_field.get().strip()
                if not lst_var:
                    self.log("⚠ リストの変数名を入力してください")
                    return
                step = {"handler": "list", "action": "length", "params": {"lst": "{{" + lst_var + "}}"}}
                store_as = self._ask_store_as()
                if store_as:
                    step["store_as"] = store_as
                self.register_step(step)
                self.log("(同じことは {{変数名.length}} の書き方でも直接できます)")

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

    # ---------- 制御構文(For/If/Goto) ----------

    def _existing_labels(self) -> list[str]:
        return [
            s["params"]["name"] for s in self.recorder.steps
            if s.get("handler") == "control" and s.get("action") == "label"
        ]

    def _build_control(self, action: str) -> None:
        f = self.form_frame
        labels = self._existing_labels()
        if labels:
            ttk.Label(f, text="これまでに置いたラベル: " + ", ".join(labels)).pack(anchor="w", pady=(0, 6))

        if action == "ラベルを置く":
            name_field = PlainField(f, "ラベル名(例: Label1)", width=20)
            name_field.pack(fill="x", pady=4)

            def on_submit():
                name = name_field.get().strip()
                if not name:
                    self.log("⚠ ラベル名を入力してください")
                    return
                if name in self._existing_labels():
                    self.log(f"⚠ 同じ名前のラベル '{name}' が既にあります")
                    return
                self.register_step({"handler": "control", "action": "label", "params": {"name": name}})

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "指定したラベルへジャンプする(goto)":
            name_field = PlainField(f, "ジャンプ先のラベル名", width=20)
            name_field.pack(fill="x", pady=4)

            def on_submit():
                name = name_field.get().strip()
                if not name:
                    self.log("⚠ ラベル名を入力してください")
                    return
                self.register_step({"handler": "control", "action": "goto", "params": {"label": name}})

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "条件を満たしたらジャンプする(IF文)":
            ttk.Label(
                f, text="条件が不成立の場合はそのまま次の手順に進むため、続く手順が\n"
                        "「ELSE」の代わりになります。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            left_field = PlainField(f, "左辺(例: {{A}})", width=20)
            left_field.pack(fill="x", pady=4)
            op_var = tk.StringVar(value="==")
            op_frame = ttk.Frame(f)
            op_frame.pack(anchor="w", pady=2)
            for op in ("==", "!=", "<", "<=", ">", ">="):
                ttk.Radiobutton(op_frame, text=op, variable=op_var, value=op).pack(side="left")
            right_field = PlainField(f, "右辺(例: {{B}})", width=20)
            right_field.pack(fill="x", pady=4)
            label_field = PlainField(f, "条件が真のときのジャンプ先ラベル名", width=20)
            label_field.pack(fill="x", pady=4)

            def on_submit():
                left, right, label = left_field.get().strip(), right_field.get().strip(), label_field.get().strip()
                if not left or not right or not label:
                    self.log("⚠ 左辺・右辺・ラベル名をすべて入力してください")
                    return
                self.register_step({
                    "handler": "control", "action": "if_goto",
                    "params": {"left": left, "op": op_var.get(), "right": right, "label": label},
                })

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "変数に値を設定する":
            var_field = PlainField(f, "設定する変数名(例: A)", width=20)
            var_field.pack(fill="x", pady=4)
            value_field = PlainField(f, "設定する値(例: {{A+1}})", width=30)
            value_field.pack(fill="x", pady=4)

            def on_submit():
                var_name = var_field.get().strip()
                value = value_field.get().strip()
                if not var_name or not value:
                    self.log("⚠ 変数名と値の両方を入力してください")
                    return
                self.register_step({
                    "handler": "control", "action": "set_value",
                    "params": {"value": value}, "store_as": var_name,
                })

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "変数の型を変換する(文字列/整数/小数)":
            ttk.Label(
                f, text="Excelのセル値等は先頭が0の値(郵便番号等)を保持するため基本的に\n"
                        "文字列として扱われます。計算に使いたい場合や、逆に確実に文字列と\n"
                        "して扱いたい場合に、変数の型を強制的に変換します。",
                foreground="#557", justify="left",
            ).pack(anchor="w", pady=(0, 6))
            var_field = PlainField(f, "変換する変数名(例: A)", width=20)
            var_field.pack(fill="x", pady=4)
            type_var = tk.StringVar(value="to_str")
            type_frame = ttk.Frame(f)
            type_frame.pack(anchor="w", pady=2)
            ttk.Radiobutton(type_frame, text="文字列(str)", variable=type_var, value="to_str").pack(side="left")
            ttk.Radiobutton(type_frame, text="整数(int)", variable=type_var, value="to_int").pack(side="left")
            ttk.Radiobutton(type_frame, text="小数(float)", variable=type_var, value="to_float").pack(side="left")

            def on_submit():
                var_name = var_field.get().strip()
                if not var_name:
                    self.log("⚠ 変数名を入力してください")
                    return
                self.register_step({
                    "handler": "control", "action": type_var.get(),
                    "params": {"value": "{{" + var_name + "}}"}, "store_as": var_name,
                })

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "繰り返しを開始する(for)":
            var_field = PlainField(f, "ループカウンタの変数名", default="i", width=15)
            var_field.pack(fill="x", pady=4)
            start_field = PlainField(f, "開始値", default="0", width=15)
            start_field.pack(fill="x", pady=4)
            end_field = PlainField(f, "終了値(例: 10、または {{リスト変数.length-1}})", width=30)
            end_field.pack(fill="x", pady=4)

            def on_submit():
                var_name = var_field.get().strip() or "i"
                start_raw = start_field.get().strip() or "0"
                end_raw = end_field.get().strip()
                if not end_raw:
                    self.log("⚠ 終了値を入力してください")
                    return

                def as_int_or_str(raw):
                    try:
                        return int(raw)
                    except ValueError:
                        return raw

                self.register_step({
                    "handler": "control", "action": "for_start",
                    "params": {"var": var_name, "start": as_int_or_str(start_raw), "end": as_int_or_str(end_raw)},
                })
                self.log("この後に繰り返したい手順を登録し、最後に「繰り返しを終了する」も忘れずに登録してください。")

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

        elif action == "繰り返しを終了する(next)":
            def on_submit():
                open_stack: list[str] = []
                for s in self.recorder.steps:
                    if s.get("handler") != "control":
                        continue
                    if s.get("action") == "for_start":
                        open_stack.append(s["params"].get("var", "i"))
                    elif s.get("action") == "for_end":
                        if open_stack:
                            open_stack.pop()
                if not open_stack:
                    self.log("⚠ 対応する「繰り返しを開始する」がまだ登録されていません")
                    return
                var_name = open_stack[-1]
                self.register_step({"handler": "control", "action": "for_end", "params": {"var": var_name}})

            ttk.Button(f, text="登録", command=on_submit).pack(pady=6)

    # ---------- 元に戻す / 保存 / 中止 ----------

    def _undo(self) -> None:
        if len(self.recorder.steps) <= self.base_step_count:
            self.log("これ以上は取り消せません。")
            return
        removed = self.recorder.steps.pop()
        if removed["handler"] == "browser" and removed["action"] == "open_registered_site":
            self.recorder._site_opened = False

        removed_refs = MacroRecorder._collect_slot_refs(removed.get("params", {}))
        for ref in removed_refs:
            slot_name = ref[2:-2].strip()
            still_used = any(
                ref in MacroRecorder._collect_slot_refs(s.get("params", {}))
                for s in self.recorder.steps
            )
            if not still_used and slot_name in self.recorder.required_slots:
                self.recorder.required_slots.remove(slot_name)

        removed_store_as = removed.get("store_as")
        if removed_store_as and removed_store_as in self.recorder.variables:
            del self.recorder.variables[removed_store_as]
            self.refresh_variables()

        self.refresh_steps()
        self.log(f"直前の操作を取り消しました: {removed['handler']}.{removed['action']}")

    def _finish(self) -> None:
        if not self.recorder.steps:
            messagebox.showinfo("保存できません", "まだ何も記録されていません。")
            return

        win = tk.Toplevel(self)
        win.title("保存")
        win.grab_set()
        ttk.Label(win, text="保存時に使う名前(半角英数字):").pack(anchor="w", padx=10, pady=(10, 0))
        name_var = tk.StringVar()
        ttk.Entry(win, textvariable=name_var, width=40).pack(padx=10)
        ttk.Label(win, text="このマクロの説明:").pack(anchor="w", padx=10, pady=(10, 0))
        desc_var = tk.StringVar()
        ttk.Entry(win, textvariable=desc_var, width=40).pack(padx=10)
        ttk.Label(win, text="呼び出しキーワード(カンマ区切り):").pack(anchor="w", padx=10, pady=(10, 0))
        kw_var = tk.StringVar()
        ttk.Entry(win, textvariable=kw_var, width=40).pack(padx=10)

        def do_save():
            macro_name = name_var.get().strip()
            description = desc_var.get().strip()
            keywords = [k.strip() for k in kw_var.get().split(",") if k.strip()]
            if not macro_name:
                messagebox.showwarning("入力不足", "保存名を入力してください")
                return

            if self.recorder._site_opened:
                self.recorder.browser.close()
                self.recorder.steps.append({"handler": "browser", "action": "close", "params": {}})

            macro_def = {
                "description": description,
                "required_slots": self.recorder.required_slots,
                "steps": self.recorder.steps,
            }
            if any(s.get("handler") == "browser" for s in self.recorder.steps):
                # このマクロを記録したときに使っていたブラウザ(chrome/edge)を保存し、
                # 実行時にこのマクロだけ自動でそのブラウザに切り替わるようにする。
                macro_def["browser"] = self.recorder.browser.browser
            self.recorder._save_macro(macro_name, macro_def)
            self.recorder._save_intent(macro_name, description, keywords)
            win.destroy()
            messagebox.showinfo("保存しました", f"マクロ '{macro_name}' を保存しました。")
            self.destroy()

        ttk.Button(win, text="保存する", command=do_save).pack(pady=12)

    def _cancel_all(self) -> None:
        if not messagebox.askyesno("中止", "記録した内容を破棄して終了しますか?"):
            return
        self._on_close()

    def _on_close(self) -> None:
        try:
            if self.recorder._site_opened:
                self.recorder.browser.close()
        except Exception:  # noqa: BLE001
            pass
        self.destroy()


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="疑似ローカルAI 操作の登録(GUI)")
    parser.add_argument(
        "--browser", choices=["chrome", "edge"], default="chrome",
        help="Web操作に使うブラウザ(既定: chrome)。EdgeはChromiumベースのためほぼ同様に動作するが、"
             "Edge本体のインストールが別途必要",
    )
    args = parser.parse_args()

    CLIP_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    app = RecorderApp(browser=args.browser)
    app.mainloop()


if __name__ == "__main__":
    main()
